"""Merge Round 3 reviewed articles + Round 0 seeds into master corpus.

Adds:
  - Round 0: 12 seed articles (all keep=1 by definition)
  - Round 3: 24 human-reviewed articles (keep read from round_03_article_review.xlsx)

Writes:
  - reports/master_corpus/master_all_articles.csv      (all rounds, full data)
  - reports/master_corpus/master_relevant_articles.csv (keep=1 only)
  - reports/master_corpus/master_irrelevant_articles.csv
  - reports/master_corpus/master_corpus_readable.csv   (keep=1, no article_text, Excel-friendly)
  - reports/master_corpus/master_corpus_summary.json
"""

from __future__ import annotations

import csv
import json
import sys
import uuid
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import urlparse

import openpyxl

sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parent.parent
RUNS = ROOT / "data" / "runs"
REPORTS = ROOT / "reports"
MASTER_DIR = REPORTS / "master_corpus"

R3_RUN = RUNS / "edible_oil_adulteration_round_03_2026-06-23" / "mediacloud" / "outputs"
R3_REVIEW_XLSX = R3_RUN / "oil_relevance" / "round_03_article_review.xlsx"
R3_ALL_CSV     = R3_RUN / "oil_relevance" / "all_articles_review.csv"
R3_META_CSV    = R3_RUN / "oil_relevance" / "metadata_all_articles_review.csv"

SEEDS_CSV = ROOT / "data" / "seeds" / "seed_urls.csv"

MASTER_COLS = [
    "round_number", "round_name", "round_description", "date_start", "date_end",
    "source_run", "final_keep", "final_human_label", "human_review_status",
    "human_review_source", "title", "source", "date", "url", "domain",
    "publication_date", "word_count", "model_final_label", "model_confidence",
    "reason", "evidence_phrase", "oil_role", "edible_oil_terms",
    "adulteration_action_terms", "negative_terms", "query_family", "query_id",
    "article_id", "file_path", "cleaned_text_path", "article_text",
]

READABLE_COLS = [
    "round_number", "final_keep", "final_human_label", "title", "source",
    "date", "url", "domain", "oil_role", "model_final_label", "model_confidence",
    "reason", "evidence_phrase", "query_family", "query_id", "human_review_status",
    "word_count",
]

R3_META = {
    "round_number": "3",
    "round_name": "round_03_mediacloud_phrase_boolean_title_proximity",
    "round_description": (
        "MediaCloud Indian collections, Jun 23 2026, phrase + boolean + title-only + proximity "
        "queries using 17 new Round-3 keywords; deduped against Rounds 1+2; human-reviewed."
    ),
    "date_start": "2021-01-01",
    "date_end": "2026-06-23",
    "source_run": "edible_oil_adulteration_round_03_2026-06-23",
    "human_review_source": "round_03_article_review.xlsx",
}

R0_META = {
    "round_number": "0",
    "round_name": "round_00_seed_articles",
    "round_description": (
        "Manually curated seed articles from oil_sample_articles.docx "
        "used to bootstrap keyword discovery."
    ),
    "date_start": "",
    "date_end": "",
    "source_run": "seed_manual_curation",
    "human_review_source": "seed_urls.csv",
}


def read_csv(p: Path) -> list[dict]:
    if not p.exists():
        return []
    with p.open(encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def write_csv(p: Path, rows: list[dict], cols: list[str]) -> None:
    p.parent.mkdir(parents=True, exist_ok=True)
    with p.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=cols, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)
    print(f"  Written: {p.name}  ({len(rows)} rows)")


def domain_of(url: str) -> str:
    try:
        return urlparse(url).netloc.lower().removeprefix("www.")
    except Exception:
        return ""


def blank_row(meta: dict, **overrides) -> dict:
    row = {c: "" for c in MASTER_COLS}
    row.update(meta)
    row.update(overrides)
    return row


# ── Load existing master ──────────────────────────────────────────────────────
existing = read_csv(MASTER_DIR / "master_all_articles.csv")
existing_urls = {r["url"] for r in existing}
print(f"Existing master rows: {len(existing)}  (unique URLs: {len(existing_urls)})")

new_rows: list[dict] = []

# ── Round 0: seeds ────────────────────────────────────────────────────────────
seeds = read_csv(SEEDS_CSV)
seed_added = 0
for s in seeds:
    url = s.get("url", "").strip()
    if not url or url in existing_urls:
        continue
    row = blank_row(
        R0_META,
        article_id=str(uuid.uuid4()),
        url=url,
        title=s.get("title_snippet", ""),
        domain=domain_of(url),
        source=s.get("domain", ""),
        date=s.get("published_date", ""),
        publication_date=s.get("published_date", ""),
        query_family="seed",
        query_id=s.get("query_used", ""),
        final_keep="1",
        final_human_label="relevant",
        human_review_status="seed_manual",
        oil_role="adulterated_product",
        model_final_label="seed",
    )
    new_rows.append(row)
    existing_urls.add(url)
    seed_added += 1
print(f"Round 0 seeds added: {seed_added}  (skipped duplicates: {len(seeds) - seed_added})")

# ── Round 3: xlsx review + full-text data ─────────────────────────────────────
wb = openpyxl.load_workbook(R3_REVIEW_XLSX, data_only=True)
ws = wb["Round 3 Review"]
headers = [c.value for c in ws[1]]
review_rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if any(v is not None for v in row):
        review_rows.append(dict(zip(headers, row)))

# Lookup tables from crawled data and metadata
r3_all_by_url  = {r["url"]: r for r in read_csv(R3_ALL_CSV)}
r3_meta_by_url = {r["url"]: r for r in read_csv(R3_META_CSV)}

r3_added = 0
for rv in review_rows:
    url  = rv.get("url", "").strip()
    keep = str(rv.get("keep", "")).strip()
    if not url or not keep:
        continue
    if url in existing_urls:
        print(f"  SKIP (already in master): {url[:80]}")
        continue

    crawled = r3_all_by_url.get(url, {})
    meta    = r3_meta_by_url.get(url, {})

    row = blank_row(
        R3_META,
        article_id=crawled.get("article_id") or meta.get("article_id") or str(uuid.uuid4()),
        url=url,
        title=rv.get("title") or crawled.get("title") or meta.get("title") or "",
        domain=domain_of(url),
        source=rv.get("source") or crawled.get("source") or meta.get("source") or "",
        date=rv.get("date") or crawled.get("date") or meta.get("date") or "",
        publication_date=crawled.get("publication_date") or rv.get("date") or meta.get("date") or "",
        word_count=crawled.get("word_count") or "",
        query_family=meta.get("query_family") or rv.get("found_by") or "",
        query_id=meta.get("query_id") or rv.get("found_by") or "",
        # Rule classifier outputs (from crawled data if available)
        model_final_label=crawled.get("final_label") or rv.get("review_tag") or "",
        model_confidence=crawled.get("confidence") or "",
        reason=crawled.get("reason") or rv.get("rule_reason") or "",
        evidence_phrase=crawled.get("evidence_phrase") or "",
        oil_role=crawled.get("oil_role") or rv.get("oil_role") or "",
        edible_oil_terms=crawled.get("edible_oil_terms") or "",
        adulteration_action_terms=crawled.get("adulteration_action_terms") or "",
        negative_terms=crawled.get("negative_terms") or "",
        # LLM outputs
        llm_label=crawled.get("llm_label") or "",
        llm_confidence=crawled.get("llm_confidence") or "",
        llm_reason=crawled.get("llm_reason") or "",
        # Human review
        final_keep=keep,
        final_human_label="relevant" if keep == "1" else "irrelevant",
        human_review_status="human_reviewed",
        file_path=crawled.get("file_path") or "",
        cleaned_text_path=crawled.get("cleaned_text_path") or "",
        article_text=crawled.get("article_text") or "",
    )
    new_rows.append(row)
    existing_urls.add(url)
    r3_added += 1

print(f"Round 3 reviewed articles added: {r3_added}")

# ── Combine and write ─────────────────────────────────────────────────────────
# Ensure MASTER_COLS includes any extra cols from crawled data (llm fields)
extra_cols = ["llm_label", "llm_confidence", "llm_reason"]
all_cols = MASTER_COLS + [c for c in extra_cols if c not in MASTER_COLS]

all_rows = existing + new_rows

# Patch existing rows to have blank extra cols if missing
for r in all_rows:
    for c in extra_cols:
        if c not in r:
            r[c] = ""

print(f"\nWriting master corpus ({len(all_rows)} total rows)...")

relevant   = [r for r in all_rows if str(r.get("final_keep")) == "1"]
irrelevant = [r for r in all_rows if str(r.get("final_keep")) == "0"]

write_csv(MASTER_DIR / "master_all_articles.csv",       all_rows,   all_cols)
write_csv(MASTER_DIR / "master_relevant_articles.csv",  relevant,   all_cols)
write_csv(MASTER_DIR / "master_irrelevant_articles.csv",irrelevant, all_cols)
write_csv(MASTER_DIR / "master_corpus_readable.csv",    relevant,   READABLE_COLS)

# ── Summary ───────────────────────────────────────────────────────────────────
by_round = Counter(str(r.get("round_number", "?")) for r in all_rows)
by_keep  = Counter(str(r.get("final_keep", "?")) for r in all_rows)
by_label = Counter(r.get("final_human_label", "?") for r in all_rows)

summary = {
    "generated_at": datetime.now(timezone.utc).isoformat(),
    "total_rows": len(all_rows),
    "total_relevant": len(relevant),
    "total_irrelevant": len(irrelevant),
    "by_round": dict(sorted(by_round.items())),
    "by_keep": dict(by_keep),
    "new_rows_added": len(new_rows),
    "seeds_added": seed_added,
    "r3_added": r3_added,
}
summary_path = MASTER_DIR / "master_corpus_summary.json"
summary_path.write_text(json.dumps(summary, indent=2, ensure_ascii=False), encoding="utf-8")
print(f"  Written: master_corpus_summary.json")

print("\n=== FINAL MASTER CORPUS ===")
print(f"  Total articles  : {len(all_rows)}")
print(f"  Relevant (keep=1): {len(relevant)}")
print(f"  Irrelevant (keep=0): {len(irrelevant)}")
print(f"  By round        : {dict(sorted(by_round.items()))}")
print(f"\n  Breakdown of relevant by round:")
for rn in sorted(by_round):
    rel = [r for r in all_rows if str(r.get("round_number")) == rn and str(r.get("final_keep")) == "1"]
    irr = [r for r in all_rows if str(r.get("round_number")) == rn and str(r.get("final_keep")) == "0"]
    print(f"    Round {rn}: {len(rel)} relevant, {len(irr)} irrelevant")
