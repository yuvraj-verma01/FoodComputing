"""Re-score the latest URL discovery with the BEST model — the 3-way weighted
ensemble (TF-IDF + e5-large + RoBERTa, weights 0.31/0.37/0.31) — and rebuild
the review Excel.

Each base model is loaded from its deployable checkpoint, run on the new
articles, and their probabilities are blended with the winning weights.
"""
from __future__ import annotations

import csv
import json
import sys
from collections import Counter
from pathlib import Path

import joblib
import numpy as np

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))
sys.stdout.reconfigure(encoding="utf-8")

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.model_training.build_text_representations import (
    title_plus_body_full, title_plus_oil_windows,
)

# ── Paths ──────────────────────────────────────────────────────────────────────
CKPT = ROOT / "reports/model_training/trained_models/final_checkpoints"
TFIDF_PATH = CKPT / "tfidf/tfidf_logreg__title_plus_body_full.joblib"
E5_CLF     = CKPT / "embeddings/emb_e5-large_lr__oil_window_embedding_clf.joblib"
E5_SCALER  = CKPT / "embeddings/emb_e5-large_lr__oil_window_embedding_scaler.joblib"
E5_CONFIG  = CKPT / "embeddings/emb_e5-large_lr__oil_window_embedding_config.json"
ROBERTA_DIR = CKPT / "transformers/hf_roberta-base_lead_512"

DISCOVERY_CSV = ROOT / "reports/full_rediscovery/full_rediscovery_crawled.csv"
OUT_XLSX      = ROOT / "reports/full_rediscovery/full_rediscovery_review_BEST_ensemble.xlsx"
OUT_CSV       = ROOT / "reports/full_rediscovery/full_rediscovery_scored_ensemble.csv"

# Winning ensemble weights (tfidf, e5, roberta), normalized
WEIGHTS = np.array([0.31, 0.37, 0.31]); WEIGHTS = WEIGHTS / WEIGHTS.sum()

# Ensemble operating points (from best_model_summary.json thresholds)
T_RELEVANT = 0.76   # high-precision point (~88% precision)
T_REVIEW   = 0.40   # surface everything from here up for human eyes


def read_csv(p):
    with p.open(encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def bucket(p):
    if p >= T_RELEVANT: return "candidate_relevant"
    if p >= T_REVIEW:   return "manual_review"
    return "candidate_irrelevant"


def main() -> int:
    rows_all = read_csv(DISCOVERY_CSV)
    rows = [r for r in rows_all if (r.get("article_text") or "").strip()]
    failed = [r for r in rows_all if not (r.get("article_text") or "").strip()]
    titles = [r.get("title", "") or "" for r in rows]
    bodies = [r.get("article_text", "") or "" for r in rows]
    print(f"Scoring {len(rows)} articles with text ({len(failed)} had no text).")

    # ── 1. TF-IDF (title + full body) ─────────────────────────────────────────
    print("  [1/3] TF-IDF ...", flush=True)
    tfidf = joblib.load(TFIDF_PATH)
    tfidf_texts = [title_plus_body_full(t, b) for t, b in zip(titles, bodies)]
    p_tfidf = tfidf.predict_proba(tfidf_texts)[:, 1]

    # ── 2. e5-large on oil-window text ────────────────────────────────────────
    print("  [2/3] e5-large embedding (oil-window) ...", flush=True)
    from sentence_transformers import SentenceTransformer
    cfg = json.loads(E5_CONFIG.read_text(encoding="utf-8"))
    prefix = cfg.get("encoder_prefix", "passage: ")
    e5 = SentenceTransformer(cfg.get("encoder", "intfloat/e5-large-v2"))
    e5.max_seq_length = 512
    oil_texts = [title_plus_oil_windows(t, b) for t, b in zip(titles, bodies)]
    emb = e5.encode([prefix + x for x in oil_texts], batch_size=32,
                    show_progress_bar=True, convert_to_numpy=True,
                    normalize_embeddings=True).astype(np.float32)
    scaler = joblib.load(E5_SCALER); e5_clf = joblib.load(E5_CLF)
    p_e5 = e5_clf.predict_proba(scaler.transform(emb))[:, 1]

    # ── 3. RoBERTa lead_512 (title + body, first 512 tokens) ──────────────────
    print("  [3/3] RoBERTa lead_512 ...", flush=True)
    import torch
    from transformers import AutoModelForSequenceClassification, AutoTokenizer
    dev = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    tok = AutoTokenizer.from_pretrained(ROBERTA_DIR)
    rob = AutoModelForSequenceClassification.from_pretrained(ROBERTA_DIR).to(dev).eval()
    p_rob = []
    with torch.no_grad():
        for i in range(0, len(tfidf_texts), 16):
            batch = tfidf_texts[i:i + 16]
            enc = tok(batch, truncation=True, padding=True, max_length=512,
                      return_tensors="pt").to(dev)
            with torch.autocast(device_type=dev.type, dtype=torch.float16, enabled=dev.type == "cuda"):
                logits = rob(**enc).logits.float()
            p_rob.extend(torch.softmax(logits, dim=1)[:, 1].cpu().numpy().tolist())
    p_rob = np.array(p_rob)

    # ── Ensemble ──────────────────────────────────────────────────────────────
    p_ens = WEIGHTS[0] * p_tfidf + WEIGHTS[1] * p_e5 + WEIGHTS[2] * p_rob

    records = []
    for r, pt, pe, pr, pen in zip(rows, p_tfidf, p_e5, p_rob, p_ens):
        records.append({
            "ensemble_prob": round(float(pen), 4),
            "bucket": bucket(float(pen)),
            "p_tfidf": round(float(pt), 3),
            "p_e5": round(float(pe), 3),
            "p_roberta": round(float(pr), 3),
            "title": r.get("title", ""), "source": r.get("source", ""),
            "date": r.get("date", ""), "url": r.get("url", ""),
            "domain": r.get("domain", ""), "query_family": r.get("query_family", ""),
            "word_count": r.get("word_count", ""),
        })
    records.sort(key=lambda x: x["ensemble_prob"], reverse=True)

    # ── Write CSV ─────────────────────────────────────────────────────────────
    fields = ["ensemble_prob", "bucket", "p_tfidf", "p_e5", "p_roberta",
              "title", "source", "date", "url", "domain", "query_family", "word_count"]
    with OUT_CSV.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader(); w.writerows(records)

    # ── Write Excel ───────────────────────────────────────────────────────────
    GREEN = PatternFill("solid", fgColor="C6EFCE")
    YELLOW = PatternFill("solid", fgColor="FFEB9C")
    RED = PatternFill("solid", fgColor="FFC7CE")
    GREY = PatternFill("solid", fgColor="D9D9D9")
    HDR = PatternFill("solid", fgColor="2F5496"); HF = Font(bold=True, color="FFFFFF")

    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Ensemble Scored"
    cols = ["keep", "ensemble_prob", "bucket", "p_tfidf", "p_e5", "p_roberta",
            "title", "source", "date", "url", "domain", "query_family", "word_count"]
    ws.append(cols)
    for c in ws[1]:
        c.font = HF; c.fill = HDR
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for rec in records:
        fill = (GREEN if rec["bucket"] == "candidate_relevant"
                else YELLOW if rec["bucket"] == "manual_review" else RED)
        ws.append(["", rec["ensemble_prob"], rec["bucket"], rec["p_tfidf"],
                   rec["p_e5"], rec["p_roberta"], rec["title"], rec["source"],
                   rec["date"], rec["url"], rec["domain"], rec["query_family"],
                   rec["word_count"]])
        for c in ws[ws.max_row]:
            c.fill = fill; c.alignment = Alignment(vertical="top")
    for i, wdt in enumerate([6, 13, 18, 8, 8, 9, 60, 20, 12, 58, 22, 16, 10], 1):
        ws.column_dimensions[get_column_letter(i)].width = wdt
    ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions

    ws2 = wb.create_sheet("Crawl Failed")
    ws2.append(["url", "source", "title", "query_family"])
    for c in ws2[1]:
        c.font = HF; c.fill = HDR
    for r in failed:
        ws2.append([r.get("url", ""), r.get("source", ""), r.get("title", ""), r.get("query_family", "")])
        for c in ws2[ws2.max_row]:
            c.fill = GREY

    ws3 = wb.create_sheet("Stats")
    bc = Counter(r["bucket"] for r in records)
    stats = [
        ("MODEL", "3-way weighted ensemble"),
        ("members", "TF-IDF logreg + e5-large LR (oil-window) + RoBERTa lead_512"),
        ("weights (tfidf/e5/roberta)", "0.31 / 0.37 / 0.31"),
        ("CV F1 of this ensemble", 0.795),
        ("thresholds", f"candidate_relevant>={T_RELEVANT}, manual_review>={T_REVIEW}"),
        ("", ""),
        ("articles scored", len(records)),
        ("crawl failed (no text)", len(failed)),
        ("candidate_relevant (>=0.76)", bc.get("candidate_relevant", 0)),
        ("manual_review (0.40-0.76)", bc.get("manual_review", 0)),
        ("candidate_irrelevant (<0.40)", bc.get("candidate_irrelevant", 0)),
    ]
    ws3.column_dimensions["A"].width = 34; ws3.column_dimensions["B"].width = 60
    for row in stats:
        ws3.append(row)

    wb.save(OUT_XLSX)

    print(f"\n{'='*60}")
    print(f"Scored {len(records)} articles with the BEST ensemble.")
    print(f"  candidate_relevant (>={T_RELEVANT}): {bc.get('candidate_relevant',0)}")
    print(f"  manual_review ({T_REVIEW}-{T_RELEVANT}):     {bc.get('manual_review',0)}")
    print(f"  candidate_irrelevant (<{T_REVIEW}):  {bc.get('candidate_irrelevant',0)}")
    print(f"Excel: {OUT_XLSX}")
    print(f"CSV:   {OUT_CSV}")
    print(f"{'='*60}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())