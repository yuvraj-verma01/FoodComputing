from __future__ import annotations

import argparse
import csv
import json
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parent.parent
DEFAULT_REPORTS_DIR = ROOT / "reports"
DEFAULT_MASTER_DIR = ROOT / "reports" / "master_corpus"


MASTER_COLUMNS = [
    "round_number",
    "round_name",
    "round_description",
    "date_start",
    "date_end",
    "source_run",
    "final_keep",
    "final_human_label",
    "human_review_status",
    "human_review_source",
    "title",
    "source",
    "date",
    "url",
    "domain",
    "publication_date",
    "word_count",
    "model_final_label",
    "model_confidence",
    "reason",
    "evidence_phrase",
    "oil_role",
    "edible_oil_terms",
    "adulteration_action_terms",
    "negative_terms",
    "query_family",
    "query_id",
    "article_id",
    "file_path",
    "cleaned_text_path",
    "article_text",
]

CONFLICT_COLUMNS = [
    "url",
    "first_round",
    "duplicate_round",
    "first_label",
    "duplicate_label",
    "title",
    "source_file",
    "final_decision",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build the cumulative master reviewed corpus from round report folders."
    )
    parser.add_argument("--reports-dir", type=Path, default=DEFAULT_REPORTS_DIR)
    parser.add_argument("--master-dir", type=Path, default=DEFAULT_MASTER_DIR)
    return parser.parse_args()


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def write_csv(path: Path, rows: list[dict[str, Any]], columns: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def discover_round_files(reports_dir: Path) -> list[Path]:
    return sorted(reports_dir.glob("edible_oil_adulteration_round_*/round_*_all_articles.csv"))


def read_conflict_resolutions(path: Path) -> dict[str, str]:
    if not path.exists():
        return {}
    resolutions: dict[str, str] = {}
    for row in read_csv(path):
        url = row.get("url", "").strip()
        decision = (
            row.get("final_decision", "")
            or row.get("final_decisio", "")
            or row.get("final", "")
        ).strip()
        if not url or decision == "":
            continue
        if decision not in {"0", "1"}:
            raise ValueError(f"Invalid final_decision for {url}: {decision!r}. Use 0 or 1.")
        if url in resolutions and resolutions[url] != decision:
            raise ValueError(f"Conflicting final_decision values for duplicate URL: {url}")
        resolutions[url] = decision
    return resolutions


def apply_conflict_resolutions(
    rows: list[dict[str, Any]],
    all_round_rows: list[dict[str, Any]],
    resolutions: dict[str, str],
) -> int:
    if not resolutions:
        return 0
    applied_urls: set[str] = set()
    for row in rows + all_round_rows:
        url = row.get("url", "")
        if url not in resolutions:
            continue
        decision = resolutions[url]
        row["final_keep"] = decision
        row["final_human_label"] = "relevant" if decision == "1" else "irrelevant"
        row["human_review_status"] = "conflict_resolved"
        row["human_review_source"] = "master_duplicate_label_conflicts.csv"
        applied_urls.add(url)
    return len(applied_urls)


def normalize_rows(round_files: list[Path]) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    rows: list[dict[str, Any]] = []
    all_round_rows: list[dict[str, Any]] = []
    duplicate_urls: list[dict[str, Any]] = []
    seen: dict[str, dict[str, Any]] = {}
    for path in round_files:
        for row in read_csv(path):
            normalized = {col: row.get(col, "") for col in MASTER_COLUMNS}
            normalized["_source_file"] = str(path)
            all_round_rows.append(normalized)
            url = normalized.get("url", "")
            if url and url in seen:
                duplicate_urls.append(
                    {
                        "url": url,
                        "first_round": seen[url].get("round_number", ""),
                        "duplicate_round": normalized.get("round_number", ""),
                        "first_label": seen[url].get("final_human_label", ""),
                        "duplicate_label": normalized.get("final_human_label", ""),
                        "title": normalized.get("title", ""),
                        "source_file": str(path),
                    }
                )
                continue
            if url:
                seen[url] = normalized
            rows.append(normalized)
    return rows, duplicate_urls, all_round_rows


def write_workbook(
    path: Path,
    rows: list[dict[str, Any]],
    duplicates: list[dict[str, Any]],
    duplicate_label_conflicts: list[dict[str, Any]],
    all_round_rows: list[dict[str, Any]],
    summary: dict[str, Any],
) -> None:
    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "summary"
    summary_ws.append(["metric", "value"])
    for key, value in summary.items():
        summary_ws.append([key, json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list)) else value])
    summary_ws.column_dimensions["A"].width = 34
    summary_ws.column_dimensions["B"].width = 100
    for cell in summary_ws[1]:
        cell.font = Font(bold=True)

    add_sheet(wb, "all_articles", rows, [col for col in MASTER_COLUMNS if col != "article_text"])
    add_sheet(wb, "all_round_articles", all_round_rows, [col for col in MASTER_COLUMNS if col != "article_text"])
    add_sheet(
        wb,
        "relevant",
        [row for row in rows if row.get("final_human_label") == "relevant"],
        [col for col in MASTER_COLUMNS if col != "article_text"],
    )
    add_sheet(
        wb,
        "irrelevant",
        [row for row in rows if row.get("final_human_label") == "irrelevant"],
        [col for col in MASTER_COLUMNS if col != "article_text"],
    )
    add_sheet(
        wb,
        "duplicate_urls_skipped",
        duplicates,
        CONFLICT_COLUMNS,
    )
    add_sheet(
        wb,
        "duplicate_label_conflicts",
        duplicate_label_conflicts,
        CONFLICT_COLUMNS,
    )
    wb.save(path)


def add_sheet(wb: Workbook, title: str, rows: list[dict[str, Any]], columns: list[str]) -> None:
    ws = wb.create_sheet(title)
    ws.append(columns)
    for row in rows:
        ws.append([row.get(col, "") for col in columns])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    keep_fill = PatternFill("solid", fgColor="E2F0D9")
    drop_fill = PatternFill("solid", fgColor="FCE4D6")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = {
        "A": 12,
        "B": 36,
        "C": 70,
        "G": 10,
        "H": 18,
        "K": 58,
        "L": 24,
        "N": 54,
        "R": 18,
        "T": 45,
        "U": 45,
        "AB": 14,
        "AD": 54,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    label_col = columns.index("final_human_label") + 1 if "final_human_label" in columns else None
    url_col = columns.index("url") + 1 if "url" in columns else None
    for row_idx in range(2, ws.max_row + 1):
        fill = None
        if label_col:
            label = ws.cell(row=row_idx, column=label_col).value
            fill = keep_fill if label == "relevant" else drop_fill if label == "irrelevant" else None
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        if url_col:
            cell = ws.cell(row=row_idx, column=url_col)
            if cell.value:
                cell.hyperlink = str(cell.value)
                cell.style = "Hyperlink"
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def main() -> None:
    args = parse_args()
    round_files = discover_round_files(args.reports_dir)
    conflict_resolutions = read_conflict_resolutions(args.master_dir / "master_duplicate_label_conflicts.csv")
    rows, duplicates, all_round_rows = normalize_rows(round_files)
    duplicate_label_conflicts = [
        row for row in duplicates
        if row.get("first_label") != row.get("duplicate_label")
    ]
    for row in duplicate_label_conflicts:
        row["final_decision"] = conflict_resolutions.get(row.get("url", ""), "")
    for row in duplicates:
        row["final_decision"] = conflict_resolutions.get(row.get("url", ""), "")
    conflict_resolutions_applied = apply_conflict_resolutions(rows, all_round_rows, conflict_resolutions)
    label_counts = Counter(row.get("final_human_label") or "" for row in rows)
    round_counts = Counter(row.get("round_number") or "" for row in rows)
    all_round_counts = Counter(row.get("round_number") or "" for row in all_round_rows)
    summary = {
        "created_at": utc_now(),
        "round_files_loaded": [str(path) for path in round_files],
        "total_reviewed_rows_including_duplicate_urls": len(all_round_rows),
        "total_unique_articles": len(rows),
        "duplicate_urls_skipped": len(duplicates),
        "duplicate_label_conflicts": len(duplicate_label_conflicts),
        "conflict_resolutions_loaded": len(conflict_resolutions),
        "conflict_resolutions_applied": conflict_resolutions_applied,
        "label_counts": dict(sorted(label_counts.items())),
        "round_counts": dict(sorted(round_counts.items())),
        "all_round_row_counts": dict(sorted(all_round_counts.items())),
        "note": "Master corpus is URL-deduped; duplicate URLs from later rounds are listed separately.",
    }

    args.master_dir.mkdir(parents=True, exist_ok=True)
    write_csv(args.master_dir / "master_all_articles.csv", rows, MASTER_COLUMNS)
    write_csv(args.master_dir / "master_all_round_articles.csv", all_round_rows, MASTER_COLUMNS)
    write_csv(
        args.master_dir / "master_relevant_articles.csv",
        [row for row in rows if row.get("final_human_label") == "relevant"],
        MASTER_COLUMNS,
    )
    write_csv(
        args.master_dir / "master_irrelevant_articles.csv",
        [row for row in rows if row.get("final_human_label") == "irrelevant"],
        MASTER_COLUMNS,
    )
    write_csv(
        args.master_dir / "master_duplicate_urls_skipped.csv",
        duplicates,
        CONFLICT_COLUMNS,
    )
    write_csv(
        args.master_dir / "master_duplicate_label_conflicts.csv",
        duplicate_label_conflicts,
        CONFLICT_COLUMNS,
    )
    (args.master_dir / "master_summary.json").write_text(
        json.dumps(summary, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    write_workbook(
        args.master_dir / "master_corpus.xlsx",
        rows,
        duplicates,
        duplicate_label_conflicts,
        all_round_rows,
        summary,
    )

    print(json.dumps(summary, indent=2, ensure_ascii=False))
    print(f"Master corpus written: {args.master_dir.resolve()}")


if __name__ == "__main__":
    main()
