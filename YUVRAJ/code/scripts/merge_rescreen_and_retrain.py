"""
Pipeline:
  0) Merge rescreen-marked articles into master corpus (round 4)
  1) Retrain classifiers (backup previous results first)
  2) Extract keywords from new round-4 relevant articles → Excel for review
"""
from __future__ import annotations
import csv, json, re, shutil, string, sys
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))
sys.stdout.reconfigure(encoding="utf-8")

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sklearn.feature_extraction.text import TfidfVectorizer
import numpy as np

MASTER_DIR   = ROOT / "reports/master_corpus"
MASTER_CSV   = MASTER_DIR / "master_all_articles.csv"
RESCREEN_CSV = ROOT / "reports/rescreen/rescreen_all_dropped.csv"
RESCREEN_XLS = ROOT / "reports/rescreen/rescreen_review.xlsx"
MODEL_OUT    = ROOT / "reports/model_training"
KW_OUT       = ROOT / "reports/rescreen/round4_keyword_candidates.xlsx"

ROUND4_META = {
    "round_number":    "4",
    "round_name":      "rescreen_dropped",
    "round_description": "Rescreened metadata-dropped URLs from rounds 1-3 using best classifier",
    "date_start":      "2026-06-24",
    "date_end":        "2026-06-24",
    "source_run":      "rescreen_all_dropped",
    "human_review_status": "human_reviewed",
    "human_review_source": "rescreen_review.xlsx",
}

# ── Domain anchors for keyword extraction ─────────────────────────────────────
_OIL_TERMS = {
    "oil", "edible", "mustard", "palm", "groundnut", "soybean", "sunflower",
    "coconut", "rice", "bran", "cooking", "vanaspati", "ghee", "adulterated",
    "adulteration", "spurious", "substandard", "contaminated", "fake", "seized",
    "raid", "fssai", "fda", "fsda", "food safety", "quality",
}
_STOP = {
    "india", "indian", "news", "article", "report", "said", "also", "new",
    "year", "time", "state", "government", "official", "officials", "police",
    "court", "district", "city", "people", "home", "public", "per", "cent",
    "lakh", "crore", "rs", "kg", "litre", "litres", "ton", "department",
    "minister", "national", "local", "high", "low", "large", "small",
    "days", "months", "years", "week", "monday", "tuesday", "wednesday",
    "thursday", "friday", "saturday", "sunday", "january", "february",
    "march", "april", "may", "june", "july", "august", "september",
    "october", "november", "december", "the", "and", "for", "are", "was",
    "were", "has", "have", "had", "been", "being", "that", "this", "with",
    "from", "they", "will", "would", "could", "should", "which", "their",
    "there", "about", "after", "before", "while", "when", "where", "who",
    "what", "how", "its", "not", "but", "more", "into", "other", "than",
    "them", "then", "some", "such", "any", "all", "two", "three", "four",
    "five", "six", "seven", "eight", "nine", "ten",
}


def read_csv(p: Path) -> list[dict]:
    with p.open(encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def write_csv(p: Path, rows: list[dict], fieldnames: list[str]) -> None:
    with p.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)
    print(f"  Written {p.name}: {len(rows)} rows")


# ═══════════════════════════════════════════════════════════════════════════════
# TASK 0 — Merge rescreen marks into master corpus
# ═══════════════════════════════════════════════════════════════════════════════
def task0_merge() -> list[dict]:
    print("\n" + "="*60)
    print("TASK 0 — Merge rescreen marks into master corpus")
    print("="*60)

    # Load marks from Excel
    wb  = openpyxl.load_workbook(RESCREEN_XLS, data_only=True)
    ws  = wb["All Scored"]
    hdrs = [c.value for c in ws[1]]
    marked = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(hdrs, row))
        k = d.get("keep")
        if k in (0, 1, "0", "1", 0.0, 1.0):
            d["keep"] = int(float(str(k)))
            marked.append(d)

    keep1 = [d for d in marked if d["keep"] == 1]
    keep0 = [d for d in marked if d["keep"] == 0]
    print(f"Marks loaded — keep=1: {len(keep1)}, keep=0: {len(keep0)}")

    # Load full text from rescreen CSV (join by URL)
    rescreen_rows = read_csv(RESCREEN_CSV)
    text_by_url = {r["url"]: r for r in rescreen_rows}

    # Load existing master corpus
    master_rows = read_csv(MASTER_CSV)
    fieldnames  = list(master_rows[0].keys())
    master_urls = {r["url"] for r in master_rows}

    # Build new rows
    now = datetime.now(timezone.utc).isoformat()
    new_rows: list[dict] = []
    skipped_dup = 0

    for d in marked:
        url = d.get("url") or ""
        if not url or url in master_urls:
            skipped_dup += 1
            continue

        full = text_by_url.get(url, {})
        text = full.get("article_text") or ""
        keep = d["keep"]

        from urllib.parse import urlparse
        domain = urlparse(url).netloc.replace("www.", "")

        row = {col: "" for col in fieldnames}
        row.update(ROUND4_META)
        row["final_keep"]        = str(keep)
        row["final_human_label"] = "relevant" if keep == 1 else "irrelevant"
        row["title"]             = str(d.get("title") or "")
        row["source"]            = str(d.get("source") or full.get("source") or "")
        row["date"]              = str(d.get("date") or full.get("date") or "")
        row["url"]               = url
        row["domain"]            = domain
        row["word_count"]        = str(full.get("word_count") or "")
        row["query_family"]      = str(d.get("query_family") or full.get("query_family") or "")
        row["article_text"]      = text
        row["model_final_label"] = d.get("bucket") or ""
        row["model_confidence"]  = str(d.get("prob") or "")
        new_rows.append(row)
        master_urls.add(url)

    print(f"New rows to add: {len(new_rows)}  (skipped {skipped_dup} already in corpus)")
    all_rows = master_rows + new_rows

    # Write all master files
    relevant   = [r for r in all_rows if str(r.get("final_keep")) == "1"]
    irrelevant = [r for r in all_rows if str(r.get("final_keep")) == "0"]

    READABLE_COLS = [
        "round_number", "final_keep", "final_human_label", "title", "source",
        "date", "url", "domain", "oil_role", "model_final_label", "model_confidence",
        "reason", "evidence_phrase", "query_family", "query_id", "human_review_status",
        "word_count",
    ]

    write_csv(MASTER_CSV, all_rows, fieldnames)
    write_csv(MASTER_DIR / "master_relevant_articles.csv",   relevant,   fieldnames)
    write_csv(MASTER_DIR / "master_irrelevant_articles.csv", irrelevant, fieldnames)
    write_csv(MASTER_DIR / "master_corpus_readable.csv",     relevant,   READABLE_COLS)

    print(f"Master corpus: {len(all_rows)} total | {len(relevant)} relevant | {len(irrelevant)} irrelevant")
    return new_rows


# ═══════════════════════════════════════════════════════════════════════════════
# TASK 1 — Retrain classifiers (backup previous results first)
# ═══════════════════════════════════════════════════════════════════════════════
def task1_retrain() -> None:
    print("\n" + "="*60)
    print("TASK 1 — Backup previous results + retrain")
    print("="*60)

    # Backup existing model_training output
    backup_dir = MODEL_OUT.parent / f"model_training_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    if MODEL_OUT.exists():
        shutil.copytree(MODEL_OUT, backup_dir)
        print(f"Backed up previous results to: {backup_dir.name}")

    import subprocess
    cmd = [
        sys.executable, "-m", "src.model_training.run_all_experiments",
        "--input",        str(MASTER_CSV),
        "--label-column", "final_keep",
        "--output-dir",   str(MODEL_OUT),
        "--skip-sbert",
    ]
    print(f"Running: {' '.join(cmd[2:])}")
    result = subprocess.run(cmd, cwd=ROOT, capture_output=False, text=True)
    if result.returncode != 0:
        print("WARNING: Training exited with non-zero code")


# ═══════════════════════════════════════════════════════════════════════════════
# TASK 2 — Extract keywords from new round-4 relevant articles
# ═══════════════════════════════════════════════════════════════════════════════
def task2_keywords(new_rows: list[dict]) -> None:
    print("\n" + "="*60)
    print("TASK 2 — Extract keyword candidates from new relevant articles")
    print("="*60)

    # Load ALL existing queries used in rounds 1-3 (to avoid repeating them)
    used_terms: set[str] = set()
    for qfile in ["config/queries.yaml", "config/queries_ddgs.yaml",
                  "config/mediacloud_seed_keywords.yaml"]:
        p = ROOT / qfile
        if p.exists():
            text = p.read_text(encoding="utf-8").lower()
            # Extract quoted phrases and bare terms
            for m in re.findall(r'"([^"]+)"', text):
                used_terms.add(m.strip())
            for m in re.findall(r"'([^']+)'", text):
                used_terms.add(m.strip())

    # New round-4 relevant articles with text
    rel_new = [r for r in new_rows if str(r.get("final_keep")) == "1"
               and (r.get("article_text") or "").strip()]
    print(f"New relevant articles with text: {len(rel_new)}")

    if not rel_new:
        print("No new relevant articles with text — skipping keyword extraction.")
        return

    # Also load ALL previous relevant articles for TF-IDF background corpus
    master_rows = read_csv(MASTER_CSV)
    r4_urls = {r["url"] for r in new_rows}
    prev_rel = [r for r in master_rows
                if str(r.get("final_keep")) == "1"
                and r["url"] not in r4_urls
                and (r.get("article_text") or "").strip()]

    all_rel_texts  = [r["article_text"] for r in prev_rel + rel_new]
    all_rel_labels = [0] * len(prev_rel) + [1] * len(rel_new)  # 1 = new

    # TF-IDF on unigrams + bigrams
    vec = TfidfVectorizer(
        ngram_range=(1, 2),
        max_features=5000,
        min_df=2,
        stop_words="english",
        sublinear_tf=True,
    )
    X = vec.fit_transform(all_rel_texts)
    features = np.array(vec.get_feature_names_out())

    # Score = mean TF-IDF in new articles
    new_idx  = [i for i, l in enumerate(all_rel_labels) if l == 1]
    X_new    = X[new_idx]
    scores   = np.asarray(X_new.mean(axis=0)).ravel()

    # Filter: must contain oil/adulteration anchor, not in stopwords
    def is_valid(term: str) -> bool:
        words = term.lower().split()
        if any(w in _STOP for w in words):
            return False
        if len(term) < 4:
            return False
        if any(ch.isdigit() for ch in term):
            return False
        # Must overlap with domain terms or be novel enforcement language
        has_anchor = any(oil in term for oil in _OIL_TERMS)
        return True  # keep all, flag anchor ones separately

    candidates = []
    for feat, score in zip(features, scores):
        if score < 0.01:
            continue
        term = feat.strip().lower()
        if not is_valid(term):
            continue
        # Count how many new articles contain this term
        df_new  = sum(1 for i in new_idx if X[i, list(features).index(feat)] > 0)
        anchored = any(oil in term for oil in _OIL_TERMS)
        already  = term in used_terms or any(term in u for u in used_terms)
        candidates.append({
            "term":          feat,
            "tfidf_score":   round(float(score), 4),
            "doc_freq_new":  df_new,
            "anchored":      "yes" if anchored else "",
            "already_used":  "yes" if already else "",
            "add_to_queries": "",   # user fills in 1/0
            "notes": "",
        })

    candidates.sort(key=lambda x: (-x["doc_freq_new"], -x["tfidf_score"]))
    print(f"Keyword candidates: {len(candidates)}")

    # Write Excel
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Keyword Candidates"

    hdr_fill = PatternFill("solid", fgColor="2E75B6")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    anch_fill = PatternFill("solid", fgColor="E2EFDA")
    used_fill = PatternFill("solid", fgColor="FFC7CE")
    thin = Border(bottom=Side(style="thin", color="DDDDDD"))

    COLS = ["add_to_queries", "term", "tfidf_score", "doc_freq_new",
            "anchored", "already_used", "notes"]
    ws.append(COLS)
    for ci, col in enumerate(COLS, 1):
        c = ws.cell(1, ci); c.font = hdr_font; c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center")

    for cand in candidates:
        ws.append([cand[c] for c in COLS])
        ri = ws.max_row
        fill = used_fill if cand["already_used"] else (anch_fill if cand["anchored"] else PatternFill())
        for ci in range(1, len(COLS)+1):
            c = ws.cell(ri, ci); c.fill = fill; c.border = thin
            c.alignment = Alignment(vertical="top")

    widths = dict(add_to_queries=14, term=40, tfidf_score=11, doc_freq_new=12,
                  anchored=10, already_used=12, notes=40)
    for ci, col in enumerate(COLS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = widths.get(col, 14)
    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Sheet 2: source articles for reference
    ws2 = wb.create_sheet("New Relevant Articles")
    COLS2 = ["title", "source", "date", "url", "word_count", "model_confidence"]
    ws2.append(COLS2)
    for ci, col in enumerate(COLS2, 1):
        c = ws2.cell(1, ci); c.font = hdr_font; c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center")
    for r in sorted(rel_new, key=lambda x: float(x.get("model_confidence") or 0), reverse=True):
        ws2.append([r.get(c, "") for c in COLS2])
    ws2.column_dimensions["A"].width = 65
    ws2.column_dimensions["D"].width = 55
    ws2.freeze_panes = "A2"

    wb.save(KW_OUT)
    print(f"Keyword candidates saved: {KW_OUT}")
    print(f"  Green  = oil/adulteration anchored terms")
    print(f"  Red    = already used in previous rounds")


# ═══════════════════════════════════════════════════════════════════════════════
# Main
# ═══════════════════════════════════════════════════════════════════════════════
def main() -> int:
    new_rows = task0_merge()
    task1_retrain()
    task2_keywords(new_rows)
    print("\nAll tasks complete.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
