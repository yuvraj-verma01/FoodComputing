"""Write full rescreen results to a formatted Excel file for human review."""
import csv, sys
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")
ROOT = Path(__file__).resolve().parent.parent

rows = list(csv.DictReader(
    (ROOT / "reports/rescreen/rescreen_all_dropped.csv").open(encoding="utf-8-sig")
))

def safe_prob(r):
    try:
        return float(r["prob"])
    except (ValueError, TypeError):
        return -1.0

scored = [r for r in rows if r.get("prob") not in ("", "None", None)]
scored.sort(key=safe_prob, reverse=True)
failed = [r for r in rows if r.get("bucket") in ("crawl_failed", "no_text")]

wb = openpyxl.Workbook()

hdr_fill = PatternFill("solid", fgColor="2E75B6")
hdr_font = Font(bold=True, color="FFFFFF", size=11)
fills = {
    "candidate_relevant":   PatternFill("solid", fgColor="C6EFCE"),
    "manual_review":        PatternFill("solid", fgColor="FFEB9C"),
    "candidate_irrelevant": PatternFill("solid", fgColor="FFC7CE"),
}
thin = Border(bottom=Side(style="thin", color="DDDDDD"))

# ── Sheet 1: All Scored ───────────────────────────────────────────────────────
ws = wb.active
ws.title = "All Scored"

COLS = ["keep", "prob", "bucket", "round", "title", "source",
        "date", "url", "query_family", "word_count"]
ws.append(COLS)
for ci, col in enumerate(COLS, 1):
    c = ws.cell(1, ci)
    c.font = hdr_font
    c.fill = hdr_fill
    c.alignment = Alignment(horizontal="center", vertical="center")

for r in scored:
    try:
        prob = round(float(r["prob"]), 3)
    except (ValueError, TypeError):
        prob = ""
    bkt = r.get("bucket", "")
    ws.append([
        "",                         # keep — user fills 0/1
        prob,
        bkt,
        r.get("round", ""),
        r.get("title", ""),
        r.get("source", ""),
        r.get("date", ""),
        r.get("url", ""),
        r.get("query_family", ""),
        r.get("word_count", ""),
    ])
    ri = ws.max_row
    f  = fills.get(bkt, PatternFill())
    for ci in range(1, len(COLS) + 1):
        c = ws.cell(ri, ci)
        c.fill      = f
        c.border    = thin
        c.alignment = Alignment(vertical="top")

widths = dict(keep=6, prob=7, bucket=22, round=6, title=60, source=22,
              date=12, url=55, query_family=14, word_count=10)
for ci, col in enumerate(COLS, 1):
    ws.column_dimensions[get_column_letter(ci)].width = widths.get(col, 14)
ws.row_dimensions[1].height = 26
ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions

# ── Sheet 2: Crawl Failed ────────────────────────────────────────────────────
ws2 = wb.create_sheet("Crawl Failed")
COLS2 = ["round", "title", "source", "date", "url", "query_family"]
ws2.append(COLS2)
for ci, col in enumerate(COLS2, 1):
    c = ws2.cell(1, ci)
    c.font = hdr_font
    c.fill = hdr_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
for r in failed:
    ws2.append([r.get(c, "") for c in COLS2])
ws2.column_dimensions["E"].width = 60
ws2.column_dimensions["B"].width = 55
ws2.freeze_panes = "A2"

# ── Sheet 3: Stats ───────────────────────────────────────────────────────────
ws3 = wb.create_sheet("Stats")
stats = [
    ("Total URLs in pool",                         3936),
    ("Crawled successfully",                       3399),
    ("Got article text + scored",                  len(scored)),
    ("Crawl failed / no text",                     len(failed)),
    ("", ""),
    ("candidate_relevant  (prob >= 0.65)",  sum(1 for r in scored if safe_prob(r) >= 0.65)),
    ("manual_review       (0.35 – 0.65)",   sum(1 for r in scored if 0.35 <= safe_prob(r) < 0.65)),
    ("candidate_irrelevant (prob < 0.35)",  sum(1 for r in scored if safe_prob(r) < 0.35)),
]
for row in stats:
    ws3.append(list(row))
for cell in ws3["A"]:
    cell.font = Font(bold=True)
ws3.column_dimensions["A"].width = 40
ws3.column_dimensions["B"].width = 12

out = ROOT / "reports/rescreen/rescreen_review.xlsx"
wb.save(out)
print(f"Saved: {out}")
print(f"Sheet 'All Scored':   {len(scored)} rows")
print(f"Sheet 'Crawl Failed': {len(failed)} rows")
