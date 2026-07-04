"""Full combined rediscovery: all queries from rounds 1-4 → crawl → classify → Excel.

Steps:
  1. Add the 1 confirmed-relevant R4 article (Chennai) to master corpus
  2. Merge all query plans from rounds 1-4, deduplicate by query text
  3. Run MediaCloud discovery with all unique queries
  4. Filter out all previously-seen URLs (4,282+)
  5. Crawl new URLs with Playwright (robots bypass, academic use)
  6. Apply best ML classifier (tfidf_phrase_svm x title_plus_body)
  7. Write Excel for human review
"""
from __future__ import annotations

import csv
import json
import sqlite3
import subprocess
import sys
import time
import uuid
from collections import Counter
from pathlib import Path

import joblib
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))
sys.stdout.reconfigure(encoding="utf-8")

from crawler.config import Config
from crawler.downloader import Downloader
from crawler.extractor import Extractor
from crawler.review_dedupe import url_keys as _url_keys
from src.model_training.build_text_representations import build_single

# ── Paths ──────────────────────────────────────────────────────────────────────
MASTER_CSV  = ROOT / "reports/master_corpus/master_all_articles.csv"
MODEL_PATH  = ROOT / "reports/model_training/trained_models/best_model.joblib"
MODEL_CFG   = ROOT / "reports/model_training/trained_models/best_model_config.json"
CRAWL_CFG   = ROOT / "config/config_edible_oils_round3.yaml"

RUN_NAME   = "edible_oil_adulteration_full_combined_2026-06-25"
RUN_DIR    = ROOT / "data/runs" / RUN_NAME
OUT_DIR    = ROOT / "reports/full_rediscovery"
COMBINED_QUERY_CSV = RUN_DIR / "proposed_mediacloud_combined_all_rounds.csv"
PREV_URLS_CSV      = RUN_DIR / "mediacloud/outputs/discovery_previously_reviewed_urls.csv"
CHECKPOINT         = OUT_DIR / "_crawl_checkpoint.json"
CRAWL_RESULTS_CSV  = OUT_DIR / "full_rediscovery_crawled.csv"
EXCEL_PATH         = OUT_DIR / "full_rediscovery_review.xlsx"

# Round query plan CSVs from previous runs
QUERY_PLAN_FILES = [
    ROOT / "data/runs/edible_oils_boolean_title_proximity_2026-06-22/proposed_mediacloud_combined_seed_queries.csv",
    ROOT / "data/runs/edible_oil_adulteration_round_02_2026-06-23/proposed_mediacloud_combined_seed_queries.csv",
    ROOT / "data/runs/edible_oil_adulteration_round_03_2026-06-23/proposed_mediacloud_round3_seed_queries.csv",
    ROOT / "data/runs/edible_oil_adulteration_round_04_2026-06-25/proposed_mediacloud_round4_seed_queries.csv",
]

# The 1 confirmed-keep R4 article
R4_ARTICLE = {
    "round_number":      "4",
    "round_name":        "round_04_fresh_discovery",
    "round_description": "Fresh R4 MediaCloud discovery with 6 new keywords (FSDA, raids, unhygienic, etc.)",
    "date_start":        "2021-01-01",
    "date_end":          "2026-06-25",
    "source_run":        RUN_NAME,
    "final_keep":        "1",
    "final_human_label": "relevant",
    "human_review_status": "human_reviewed",
    "human_review_source": "user_confirmed_keep=1",
    "title":             "Over 4,000 litre of adulterated oil seized from Chennai store",
    "source":            "indianexpress.com",
    "date":              "2022-08-23",
    "url":               "https://indianexpress.com/article/cities/chennai/over-4-litre-of-adulterated-oil-seized-from-chennai-store-8107733/",
    "domain":            "indianexpress.com",
    "publication_date":  "2022-08-23",
    "word_count":        "",
    "model_final_label": "relevant",
    "model_confidence":  "0.94",
    "reason":            "Pattern says edible oil is the adulterated/seized/failed product.",
    "evidence_phrase":   "Over 4,000 litre of adulterated oil seized from Chennai store",
    "oil_role":          "adulterated_product",
    "edible_oil_terms":  "",
    "adulteration_action_terms": "adulterated; contaminated; seized; raided; food safety",
    "negative_terms":    "",
    "query_family":      "proximity",
    "query_id":          "r4_proximity_unhygienic_oil",
    "article_id":        "e8bb54fe-96a5-4275-829a-5f7823de6c8d",
    "file_path":         str(ROOT / "data/runs/edible_oil_adulteration_round_04_2026-06-25/mediacloud/cleaned_text/indianexpress.com/757ad69161add26e.txt"),
    "cleaned_text_path": str(ROOT / "data/runs/edible_oil_adulteration_round_04_2026-06-25/mediacloud/cleaned_text/indianexpress.com/757ad69161add26e.txt"),
    "article_text":      "",
    "llm_label":         "",
    "llm_confidence":    "",
    "llm_reason":        "",
}

HIGH_RECALL_T    = 0.35
HIGH_PRECISION_T = 0.65


def read_csv(p: Path) -> list[dict]:
    with p.open(encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def write_csv(p: Path, rows: list[dict], fieldnames: list[str]) -> None:
    with p.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)
    print(f"  Written: {p.name} ({len(rows)} rows)")


def load_checkpoint() -> dict[str, dict]:
    if CHECKPOINT.exists():
        return json.loads(CHECKPOINT.read_text(encoding="utf-8"))
    return {}


def save_checkpoint(done: dict[str, dict]) -> None:
    CHECKPOINT.write_text(json.dumps(done, ensure_ascii=False), encoding="utf-8")


def bucket(prob: float) -> str:
    if prob >= HIGH_PRECISION_T:
        return "candidate_relevant"
    if prob >= HIGH_RECALL_T:
        return "manual_review"
    return "candidate_irrelevant"


# ── Step 1: Add R4 article to master corpus ────────────────────────────────────
def add_r4_article_to_master() -> None:
    rows = read_csv(MASTER_CSV)
    existing_urls = {r.get("url", "").strip() for r in rows}
    if R4_ARTICLE["url"] in existing_urls:
        print(f"[Step 1] R4 article already in master corpus, skipping.")
        return

    # Backfill article_text from cleaned text file
    cleaned_path = Path(R4_ARTICLE["cleaned_text_path"])
    if cleaned_path.exists():
        R4_ARTICLE["article_text"] = cleaned_path.read_text(encoding="utf-8", errors="replace")
        R4_ARTICLE["word_count"]   = str(len(R4_ARTICLE["article_text"].split()))

    fieldnames = list(rows[0].keys()) if rows else list(R4_ARTICLE.keys())
    rows.append({k: R4_ARTICLE.get(k, "") for k in fieldnames})
    write_csv(MASTER_CSV, rows, fieldnames)
    print(f"[Step 1] Added R4 Chennai article to master corpus ({len(rows)} total rows).")


# ── Step 2: Build combined deduplicated query plan ─────────────────────────────
def build_combined_query_plan() -> list[dict]:
    seen_queries: set[str] = set()
    combined: list[dict] = []
    for plan_path in QUERY_PLAN_FILES:
        if not plan_path.exists():
            print(f"  [WARN] Query plan not found: {plan_path}")
            continue
        rows = read_csv(plan_path)
        new = 0
        for row in rows:
            q = row.get("query", "").strip()
            if q and q not in seen_queries:
                seen_queries.add(q)
                combined.append(row)
                new += 1
        print(f"  {plan_path.name}: {len(rows)} queries, {new} unique new")

    # Re-number
    for i, row in enumerate(combined, 1):
        row["query_number"] = str(i)

    RUN_DIR.mkdir(parents=True, exist_ok=True)
    fieldnames = ["query_number", "query_id", "query_family", "query",
                  "signal_mode", "reason", "date_start", "date_end",
                  "collection_count", "hard_exclusion", "breadth"]
    write_csv(COMBINED_QUERY_CSV, combined, fieldnames)
    print(f"[Step 2] Combined query plan: {len(combined)} unique queries -> {COMBINED_QUERY_CSV.name}")
    return combined


# ── Step 3: Run MediaCloud discovery ──────────────────────────────────────────
def run_mediacloud_discovery() -> int:
    config_path = ROOT / "config/config_edible_oils_round4.yaml"
    # Patch config to point to our new run dir — write a temporary config
    import yaml
    with config_path.open(encoding="utf-8") as f:
        cfg_data = yaml.safe_load(f)

    mc_out = RUN_DIR / "mediacloud/outputs"
    cfg_data["paths"] = {
        "raw_html":        str(RUN_DIR / "mediacloud/raw_html"),
        "cleaned_text":    str(RUN_DIR / "mediacloud/cleaned_text"),
        "outputs":         str(mc_out),
        "db":              str(mc_out / "articles.db"),
        "discovered_urls": str(mc_out / "discovered_urls.jsonl"),
        "articles_jsonl":  str(mc_out / "articles.jsonl"),
        "articles_csv":    str(mc_out / "articles.csv"),
        "report":          str(mc_out / "report.json"),
    }
    cfg_data["discovery"]["mediacloud"]["previously_reviewed_urls_csv"] = str(PREV_URLS_CSV)

    tmp_config = RUN_DIR / "config_full_combined.yaml"
    tmp_config.write_text(yaml.safe_dump(cfg_data, sort_keys=False, allow_unicode=False), encoding="utf-8")

    baseline_run = ROOT / "data/runs/edible_oil_adulteration_round_03_2026-06-23"
    cmd = [
        sys.executable, "scripts/run_combined_mediacloud_discovery.py",
        "--config", str(tmp_config),
        "--query-plan", str(COMBINED_QUERY_CSV),
        "--run-dir", str(RUN_DIR),
        "--baseline-run-dir", str(baseline_run),
    ]
    print(f"\n[Step 3] Running MediaCloud discovery ({len(read_csv(COMBINED_QUERY_CSV))} queries) ...")
    result = subprocess.run(cmd, cwd=ROOT)
    if result.returncode != 0:
        raise RuntimeError(f"Discovery failed (exit {result.returncode})")

    db_path = mc_out / "articles.db"
    with sqlite3.connect(db_path) as con:
        n = con.execute("SELECT COUNT(*) FROM discovered_urls").fetchone()[0]
    print(f"[Step 3] Discovery complete: {n} URLs in DB")
    return n


# ── Step 4: Collect all previously-seen URLs + write dedup file ────────────────
def build_seen_urls() -> set[str]:
    seen: set[str] = set()
    for row in read_csv(MASTER_CSV):
        u = row.get("url", "").strip()
        if u:
            seen.update(_url_keys(u))

    for p in (ROOT / "data/runs").rglob("discovery_url_review.csv"):
        for row in read_csv(p):
            u = row.get("url", "").strip()
            if u:
                seen.update(_url_keys(u))

    for p in (ROOT / "data/runs").rglob("metadata_all_articles_review.csv"):
        for row in read_csv(p):
            u = row.get("url", "").strip()
            if u:
                seen.update(_url_keys(u))

    rescreen = ROOT / "reports/rescreen/rescreen_all_dropped.csv"
    if rescreen.exists():
        for row in read_csv(rescreen):
            u = row.get("url", "").strip()
            if u:
                seen.update(_url_keys(u))

    PREV_URLS_CSV.parent.mkdir(parents=True, exist_ok=True)
    with PREV_URLS_CSV.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=["url"])
        w.writeheader()
        for u in sorted(seen):
            w.writerow({"url": u})

    print(f"[Step 4] Previously-seen URL keys: {len(seen)} -> {PREV_URLS_CSV.name}")
    return seen


# ── Step 5: Get newly discovered URLs not in seen set ─────────────────────────
def get_new_discovered_urls(seen_keys: set[str]) -> list[dict]:
    db_path = RUN_DIR / "mediacloud/outputs/articles.db"
    with sqlite3.connect(db_path) as con:
        con.row_factory = sqlite3.Row
        rows = con.execute(
            "SELECT id, url, discovery_method, query_used, title_snippet, source, domain, published_date "
            "FROM discovered_urls ORDER BY published_date DESC, id ASC"
        ).fetchall()

    new: list[dict] = []
    for row in rows:
        d = dict(row)
        u = d.get("url", "").strip()
        if u and not (_url_keys(u) & seen_keys):
            query_family = d.get("discovery_method", "").removeprefix("mediacloud_")
            new.append({
                "url":          u,
                "title":        d.get("title_snippet", ""),
                "source":       d.get("source", ""),
                "date":         d.get("published_date", ""),
                "query_family": query_family,
                "query_used":   d.get("query_used", ""),
                "domain":       d.get("domain", ""),
            })

    print(f"[Step 5] New URLs after dedup: {len(new)} (from {len(rows)} discovered)")
    return new


# ── Step 6 & 7: Crawl + classify ──────────────────────────────────────────────
def crawl_and_classify(candidates: list[dict]) -> dict[str, dict]:
    model_cfg = json.loads(MODEL_CFG.read_text(encoding="utf-8"))
    repr_name = model_cfg.get("representation", "title_plus_body")
    pipeline  = joblib.load(MODEL_PATH)
    print(f"[Step 6] Model: {model_cfg.get('model_name')} x {repr_name}")

    cfg = Config(CRAWL_CFG)
    cfg.raw.setdefault("crawl", {}).update({
        "respect_robots_txt": False,
        "use_playwright":     True,
        "playwright_first":   True,
        "timeout_seconds":    45,
        "delay_seconds":      1.5,
        "max_retries":        1,
    })
    downloader = Downloader(cfg)
    extractor  = Extractor(cfg)

    done = load_checkpoint()
    remaining = [c for c in candidates if c["url"] not in done]
    print(f"[Step 6] Crawling {len(remaining)} URLs (checkpoint: {len(done)} already done) ...")

    for i, cand in enumerate(remaining, 1):
        url   = cand["url"]
        title = cand.get("title", "")
        print(f"[{i + len(done)}/{len(candidates)}] {url[:90]}", flush=True)

        result = downloader.download(url)
        status = result.get("status", "failed")

        if status != "success":
            record = {**cand, "crawl_status": status,
                      "article_text": "", "word_count": 0,
                      "prob": None, "bucket": "crawl_failed"}
        else:
            html = result.get("raw_html") or ""
            final_url = result.get("url") or url
            ext  = extractor.extract(final_url, html,
                                     raw_html_path=result.get("raw_html_path"))
            text = ext.get("article_text") or ""
            wc   = ext.get("word_count") or 0

            if text:
                rep  = build_single(title, text, repr_name)
                prob = float(pipeline.predict_proba([rep])[0, 1])
                bkt  = bucket(prob)
                print(f"  -> {wc}w  prob={prob:.3f}  [{bkt}]", flush=True)
            else:
                prob, bkt = None, "no_text"
                print(f"  -> crawled, no text", flush=True)

            record = {**cand, "crawl_status": "success",
                      "article_text": text, "word_count": wc,
                      "prob": prob, "bucket": bkt}

        done[url] = record
        if i % 10 == 0:
            save_checkpoint(done)

    downloader.close()
    save_checkpoint(done)
    return done


# ── Step 8: Write Excel ────────────────────────────────────────────────────────
def write_excel(done: dict[str, dict]) -> None:
    records = sorted(
        [r for r in done.values() if r.get("bucket") != "crawl_failed"],
        key=lambda r: (r.get("prob") or 0),
        reverse=True,
    )
    failed = [r for r in done.values() if r.get("bucket") == "crawl_failed"]

    GREEN  = PatternFill("solid", fgColor="C6EFCE")
    YELLOW = PatternFill("solid", fgColor="FFEB9C")
    RED    = PatternFill("solid", fgColor="FFC7CE")
    GREY   = PatternFill("solid", fgColor="D9D9D9")
    HEADER = PatternFill("solid", fgColor="2F5496")
    HDR_FT = Font(bold=True, color="FFFFFF")

    wb = openpyxl.Workbook()

    # ── Sheet 1: All Scored ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = "All Scored"
    cols = ["keep", "prob", "bucket", "query_family", "title", "source", "date",
            "url", "domain", "word_count", "query_used"]
    ws.append(cols)
    for cell in ws[1]:
        cell.font = HDR_FT
        cell.fill = HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in records:
        prob = row.get("prob")
        bkt  = row.get("bucket", "")
        fill = GREEN if bkt == "candidate_relevant" else (
               YELLOW if bkt == "manual_review" else RED)
        values = [
            "",
            round(prob, 3) if prob is not None else "",
            bkt,
            row.get("query_family", ""),
            row.get("title", ""),
            row.get("source", ""),
            row.get("date", ""),
            row.get("url", ""),
            row.get("domain", ""),
            row.get("word_count", ""),
            (row.get("query_used") or "")[:120],
        ]
        ws.append(values)
        for cell in ws[ws.max_row]:
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=False, vertical="top")

    # Column widths
    widths = [6, 7, 18, 14, 60, 22, 12, 60, 24, 10, 60]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ── Sheet 2: Crawl Failed ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Crawl Failed")
    ws2.append(["url", "source", "title", "query_family", "crawl_status"])
    for cell in ws2[1]:
        cell.font = HDR_FT
        cell.fill = HEADER
    for row in failed:
        ws2.append([row.get("url",""), row.get("source",""),
                    row.get("title",""), row.get("query_family",""),
                    row.get("crawl_status","")])
        for cell in ws2[ws2.max_row]:
            cell.fill = GREY

    # ── Sheet 3: Stats ────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Stats")
    bkt_counts = Counter(r.get("bucket", "") for r in done.values())
    fam_counts = Counter(r.get("query_family", "") for r in records)
    stats_rows = [
        ("Total discovered + new after dedup", len(done)),
        ("Crawl failed / robots blocked", len(failed)),
        ("Scored articles", len(records)),
        ("candidate_relevant (prob >= 0.65)", bkt_counts.get("candidate_relevant", 0)),
        ("manual_review (0.35 <= prob < 0.65)", bkt_counts.get("manual_review", 0)),
        ("candidate_irrelevant (prob < 0.35)", bkt_counts.get("candidate_irrelevant", 0)),
        ("no_text", bkt_counts.get("no_text", 0)),
        ("", ""),
        ("Query family breakdown", ""),
    ]
    for fam, cnt in fam_counts.most_common():
        stats_rows.append((f"  {fam}", cnt))
    ws3.column_dimensions["A"].width = 45
    ws3.column_dimensions["B"].width = 12
    for row in stats_rows:
        ws3.append(row)

    EXCEL_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(EXCEL_PATH)
    print(f"[Step 8] Excel written: {EXCEL_PATH} ({len(records)} scored, {len(failed)} crawl-failed)")


def main() -> int:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    RUN_DIR.mkdir(parents=True, exist_ok=True)

    # 1. Add R4 article to master corpus
    add_r4_article_to_master()

    # 2. Build combined query plan
    build_combined_query_plan()

    # 4. Build seen-URL set (BEFORE discovery so we can write the CSV the pipeline reads)
    seen_keys = build_seen_urls()

    # 3. Run MediaCloud discovery
    run_mediacloud_discovery()

    # 5. Get newly discovered URLs
    new_urls = get_new_discovered_urls(seen_keys)

    if not new_urls:
        print("[Done] No new URLs discovered after dedup. Nothing to crawl.")
        return 0

    # Write CSV of candidates
    cand_fieldnames = ["url", "title", "source", "date", "query_family", "query_used", "domain"]
    write_csv(OUT_DIR / "new_discovered_urls.csv", new_urls, cand_fieldnames)

    # 6 & 7. Crawl + classify
    done = crawl_and_classify(new_urls)

    # Write full results CSV
    fieldnames = ["url", "title", "source", "date", "query_family", "domain",
                  "crawl_status", "word_count", "prob", "bucket", "article_text"]
    all_records = sorted(done.values(), key=lambda r: r.get("prob") or 0, reverse=True)
    write_csv(CRAWL_RESULTS_CSV, all_records, fieldnames)

    # 8. Write Excel
    write_excel(done)

    # Summary
    bkt_counts = Counter(r.get("bucket", "") for r in done.values())
    print(f"\n{'='*60}")
    print(f"New URLs found:           {len(new_urls)}")
    print(f"Crawled successfully:      {sum(1 for r in done.values() if r.get('crawl_status')=='success')}")
    print(f"candidate_relevant:        {bkt_counts.get('candidate_relevant', 0)}")
    print(f"manual_review:             {bkt_counts.get('manual_review', 0)}")
    print(f"candidate_irrelevant:      {bkt_counts.get('candidate_irrelevant', 0)}")
    print(f"Excel: {EXCEL_PATH}")
    print(f"{'='*60}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
