from __future__ import annotations

import argparse
import csv
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parent.parent
DEFAULT_INPUT = (
    ROOT
    / "reports"
    / "edible_oil_adulteration_round_01"
    / "not_terms_review"
    / "round_01_not_term_overlap_risk.csv"
)
DEFAULT_OUTPUT = (
    ROOT
    / "reports"
    / "edible_oil_adulteration_round_01"
    / "not_terms_review"
    / "round_01_not_terms_review.xlsx"
)

COLUMNS = [
    ("keep", "keep_as_NOT", 12),
    ("not_candidate", "candidate_not_term", 30),
    ("risk_flag", "risk_flag", 30),
    ("irrelevant_category", "category", 22),
    ("irrelevant_composite_score", "score", 12),
    ("irrelevant_document_frequency", "doc_freq_irrelevant", 18),
    ("irrelevant_total_frequency", "phrase_freq_irrelevant", 20),
    ("positive_exact_kept", "also_positive_kept", 18),
    ("positive_overlap_terms", "positive_overlap_terms", 38),
    ("example_irrelevant_title", "example_irrelevant_title", 55),
    ("example_context", "example_context", 75),
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Create a review workbook for candidate NOT terms.")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    return parser.parse_args()


def read_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def main() -> None:
    args = parse_args()
    rows = read_rows(args.input)
    wb = Workbook()
    ws = wb.active
    ws.title = "NOT Term Review"
    write_review_sheet(ws, rows)
    write_summary_sheet(wb.create_sheet("Summary"), rows, args.input, args.output)
    write_instructions_sheet(wb.create_sheet("Instructions"))
    args.output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)
    print(f"Wrote {args.output}")
    print(f"Rows: {len(rows)}")


def write_review_sheet(ws, rows: list[dict[str, str]]) -> None:
    thin = thin_border()
    header_fill = PatternFill("solid", fgColor="1F4E78")
    keep_fill = PatternFill("solid", fgColor="FFF2CC")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, (_key, header, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws["A1"].comment = Comment(
        "Enter 1 only if this should be used as a NOT/exclusion term. Use 0 if it could remove relevant edible-oil adulteration articles.",
        "Codex",
    )
    validation = DataValidation(
        type="list",
        formula1='"0,1"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Use 0 or 1",
        error="Enter 1 to keep as NOT term or 0 to reject.",
    )
    ws.add_data_validation(validation)

    for row_idx, row in enumerate(rows, start=2):
        row_fill = risk_fill(row.get("risk_flag", ""))
        for col_idx, (key, _header, _width) in enumerate(COLUMNS, start=1):
            value = "" if key == "keep" else row.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.font = Font(size=10)
            cell.fill = keep_fill if key == "keep" else row_fill
        ws.row_dimensions[row_idx].height = 58

    if rows:
        validation.add(f"A2:A{len(rows) + 1}")
        ws.conditional_formatting.add(
            f"A2:A{len(rows) + 1}",
            CellIsRule(operator="equal", formula=["1"], fill=PatternFill("solid", fgColor="C6EFCE")),
        )
        ws.conditional_formatting.add(
            f"A2:A{len(rows) + 1}",
            CellIsRule(operator="equal", formula=["0"], fill=PatternFill("solid", fgColor="F4CCCC")),
        )
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions


def risk_fill(risk: str) -> PatternFill:
    if risk == "possible_not_candidate":
        return PatternFill("solid", fgColor="E2F0D9")
    if risk == "risky_domain_term_may_drop_relevant":
        return PatternFill("solid", fgColor="FFF2CC")
    if risk == "risky_overlap_with_positive_keyword":
        return PatternFill("solid", fgColor="FCE4D6")
    if risk == "do_not_not_exact_positive_kept":
        return PatternFill("solid", fgColor="F4CCCC")
    return PatternFill("solid", fgColor="FFFFFF")


def write_summary_sheet(ws, rows: list[dict[str, str]], input_path: Path, output_path: Path) -> None:
    ws["A1"] = "NOT Term Review"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A3"] = "Input CSV"
    ws["B3"] = str(input_path)
    ws["A4"] = "Output workbook"
    ws["B4"] = str(output_path)
    ws["A5"] = "Coding"
    ws["B5"] = "keep_as_NOT: 1 = use as NOT/exclusion term; 0 = do not use; blank = unreviewed"

    start = 7
    for title, counts in [
        ("Risk Flag Counts", Counter(row.get("risk_flag", "") for row in rows)),
        ("Category Counts", Counter(row.get("irrelevant_category", "") for row in rows)),
    ]:
        ws.cell(start, 1, title)
        ws.cell(start, 1).font = Font(bold=True, size=12)
        start += 1
        ws.cell(start, 1, "Value")
        ws.cell(start, 2, "Count")
        for cell in ws[start]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
        start += 1
        for value, count in sorted(counts.items()):
            ws.cell(start, 1, value)
            ws.cell(start, 2, count)
            start += 1
        start += 2
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 100


def write_instructions_sheet(ws) -> None:
    lines = [
        ("Goal", "Pick terms that safely exclude junk Round 2 articles."),
        ("Use 1", "Use 1 only for terms that clearly indicate out-of-scope topics, e.g. export/price/international commodity stories, non-food oil, reused-oil safety-only topics."),
        ("Use 0", "Use 0 for edible-oil product terms, adulteration terms, enforcement terms, and anything that could appear in a real relevant article."),
        ("Risk flag", "`possible_not_candidate` is safer. Anything marked risky or exact positive overlap should usually be 0."),
        ("Conservative rule", "A missing NOT term is better than excluding true relevant articles."),
    ]
    ws["A1"] = "Instructions"
    ws["A1"].font = Font(bold=True, size=14)
    for idx, (label, text) in enumerate(lines, start=3):
        ws.cell(idx, 1, label)
        ws.cell(idx, 2, text)
        ws.cell(idx, 1).font = Font(bold=True)
        ws.cell(idx, 1).alignment = Alignment(vertical="top")
        ws.cell(idx, 2).alignment = Alignment(vertical="top", wrap_text=True)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 100


def thin_border() -> Border:
    side = Side(style="thin", color="BFBFBF")
    return Border(left=side, right=side, top=side, bottom=side)


if __name__ == "__main__":
    main()
