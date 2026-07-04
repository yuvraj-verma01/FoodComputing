from __future__ import annotations

import argparse
import csv
import json
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


DEFAULT_OUTPUT_DIR = Path(
    "data/runs/edible_oils_boolean_title_proximity_2026-06-22/"
    "mediacloud/outputs/oil_relevance"
)

REVIEW_SHEETS = ["relevant", "manual_review", "irrelevant"]
OUTPUT_COLUMNS = [
    "human_keep",
    "human_label",
    "human_review_status",
    "human_review_source",
    "model_final_label",
    "model_confidence",
    "title",
    "source",
    "date",
    "url",
    "reason",
    "evidence_phrase",
    "llm_label",
    "llm_confidence",
    "llm_reason",
    "oil_role",
    "edible_oil_terms",
    "adulteration_action_terms",
    "negative_terms",
    "query_family",
    "query_id",
    "article_id",
    "file_path",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Merge human 1/0 review marks into final edible-oil corpus outputs."
    )
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument(
        "--main-review",
        type=Path,
        default=DEFAULT_OUTPUT_DIR / "oil_relevance_review.xlsx",
    )
    parser.add_argument(
        "--safety-review",
        type=Path,
        default=DEFAULT_OUTPUT_DIR / "rejected_safety_review.xlsx",
    )
    return parser.parse_args()


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def write_csv(path: Path, rows: list[dict[str, Any]], columns: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def norm_keep(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text in {"1", "1.0"}:
        return "1"
    if text in {"0", "0.0"}:
        return "0"
    return ""


def sheet_records(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    try:
        headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    except StopIteration:
        return []
    records = []
    for row in rows:
        records.append({headers[idx]: value for idx, value in enumerate(row) if idx < len(headers)})
    return records


def collect_main_labels(path: Path) -> dict[str, list[dict[str, str]]]:
    labels: dict[str, list[dict[str, str]]] = defaultdict(list)
    if not path.exists():
        return labels
    for sheet in REVIEW_SHEETS:
        for row in sheet_records(path, sheet):
            keep = norm_keep(row.get("keep") if "keep" in row else row.get("Keep"))
            url = str(row.get("url") or "").strip()
            if not keep or not url:
                continue
            labels[url].append(
                {
                    "keep": keep,
                    "source": f"{path.name}:{sheet}",
                    "title": str(row.get("title") or ""),
                }
            )
    return labels


def collect_safety_labels(path: Path) -> dict[str, list[dict[str, str]]]:
    labels: dict[str, list[dict[str, str]]] = defaultdict(list)
    if not path.exists():
        return labels
    for row in sheet_records(path, "safety_review"):
        keep = norm_keep(row.get("Keep") if "Keep" in row else row.get("keep"))
        url = str(row.get("url") or "").strip()
        if not keep or not url:
            continue
        labels[url].append(
            {
                "keep": keep,
                "source": f"{path.name}:safety_review",
                "title": str(row.get("title") or ""),
            }
        )
    return labels


def merge_label_sources(
    main_labels: dict[str, list[dict[str, str]]],
    safety_labels: dict[str, list[dict[str, str]]],
) -> tuple[dict[str, dict[str, str]], list[dict[str, str]], dict[str, int]]:
    merged: dict[str, dict[str, str]] = {}
    conflicts: list[dict[str, str]] = []
    counts = Counter()
    all_urls = set(main_labels) | set(safety_labels)
    for url in sorted(all_urls):
        entries = main_labels.get(url, []) + safety_labels.get(url, [])
        keep_values = sorted({entry["keep"] for entry in entries})
        if len(keep_values) > 1:
            conflicts.append(
                {
                    "url": url,
                    "keep_values": "; ".join(keep_values),
                    "sources": "; ".join(f"{e['source']}={e['keep']}" for e in entries),
                    "title": next((e["title"] for e in entries if e.get("title")), ""),
                }
            )
            counts["conflicts"] += 1
            continue
        keep = keep_values[0]
        merged[url] = {
            "human_keep": keep,
            "human_label": "relevant" if keep == "1" else "irrelevant",
            "human_review_source": "; ".join(entry["source"] for entry in entries),
        }
        counts[f"keep_{keep}"] += 1
    return merged, conflicts, dict(counts)


def apply_labels(
    article_rows: list[dict[str, str]],
    merged_labels: dict[str, dict[str, str]],
) -> list[dict[str, Any]]:
    rows = []
    for row in article_rows:
        url = row.get("url", "")
        human = merged_labels.get(url)
        if human:
            human_keep = human["human_keep"]
            human_label = human["human_label"]
            status = "human_reviewed"
            source = human["human_review_source"]
        else:
            human_keep = ""
            human_label = ""
            status = "needs_review"
            source = ""
        rows.append(
            {
                "human_keep": human_keep,
                "human_label": human_label,
                "human_review_status": status,
                "human_review_source": source,
                "model_final_label": row.get("final_label", ""),
                "model_confidence": row.get("confidence", ""),
                **row,
            }
        )
    return rows


def add_sheet(wb: Workbook, title: str, rows: list[dict[str, Any]], columns: list[str]) -> None:
    ws = wb.create_sheet(title)
    ws.append(columns)
    for row in rows:
        ws.append([row.get(col, "") for col in columns])
    header_fill = PatternFill("solid", fgColor="1F4E78")
    keep_fill = PatternFill("solid", fgColor="E2F0D9")
    drop_fill = PatternFill("solid", fgColor="FCE4D6")
    review_fill = PatternFill("solid", fgColor="FFF2CC")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    widths = {
        "A": 11,
        "B": 16,
        "C": 18,
        "D": 38,
        "E": 16,
        "F": 12,
        "G": 54,
        "H": 22,
        "I": 14,
        "J": 52,
        "K": 42,
        "L": 42,
        "M": 14,
        "N": 14,
        "O": 42,
        "P": 18,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    keep_col = columns.index("human_keep") + 1 if "human_keep" in columns else None
    url_col = columns.index("url") + 1 if "url" in columns else None
    for row_idx in range(2, ws.max_row + 1):
        keep = ws.cell(row=row_idx, column=keep_col).value if keep_col else ""
        fill = keep_fill if str(keep) == "1" else drop_fill if str(keep) == "0" else review_fill
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        if url_col:
            cell = ws.cell(row=row_idx, column=url_col)
            if cell.value:
                cell.hyperlink = str(cell.value)
                cell.style = "Hyperlink"
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def write_workbook(
    path: Path,
    rows: list[dict[str, Any]],
    conflicts: list[dict[str, str]],
    summary: dict[str, Any],
) -> None:
    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "summary"
    summary_ws.append(["metric", "value"])
    for key, value in summary.items():
        summary_ws.append([key, json.dumps(value, ensure_ascii=False) if isinstance(value, dict) else value])
    summary_ws.column_dimensions["A"].width = 34
    summary_ws.column_dimensions["B"].width = 90
    for cell in summary_ws[1]:
        cell.font = Font(bold=True)

    add_sheet(wb, "all_articles", rows, OUTPUT_COLUMNS)
    add_sheet(wb, "human_relevant", [r for r in rows if r.get("human_keep") == "1"], OUTPUT_COLUMNS)
    add_sheet(wb, "human_irrelevant", [r for r in rows if r.get("human_keep") == "0"], OUTPUT_COLUMNS)
    add_sheet(wb, "needs_review", [r for r in rows if not r.get("human_keep")], OUTPUT_COLUMNS)
    add_sheet(wb, "conflicts", conflicts, ["title", "url", "keep_values", "sources"])
    wb.save(path)


def main() -> None:
    args = parse_args()
    output_dir = args.output_dir
    article_rows = read_csv(output_dir / "all_articles_review.csv")
    main_labels = collect_main_labels(args.main_review)
    safety_labels = collect_safety_labels(args.safety_review)
    merged_labels, conflicts, label_counts = merge_label_sources(main_labels, safety_labels)
    final_rows = apply_labels(article_rows, merged_labels)

    final_counts = Counter(row["human_label"] or "needs_review" for row in final_rows)
    model_counts = Counter(row.get("model_final_label") or "" for row in final_rows)
    reviewed_urls_not_in_current = sorted(set(merged_labels) - {row.get("url", "") for row in article_rows})
    summary = {
        "created_at": utc_now(),
        "current_article_rows": len(article_rows),
        "main_review_marked_urls": len(main_labels),
        "safety_review_marked_urls": len(safety_labels),
        "merged_human_label_urls": len(merged_labels),
        "reviewed_urls_not_in_current_articles": len(reviewed_urls_not_in_current),
        "conflicts": len(conflicts),
        "human_label_counts_on_current_articles": dict(final_counts),
        "model_label_counts_on_current_articles": dict(model_counts),
        "review_label_counts_before_current_match": label_counts,
    }

    write_csv(output_dir / "human_reviewed_all_articles.csv", final_rows, OUTPUT_COLUMNS)
    write_csv(
        output_dir / "human_relevant_articles.csv",
        [row for row in final_rows if row.get("human_keep") == "1"],
        OUTPUT_COLUMNS,
    )
    write_csv(
        output_dir / "human_irrelevant_articles.csv",
        [row for row in final_rows if row.get("human_keep") == "0"],
        OUTPUT_COLUMNS,
    )
    write_csv(
        output_dir / "human_needs_review_articles.csv",
        [row for row in final_rows if not row.get("human_keep")],
        OUTPUT_COLUMNS,
    )
    write_csv(output_dir / "human_review_conflicts.csv", conflicts, ["title", "url", "keep_values", "sources"])
    (output_dir / "human_review_summary.json").write_text(
        json.dumps(summary, indent=2),
        encoding="utf-8",
    )
    write_workbook(output_dir / "human_reviewed_corpus.xlsx", final_rows, conflicts, summary)

    print(json.dumps(summary, indent=2))
    print(f"Workbook written: {(output_dir / 'human_reviewed_corpus.xlsx').resolve()}")


if __name__ == "__main__":
    main()
