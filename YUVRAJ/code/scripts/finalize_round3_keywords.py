"""Read final keep=1 decisions from round_03_keyword_review.xlsx and write CSVs."""
from __future__ import annotations
import csv, json
from datetime import datetime, timezone
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
OUT  = ROOT / "reports/edible_oil_adulteration_round_02/round_03_keyword_review"
XLSX = OUT / "round_03_keyword_review.xlsx"

wb = load_workbook(XLSX, data_only=True)

# Positive keywords
ws = wb["Positive Keywords"]
final_pos = []
for row in ws.iter_rows(min_row=2, values_only=True):
    keep = str(row[0]).strip() if row[0] is not None else ""
    if keep == "1":
        final_pos.append({
            "keyword":         str(row[1] or "").strip(),
            "suggested_label": str(row[2] or "").strip(),
            "source":          str(row[3] or "").strip(),
            "category":        str(row[4] or "").strip(),
        })

# NOT terms
ws2 = wb["NOT Terms"]
final_not = []
for row in ws2.iter_rows(min_row=2, values_only=True):
    keep = str(row[0]).strip() if row[0] is not None else ""
    if keep == "1":
        final_not.append({
            "not_term":  str(row[1] or "").strip(),
            "risk_flag": str(row[2] or "").strip(),
            "category":  str(row[3] or "").strip(),
        })

def write_csv(p: Path, rows: list[dict]) -> None:
    with p.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader(); w.writerows(rows)

write_csv(OUT / "round_03_final_positive_keywords.csv", final_pos)
write_csv(OUT / "round_03_final_not_terms.csv", final_not)

summary = {
    "created_at": datetime.now(timezone.utc).isoformat(),
    "final_positive_keywords": len(final_pos),
    "final_not_terms": len(final_not),
    "positive_keywords": [r["keyword"] for r in final_pos],
    "not_terms": [r["not_term"] for r in final_not],
}
(OUT / "round_03_final_keyword_summary.json").write_text(
    json.dumps(summary, indent=2, ensure_ascii=False), encoding="utf-8")

print(f"Final positive keywords: {len(final_pos)}")
for r in final_pos:
    print(f"  [{r['suggested_label']}] {r['keyword']}")
print(f"\nFinal NOT terms: {len(final_not)}")
for r in final_not:
    print(f"  {r['not_term']}")
