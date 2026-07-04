from __future__ import annotations

import argparse
import csv
import json
import shutil
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parent.parent
DEFAULT_OUTPUT_DIR = ROOT / Path(
    "data/runs/edible_oils_boolean_title_proximity_2026-06-22/"
    "mediacloud/outputs/oil_relevance"
)
DEFAULT_REPORT_DIR = ROOT / "reports" / "edible_oil_adulteration_round_01"

ROUND_INFO = {
    "round_number": "1",
    "round_name": "round_01_mediacloud_boolean_title_proximity",
    "round_description": "MediaCloud Indian collections, Jan 1 2021 to Jun 22 2026, boolean + title-only + proximity queries; ghee/vanaspati excluded; human-reviewed relevance labels.",
    "date_start": "2021-01-01",
    "date_end": "2026-06-22",
    "source_run": "edible_oils_boolean_title_proximity_2026-06-22",
}

ROUND_COLUMNS = [
    "round_number",
    "round_name",
    "round_description",
    "date_start",
    "date_end",
    "source_run",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Package a final edible-oil corpus round report.")
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--report-dir", type=Path, default=DEFAULT_REPORT_DIR)
    parser.add_argument("--round-number", default=ROUND_INFO["round_number"])
    parser.add_argument("--round-name", default=ROUND_INFO["round_name"])
    parser.add_argument("--round-description", default=ROUND_INFO["round_description"])
    parser.add_argument("--date-start", default=ROUND_INFO["date_start"])
    parser.add_argument("--date-end", default=ROUND_INFO["date_end"])
    parser.add_argument("--source-run", default=ROUND_INFO["source_run"])
    return parser.parse_args()


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def round_prefix(round_number: str) -> str:
    try:
        return f"round_{int(round_number):02d}"
    except ValueError:
        normalized = "".join(ch if ch.isalnum() else "_" for ch in round_number.lower()).strip("_")
        return f"round_{normalized}"


def round_info(args: argparse.Namespace) -> dict[str, str]:
    return {
        "round_number": str(args.round_number),
        "round_name": str(args.round_name),
        "round_description": str(args.round_description),
        "date_start": str(args.date_start),
        "date_end": str(args.date_end),
        "source_run": str(args.source_run),
    }


def add_round_columns(rows: list[dict[str, str]], info: dict[str, str]) -> list[dict[str, str]]:
    return [{**info, **row} for row in rows]


def write_csv(path: Path, rows: list[dict[str, Any]], columns: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def write_jsonl(path: Path, rows: list[dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8") as handle:
        for row in rows:
            handle.write(json.dumps(row, ensure_ascii=False) + "\n")


def add_sheet(wb: Workbook, title: str, rows: list[dict[str, Any]], columns: list[str]) -> None:
    ws = wb.create_sheet(title)
    ws.append(columns)
    for row in rows:
        ws.append([row.get(col, "") for col in columns])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    keep_fill = PatternFill("solid", fgColor="E2F0D9")
    drop_fill = PatternFill("solid", fgColor="FCE4D6")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = {
        "A": 12,
        "B": 34,
        "C": 72,
        "D": 13,
        "E": 13,
        "F": 36,
        "G": 10,
        "H": 18,
        "I": 18,
        "J": 40,
        "K": 16,
        "L": 12,
        "M": 58,
        "N": 24,
        "O": 14,
        "P": 54,
        "Q": 22,
        "R": 15,
        "S": 12,
        "T": 44,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    keep_col = columns.index("final_keep") + 1 if "final_keep" in columns else None
    url_col = columns.index("url") + 1 if "url" in columns else None
    for row_idx in range(2, ws.max_row + 1):
        fill = None
        if keep_col:
            keep = str(ws.cell(row=row_idx, column=keep_col).value or "")
            fill = keep_fill if keep == "1" else drop_fill if keep == "0" else None
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        if url_col:
            cell = ws.cell(row=row_idx, column=url_col)
            if cell.value:
                cell.hyperlink = str(cell.value)
                cell.style = "Hyperlink"
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def write_workbook(path: Path, all_rows: list[dict[str, Any]], summary: dict[str, Any]) -> None:
    public_columns = [col for col in all_rows[0].keys() if col != "article_text"] if all_rows else []
    relevant = [row for row in all_rows if row.get("final_keep") == "1"]
    irrelevant = [row for row in all_rows if row.get("final_keep") == "0"]

    wb = Workbook()
    ws = wb.active
    ws.title = "summary"
    ws.append(["metric", "value"])
    for key, value in summary.items():
        ws.append([key, json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list)) else value])
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 95
    for cell in ws[1]:
        cell.font = Font(bold=True)

    add_sheet(wb, "all_articles", all_rows, public_columns)
    add_sheet(wb, "relevant", relevant, public_columns)
    add_sheet(wb, "irrelevant", irrelevant, public_columns)
    wb.save(path)


def copy_if_exists(src: Path, dst: Path) -> str:
    if not src.exists():
        return "missing"
    dst.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(src, dst)
    return "copied"


def write_readme(path: Path, summary: dict[str, Any], info: dict[str, str], prefix: str) -> None:
    lines = [
        f"# Edible Oil Adulteration Corpus - {prefix.replace('_', ' ').title()}",
        "",
        "This folder is a cleaned, human-reviewed output package for one MediaCloud run.",
        "",
        "## Scope",
        "",
        f"- Round number: {info['round_number']}",
        f"- Round name: `{info['round_name']}`",
        f"- Date range: {info['date_start']} to {info['date_end']}",
        "- Geography: India-focused MediaCloud national and state/local collections",
        "- Food item: edible/cooking oils only; ghee and vanaspati excluded",
        "- Reused/used cooking oil without adulteration/fraud context is treated as irrelevant for this corpus.",
        "",
        "## Counts",
        "",
        f"- Total reviewed articles: {summary['total_articles']}",
        f"- Relevant articles: {summary['relevant_articles']}",
        f"- Irrelevant articles: {summary['irrelevant_articles']}",
        f"- Unresolved articles: {summary['unresolved_articles']}",
        "",
        "## Key Files",
        "",
        f"- `{prefix}_all_articles.csv`: all human-reviewed rows, with article text.",
        f"- `{prefix}_relevant_articles.csv`: kept articles, with article text.",
        f"- `{prefix}_irrelevant_articles.csv`: dropped articles, with article text.",
        f"- `{prefix}_relevant_articles.jsonl`: kept articles in JSONL form.",
        f"- `{prefix}_review_workbook.xlsx`: formatted Excel workbook without full article text.",
        f"- `{prefix}_summary.json`: machine-readable summary and provenance.",
        "",
    ]
    path.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    args = parse_args()
    output_dir = args.output_dir
    report_dir = args.report_dir
    report_dir.mkdir(parents=True, exist_ok=True)
    info = round_info(args)
    prefix = round_prefix(info["round_number"])

    final_all = add_round_columns(read_csv(output_dir / "final_human_reviewed_all_articles.csv"), info)
    relevant = [row for row in final_all if row.get("final_keep") == "1"]
    irrelevant = [row for row in final_all if row.get("final_keep") == "0"]
    unresolved = [row for row in final_all if row.get("final_keep") not in {"0", "1"}]

    columns = ROUND_COLUMNS + [
        col for col in final_all[0].keys() if col not in ROUND_COLUMNS
    ]
    summary = {
        "created_at": utc_now(),
        **info,
        "total_articles": len(final_all),
        "relevant_articles": len(relevant),
        "irrelevant_articles": len(irrelevant),
        "unresolved_articles": len(unresolved),
        "label_counts": dict(Counter(row.get("final_human_label") or "unresolved" for row in final_all)),
        "outputs": {
            "all_csv": f"{prefix}_all_articles.csv",
            "relevant_csv": f"{prefix}_relevant_articles.csv",
            "irrelevant_csv": f"{prefix}_irrelevant_articles.csv",
            "relevant_jsonl": f"{prefix}_relevant_articles.jsonl",
            "workbook": f"{prefix}_review_workbook.xlsx",
        },
    }

    write_csv(report_dir / f"{prefix}_all_articles.csv", final_all, columns)
    write_csv(report_dir / f"{prefix}_relevant_articles.csv", relevant, columns)
    write_csv(report_dir / f"{prefix}_irrelevant_articles.csv", irrelevant, columns)
    write_jsonl(report_dir / f"{prefix}_relevant_articles.jsonl", relevant)
    write_workbook(report_dir / f"{prefix}_review_workbook.xlsx", final_all, summary)
    (report_dir / f"{prefix}_summary.json").write_text(
        json.dumps(summary, indent=2),
        encoding="utf-8",
    )
    write_readme(report_dir / "README.md", summary, info, prefix)

    reference_dir = report_dir / "reference"
    copied = {
        "final_review_summary": copy_if_exists(
            output_dir / "final_review_summary.json",
            reference_dir / "final_review_summary.json",
        ),
        "human_edited_workbook": copy_if_exists(
            output_dir / "human_reviewed_corpus.xlsx",
            reference_dir / "human_reviewed_corpus.xlsx",
        ),
        "frozen_workbook": copy_if_exists(
            output_dir / "final_human_reviewed_corpus.xlsx",
            reference_dir / "final_human_reviewed_corpus.xlsx",
        ),
    }
    summary["reference_files"] = copied
    (report_dir / f"{prefix}_summary.json").write_text(
        json.dumps(summary, indent=2),
        encoding="utf-8",
    )
    print(json.dumps(summary, indent=2))
    print(f"Report written: {report_dir.resolve()}")


if __name__ == "__main__":
    main()
