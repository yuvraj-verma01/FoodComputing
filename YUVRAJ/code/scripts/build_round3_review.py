"""Build the Round 3 human-review workbook.

One Excel with a blank `keep` column (mark 1/0) covering everything worth a look:
  - auto_relevant         : rule+LLM agreed relevant
  - auto_manual_review    : rule/LLM unsure, needs a human call
  - crawl_failed_review   : high-value by title but the site blocked crawling
                            (news18 / ndtv / business-standard etc.) -> read manually
A second sheet lists the auto-irrelevant rows for transparency.

Targets were already deduplicated against every URL you marked in Rounds 1-2
(metadata stage excluded 122 previously-reviewed URLs).
"""

from __future__ import annotations

import csv
import re
import sqlite3
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from crawler.config import Config

RUN_NAME = "edible_oil_adulteration_round_03_2026-06-23"
RUN_DIR  = ROOT / "data" / "runs" / RUN_NAME
CONFIG   = ROOT / "config" / "config_edible_oils_round3.yaml"
REL_DIR  = RUN_DIR / "mediacloud" / "outputs" / "oil_relevance"
META_CSV = REL_DIR / "metadata_all_articles_review.csv"
QUEUE_CSV = REL_DIR / "crawl_queue.csv"
ALL_CSV  = REL_DIR / "all_articles_review.csv"
OUT_XLSX = REL_DIR / "round_03_article_review.xlsx"

TIGHT_PHRASES = {
    "fake oil", "edible oil traders", "seized adulterated food items",
    "edible oil samples", "substandard edible oil", "collected edible oil samples",
    "unfit edible oil", "mislabelled oils",
}


def read_csv(p: Path) -> list[dict]:
    if not p.exists():
        return []
    with p.open(encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def extract_phrase(q: str) -> str:
    m = re.match(r'\("(.+?)" AND', q or "")
    return m.group(1) if m else ""


def db_urls_with_text(db_path: Path) -> set[str]:
    with sqlite3.connect(db_path) as con:
        rows = con.execute(
            "SELECT url FROM articles WHERE article_text IS NOT NULL AND article_text != ''"
        ).fetchall()
    return {r[0] for r in rows if r[0]}


def gather_targets(meta_rows: list[dict]) -> dict[str, dict]:
    queue = read_csv(QUEUE_CSV)
    targets: dict[str, dict] = {r["url"]: r for r in queue if r.get("url")}
    for r in meta_rows:
        if (
            r.get("query_family") == "phrase"
            and r.get("crawl_priority") == "drop"
            and extract_phrase(r.get("query_used", "")) in TIGHT_PHRASES
            and r.get("oil_role") != "non_food_oil"
        ):
            targets.setdefault(r["url"], r)
    return targets


# Off-topic by title -> still listed but pre-noted so the user can skip fast
OFF_TOPIC_HINTS = ("jennifer-aniston", "almond-oil", "climate-crisis", "fossil-fuel")


def looks_off_topic(url: str) -> bool:
    return any(h in url.lower() for h in OFF_TOPIC_HINTS)


def main() -> int:
    cfg = Config(CONFIG)
    db_path = cfg.path("db")
    meta_rows = read_csv(META_CSV)
    meta_by_url = {r["url"]: r for r in meta_rows}
    all_rows = read_csv(ALL_CSV)

    have_text = db_urls_with_text(db_path)
    targets = gather_targets(meta_rows)
    failed_urls = [u for u in targets if u not in have_text]

    relevant = [r for r in all_rows if r.get("final_label") == "relevant"]
    manual   = [r for r in all_rows if r.get("final_label") == "manual_review"]
    irrelev  = [r for r in all_rows if r.get("final_label") == "irrelevant"]

    def review_row(r: dict, tag: str) -> dict:
        return {
            "keep": "",
            "review_tag": tag,
            "title": r.get("title", ""),
            "url": r.get("url", ""),
            "date": r.get("date", ""),
            "source": r.get("source", ""),
            "found_by": r.get("query_id", "") or extract_phrase(meta_by_url.get(r.get("url",""), {}).get("query_used","")),
            "oil_role": r.get("oil_role", ""),
            "rule_conf": r.get("confidence", ""),
            "rule_reason": r.get("reason", ""),
            "llm_label": r.get("llm_label", ""),
            "llm_conf": r.get("llm_confidence", ""),
            "llm_reason": r.get("llm_reason", ""),
        }

    def failed_row(url: str) -> dict:
        meta = meta_by_url.get(url, {})
        ph = extract_phrase(meta.get("query_used", ""))
        tag = "crawl_failed_review"
        note = "LIKELY OFF-TOPIC" if looks_off_topic(url) else ""
        return {
            "keep": "",
            "review_tag": tag,
            "title": meta.get("title", "") or "(title not captured - open URL)",
            "url": url,
            "date": meta.get("date", ""),
            "source": meta.get("source", ""),
            "found_by": meta.get("query_id", "") or ph,
            "oil_role": "NOT CRAWLED",
            "rule_conf": "",
            "rule_reason": note or "Site blocked crawling - read the article manually.",
            "llm_label": "",
            "llm_conf": "",
            "llm_reason": "",
        }

    review = (
        [review_row(r, "auto_relevant") for r in relevant]
        + [review_row(r, "auto_manual_review") for r in manual]
        + [failed_row(u) for u in failed_urls]
    )

    cols = ["keep", "review_tag", "title", "url", "date", "source", "found_by",
            "oil_role", "rule_conf", "rule_reason", "llm_label", "llm_conf", "llm_reason"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Round 3 Review"
    _write_sheet(ws, cols, review)

    ws2 = wb.create_sheet("Auto Irrelevant")
    _write_sheet(ws2, cols, [review_row(r, "auto_irrelevant") for r in irrelev])

    ws3 = wb.create_sheet("Instructions")
    _instructions(ws3, len(relevant), len(manual), len(failed_urls), len(irrelev))

    wb.save(OUT_XLSX)

    print(f"Wrote: {OUT_XLSX}")
    print(f"  auto_relevant       : {len(relevant)}")
    print(f"  auto_manual_review  : {len(manual)}")
    print(f"  crawl_failed_review : {len(failed_urls)}  "
          f"(of which likely off-topic: {sum(looks_off_topic(u) for u in failed_urls)})")
    print(f"  auto_irrelevant     : {len(irrelev)}  (separate sheet)")
    print(f"  TOTAL to review     : {len(review)}")
    return 0


def _write_sheet(ws, cols: list[str], rows: list[dict]) -> None:
    header_fill = PatternFill("solid", fgColor="305496")
    header_font = Font(bold=True, color="FFFFFF")
    tag_fills = {
        "auto_relevant": PatternFill("solid", fgColor="C6EFCE"),
        "auto_manual_review": PatternFill("solid", fgColor="FFEB9C"),
        "crawl_failed_review": PatternFill("solid", fgColor="FFC7CE"),
        "auto_irrelevant": PatternFill("solid", fgColor="F2F2F2"),
    }
    ws.append(cols)
    for ci, _ in enumerate(cols, 1):
        c = ws.cell(row=1, column=ci)
        c.fill = header_fill
        c.font = header_font
    for r in rows:
        ws.append([r.get(c, "") for c in cols])
        row_idx = ws.max_row
        fill = tag_fills.get(r.get("review_tag", ""))
        if fill:
            ws.cell(row=row_idx, column=2).fill = fill
    widths = {"keep": 6, "review_tag": 20, "title": 60, "url": 70, "date": 12,
              "source": 18, "found_by": 26, "oil_role": 16, "rule_conf": 9,
              "rule_reason": 50, "llm_label": 11, "llm_conf": 9, "llm_reason": 50}
    for ci, col in enumerate(cols, 1):
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = widths.get(col, 15)
    ws.freeze_panes = "A2"


def _instructions(ws, n_rel: int, n_man: int, n_fail: int, n_irr: int) -> None:
    lines = [
        ("Round 3 Article Review", True),
        ("", False),
        ("Mark the `keep` column 1 (relevant) or 0 (not relevant) on the 'Round 3 Review' sheet.", False),
        ("", False),
        ("review_tag legend:", True),
        (f"  auto_relevant ({n_rel}) - GREEN - rule + Ollama LLM both said edible oil is the adulterated/seized product.", False),
        (f"  auto_manual_review ({n_man}) - YELLOW - classifier unsure; your call.", False),
        (f"  crawl_failed_review ({n_fail}) - RED - site blocked crawling (news18/ndtv/business-standard etc).", False),
        ("      Title looks relevant but we could not read the body. Open the URL and judge manually.", False),
        ("      Rows flagged 'LIKELY OFF-TOPIC' can usually be marked 0 quickly.", False),
        ("", False),
        (f"'Auto Irrelevant' sheet ({n_irr}) - classifier rejected these. Skim only if you want to double-check.", False),
        ("", False),
        ("All URLs here are NEW - already deduplicated against everything you marked in Rounds 1 and 2.", False),
    ]
    for i, (text, bold) in enumerate(lines, 1):
        c = ws.cell(row=i, column=1, value=text)
        if bold:
            c.font = Font(bold=True, size=12)
    ws.column_dimensions["A"].width = 110


if __name__ == "__main__":
    raise SystemExit(main())
