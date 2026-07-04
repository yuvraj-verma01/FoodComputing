from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from crawler.review_dedupe import load_reviewed_url_keys, split_new_review_rows

DEFAULT_OUTPUT_DIR = Path(
    "data/runs/edible_oil_adulteration_round_02_2026-06-23/"
    "mediacloud/outputs/oil_relevance"
)

REVIEW_COLUMNS = [
    "keep",
    "final_label",
    "confidence",
    "title",
    "source",
    "date",
    "url",
    "reason",
    "evidence_phrase",
    "llm_label",
    "llm_confidence",
    "llm_reason",
    "oil_role",
    "edible_oil_terms",
    "adulteration_action_terms",
    "negative_terms",
    "query_family",
    "query_id",
    "article_id",
    "file_path",
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
RELEVANT_FILL = PatternFill("solid", fgColor="E2F0D9")
MANUAL_FILL = PatternFill("solid", fgColor="FFF2CC")
IRRELEVANT_FILL = PatternFill("solid", fgColor="FCE4D6")
SUMMARY_FILL = PatternFill("solid", fgColor="D9EAF7")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Create a formatted Excel workbook for oil relevance review."
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=DEFAULT_OUTPUT_DIR,
        help="Directory containing all_articles_review.csv and filtering_summary.json.",
    )
    parser.add_argument(
        "--workbook",
        type=Path,
        default=None,
        help="Workbook path to write. Defaults to <output-dir>/oil_relevance_review.xlsx.",
    )
    parser.add_argument(
        "--include-reviewed-urls",
        action="store_true",
        help="Keep URLs already human-marked in previous rounds in the workbook.",
    )
    return parser.parse_args()


def clean_cell(value: Any) -> Any:
    if value is None:
        return ""
    if not isinstance(value, str):
        return value
    # Excel rejects ASCII control characters other than tab/newline/carriage return.
    return re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", value)


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def load_summary(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def add_table_sheet(wb: Workbook, title: str, rows: list[dict[str, str]]) -> None:
    ws = wb.create_sheet(title)
    ws.append(REVIEW_COLUMNS)
    for row in rows:
        ws.append([clean_cell(row.get(col, "")) if col != "keep" else "" for col in REVIEW_COLUMNS])

    style_sheet(ws)
    apply_label_fills(ws)
    add_url_hyperlinks(ws)


def style_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = {
        "A": 8,
        "B": 16,
        "C": 10,
        "D": 52,
        "E": 22,
        "F": 14,
        "G": 42,
        "H": 44,
        "I": 44,
        "J": 14,
        "K": 14,
        "L": 36,
        "M": 18,
        "N": 28,
        "O": 32,
        "P": 28,
        "Q": 18,
        "R": 18,
        "S": 12,
        "T": 40,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def apply_label_fills(ws) -> None:
    label_col = REVIEW_COLUMNS.index("final_label") + 1
    fills = {
        "relevant": RELEVANT_FILL,
        "manual_review": MANUAL_FILL,
        "irrelevant": IRRELEVANT_FILL,
    }
    for row_idx in range(2, ws.max_row + 1):
        label = ws.cell(row=row_idx, column=label_col).value
        fill = fills.get(label)
        if not fill:
            continue
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col_idx).fill = fill


def add_url_hyperlinks(ws) -> None:
    url_col = REVIEW_COLUMNS.index("url") + 1
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=url_col)
        if cell.value:
            cell.hyperlink = str(cell.value)
            cell.style = "Hyperlink"


def add_summary_sheet(
    wb: Workbook,
    summary: dict[str, Any],
    output_dir: Path,
    previously_reviewed_rows: int = 0,
) -> None:
    ws = wb.active
    ws.title = "summary"
    ws.append(["field", "value"])
    for key, value in summary.items():
        ws.append([key, json.dumps(value, ensure_ascii=False) if isinstance(value, dict) else value])
    ws.append(["output_dir", str(output_dir)])
    ws.append(["previously_reviewed_urls_omitted", previously_reviewed_rows])
    ws.append(["review_note", "Use keep=1 to keep, keep=0 to drop, blank if undecided."])

    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 90
    for cell in ws[1]:
        cell.fill = SUMMARY_FILL
        cell.font = Font(bold=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def main() -> None:
    args = parse_args()
    output_dir = args.output_dir
    workbook_path = args.workbook or output_dir / "oil_relevance_review.xlsx"

    all_rows = read_csv_rows(output_dir / "all_articles_review.csv")
    previously_reviewed_rows: list[dict[str, str]] = []
    if not args.include_reviewed_urls:
        all_rows, previously_reviewed_rows = split_new_review_rows(all_rows, load_reviewed_url_keys())
    summary = load_summary(output_dir / "filtering_summary.json")

    wb = Workbook()
    add_summary_sheet(wb, summary, output_dir, len(previously_reviewed_rows))
    add_table_sheet(wb, "all_extracted", all_rows)
    add_table_sheet(wb, "relevant", [row for row in all_rows if row.get("final_label") == "relevant"])
    add_table_sheet(
        wb,
        "manual_review",
        [row for row in all_rows if row.get("final_label") == "manual_review"],
    )
    add_table_sheet(wb, "irrelevant", [row for row in all_rows if row.get("final_label") == "irrelevant"])

    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(workbook_path)
    print(f"Workbook written: {workbook_path.resolve()}")


if __name__ == "__main__":
    main()
