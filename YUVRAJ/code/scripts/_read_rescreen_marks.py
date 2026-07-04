"""Read marked rows from rescreen_review.xlsx."""
import sys
sys.stdout.reconfigure(encoding="utf-8")
import openpyxl
from collections import Counter
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
wb = openpyxl.load_workbook(ROOT / "reports/rescreen/rescreen_review.xlsx", data_only=True)
ws = wb["All Scored"]
headers = [c.value for c in ws[1]]
print("Columns:", headers)
print("Total rows (excl header):", ws.max_row - 1)
print()

marked = []
for row in ws.iter_rows(min_row=2, values_only=True):
    d = dict(zip(headers, row))
    k = d.get("keep")
    if k in (0, 1, "0", "1", 0.0, 1.0):
        d["keep"] = int(float(str(k)))
        marked.append(d)

keeps = Counter(d["keep"] for d in marked)
print(f"Marked rows : {len(marked)}")
print(f"  keep=1    : {keeps[1]}")
print(f"  keep=0    : {keeps[0]}")
print()
print("Sample marked rows:")
for d in marked[:10]:
    title = str(d.get("title") or "")[:70]
    print(f"  keep={d['keep']}  prob={d['prob']}  {title}")
