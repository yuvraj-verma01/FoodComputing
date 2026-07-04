from __future__ import annotations

import argparse
import csv
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from crawler.review_dedupe import load_reviewed_url_keys, split_new_review_rows

DEFAULT_OUTPUT_DIR = Path(
    "data/runs/edible_oils_boolean_title_proximity_2026-06-22/"
    "mediacloud/outputs/oil_relevance"
)

OIL_RE = re.compile(
    r"\b(edible oil|cooking oil|mustard oil|coconut oil|palm oil|soyabean oil|soybean oil|refined oil|vegetable oil|oil)\b",
    re.I,
)
RISK_RE = re.compile(
    r"\b(adulterat|fake|spurious|unsafe|substandard|misbrand|failed|reheated|reused|non-certified|food safety|fssai|fda|raid|seiz|sample|crackdown)\b",
    re.I,
)
EXCLUDE_RE = re.compile(
    r"\b(price|prices|export|import|ban|futures|stock|stocks|market|inflation|theft|stolen|looted|fire|warehouse|godown|diesel|cheese|soda|brominated|china|indonesia|sri lanka|kuwait|trump)\b",
    re.I,
)

COLUMNS = [
    "safety_flag",
    "review_reason",
    "audit_decision",
    "audit_label",
    "audit_confidence",
    "title",
    "source",
    "url",
    "audit_reason",
    "audit_evidence",
    "oil_role",
    "original_reason",
    "word_count",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Create compact rejected-article safety review.")
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument(
        "--include-reviewed-urls",
        action="store_true",
        help="Keep URLs already human-marked in previous rounds in the safety review.",
    )
    return parser.parse_args()


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def title_high_risk(title: str) -> bool:
    return bool(OIL_RE.search(title) and RISK_RE.search(title) and not EXCLUDE_RE.search(title))


def safety_reason(row: dict[str, str]) -> tuple[str, str] | None:
    reasons = []
    if row.get("audit_decision") == "rescue_review":
        reasons.append("full-text audit flagged rescue_review")
    if title_high_risk(row.get("title", "")):
        reasons.append("title has oil + adulteration/enforcement signal")
    if not reasons:
        return None
    return "manual_review_before_reject", "; ".join(reasons)


def write_csv(path: Path, rows: list[dict[str, str]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=COLUMNS)
        writer.writeheader()
        writer.writerows(rows)


def write_workbook(path: Path, rows: list[dict[str, str]], previously_reviewed_rows: int = 0) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "safety_review"
    ws.append(COLUMNS)
    for row in rows:
        ws.append([row.get(col, "") for col in COLUMNS])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    row_fill = PatternFill("solid", fgColor="FFF2CC")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = {
        "A": 24,
        "B": 42,
        "C": 18,
        "D": 14,
        "E": 12,
        "F": 56,
        "G": 22,
        "H": 52,
        "I": 42,
        "J": 42,
        "K": 18,
        "L": 42,
        "M": 12,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.fill = row_fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    url_col = COLUMNS.index("url") + 1
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=url_col)
        if cell.value:
            cell.hyperlink = str(cell.value)
            cell.style = "Hyperlink"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    summary = wb.create_sheet("summary", 0)
    summary.append(["metric", "value"])
    summary.append(["manual_review_before_reject", len(rows)])
    summary.append(["previously_reviewed_urls_omitted", previously_reviewed_rows])
    summary.append(["note", "These were originally rejected but should be checked before final exclusion."])
    summary.column_dimensions["A"].width = 32
    summary.column_dimensions["B"].width = 80
    for cell in summary[1]:
        cell.font = Font(bold=True)

    wb.save(path)


def main() -> None:
    args = parse_args()
    output_dir = args.output_dir
    audit_rows = read_csv(output_dir / "rejected_fulltext_audit.csv")
    safety_rows = []
    for row in audit_rows:
        decision = safety_reason(row)
        if not decision:
            continue
        flag, reason = decision
        safety_rows.append(
            {
                "safety_flag": flag,
                "review_reason": reason,
                "audit_decision": row.get("audit_decision", ""),
                "audit_label": row.get("audit_label", ""),
                "audit_confidence": row.get("audit_confidence", ""),
                "title": row.get("title", ""),
                "source": row.get("source", ""),
                "url": row.get("url", ""),
                "audit_reason": row.get("audit_reason", ""),
                "audit_evidence": row.get("audit_evidence", ""),
                "oil_role": row.get("oil_role", ""),
                "original_reason": row.get("reason", ""),
                "word_count": row.get("word_count", ""),
            }
        )
    previously_reviewed_rows: list[dict[str, str]] = []
    if not args.include_reviewed_urls:
        safety_rows, previously_reviewed_rows = split_new_review_rows(safety_rows, load_reviewed_url_keys())
    write_csv(output_dir / "rejected_safety_review.csv", safety_rows)
    write_workbook(output_dir / "rejected_safety_review.xlsx", safety_rows, len(previously_reviewed_rows))
    print(f"Safety-review rows: {len(safety_rows)}")
    if previously_reviewed_rows:
        print(f"Previously reviewed URLs omitted: {len(previously_reviewed_rows)}")
    print(f"Workbook written: {(output_dir / 'rejected_safety_review.xlsx').resolve()}")


if __name__ == "__main__":
    main()
