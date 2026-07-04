"""Ghee Round 1 full-text crawl + LLM review pipeline.

This mirrors the edible-oil workflow:
1. rule/title metadata filtering on every discovered URL
2. crawl all high/medium candidate URLs
3. send every successfully crawled candidate article to local Ollama with full text
4. write a human 0/1 review workbook

The crawl and LLM stages are checkpointed. Re-running the same command resumes.
"""

from __future__ import annotations

import argparse
import csv
import json
import sqlite3
import sys
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent.parent
SCRIPT_DIR = Path(__file__).resolve().parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

from crawler.config import Config
from crawler.downloader import Downloader
from crawler.extractor import Extractor
from crawler.storage import Storage

import run_ghee_metadata_llm_review as meta


RUN_DIR = ROOT / "data" / "runs" / "ghee_adulteration_round_01_2026-06-30"
CONFIG = ROOT / "config" / "config_ghee_round1.yaml"
OUT_DIR = RUN_DIR / "mediacloud" / "outputs" / "ghee_relevance_fulltext"
CRAWL_LOG = OUT_DIR / "ghee_fulltext_crawl_log.csv"
LLM_CHECKPOINT = OUT_DIR / "ghee_fulltext_llm_results.jsonl"

REVIEW_COLUMNS = [
    "keep",
    "llm_label",
    "llm_score",
    "llm_confidence",
    "title",
    "url",
    "source",
    "date",
    "word_count",
    "llm_reason",
    "llm_evidence",
    "metadata_label",
    "metadata_score",
    "rule_reason",
    "query_family",
    "query_id",
    "query_used",
    "file_path",
]


def main() -> int:
    args = parse_args()
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    cfg = Config(CONFIG)
    configure_crawl(cfg, args)

    metadata_rows = build_metadata_candidates()
    candidates = [
        row
        for row in metadata_rows
        if row.get("review_priority") in {"high", "medium"} and row.get("llm_queue") == "1"
    ]
    write_csv(OUT_DIR / "metadata_candidates.csv", candidates, list(candidates[0].keys()) if candidates else [])
    write_json(
        OUT_DIR / "metadata_candidate_summary.json",
        {
            "created_at": utc_now(),
            "all_discovered_urls": len(metadata_rows),
            "candidate_urls": len(candidates),
            "metadata_label_counts": dict(Counter(row.get("metadata_label", "") for row in metadata_rows)),
            "candidate_priority_counts": dict(Counter(row.get("review_priority", "") for row in candidates)),
            "cross_corpus_oil_dedupe": "not applied",
        },
    )

    if args.stage in {"all", "crawl"}:
        crawl_candidates(cfg, candidates, retry_failed=args.retry_failed)

    if args.stage in {"all", "llm"}:
        llm_rows = run_fulltext_llm(cfg, candidates, args)
        write_outputs(llm_rows, candidates)

    print(f"Output dir: {OUT_DIR}")
    return 0


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--stage", choices=["all", "crawl", "llm"], default="all")
    parser.add_argument("--model", default="llama3.1:8b-instruct-q4_K_M")
    parser.add_argument("--timeout-seconds", type=int, default=240)
    parser.add_argument("--crawl-timeout", type=int, default=45)
    parser.add_argument("--crawl-delay", type=float, default=1.5)
    parser.add_argument("--crawl-max-retries", type=int, default=1)
    parser.add_argument("--limit", type=int, default=0, help="Limit candidate rows for testing.")
    parser.add_argument("--max-text-chars", type=int, default=20000)
    parser.add_argument("--ignore-robots", action="store_true")
    parser.add_argument("--use-playwright", action="store_true")
    parser.add_argument("--playwright-first", action="store_true")
    parser.add_argument("--retry-failed", action="store_true")
    parser.add_argument("--retry-llm-errors", action="store_true")
    return parser.parse_args()


def configure_crawl(cfg: Config, args: argparse.Namespace) -> None:
    crawl = cfg.raw.setdefault("crawl", {})
    crawl["respect_robots_txt"] = not args.ignore_robots
    crawl["use_playwright"] = bool(args.use_playwright)
    crawl["playwright_first"] = bool(args.playwright_first)
    crawl["timeout_seconds"] = int(args.crawl_timeout)
    crawl["delay_seconds"] = float(args.crawl_delay)
    crawl["max_retries"] = int(args.crawl_max_retries)


def build_metadata_candidates() -> list[dict[str, str]]:
    query_lookup = meta.load_query_plan(meta.QUERY_PLAN)
    records = meta.load_discovered(meta.DB_PATH)
    return meta.build_metadata_rows(records, query_lookup)


def crawl_candidates(cfg: Config, candidates: list[dict[str, str]], retry_failed: bool) -> None:
    storage = Storage(cfg)
    downloader = Downloader(cfg)
    extractor = Extractor(cfg)
    logged = load_crawl_log(CRAWL_LOG)
    fieldnames = [
        "url",
        "final_url",
        "download_status",
        "http_status",
        "download_error",
        "extraction_status",
        "extraction_method",
        "word_count",
        "raw_html_path",
        "cleaned_text_path",
        "title",
        "source",
        "date",
        "query_family",
        "query_id",
        "logged_at",
    ]
    CRAWL_LOG.parent.mkdir(parents=True, exist_ok=True)
    with CRAWL_LOG.open("a", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        if handle.tell() == 0:
            writer.writeheader()
        for index, row in enumerate(candidates, start=1):
            url = row["url"]
            existing = storage.get_article(url)
            if existing and existing.get("article_text"):
                continue
            if url in logged and not retry_failed:
                continue

            print(f"[crawl {index}/{len(candidates)}] {safe_console(row.get('title', '')[:110])}", flush=True)
            result = downloader.download(url)
            status = result.get("status") or "failed"
            storage.mark_discovered_status(url, status)

            final_url = result.get("url") or url
            extracted: dict[str, Any] = {}
            if status == "success":
                extracted = extractor.extract(final_url, result.get("raw_html") or "", result.get("raw_html_path"))
                article = {
                    **extracted,
                    "url": final_url,
                    "canonical_url": extracted.get("canonical_url") or final_url,
                    "source": row.get("source"),
                    "publication_date": extracted.get("publication_date") or row.get("date"),
                    "query_used": row.get("query_used"),
                    "discovery_method": "ghee_round1_fulltext_pipeline",
                    "raw_html_path": result.get("raw_html_path"),
                    "discovered_at": "",
                }
                if storage.save_article(article) is False:
                    storage.update_article(final_url, article)
                if extracted.get("article_text"):
                    storage.mark_discovered_status(url, "downloaded_ghee_fulltext")

            log_row = {
                "url": url,
                "final_url": final_url,
                "download_status": status,
                "http_status": result.get("http_status") or "",
                "download_error": result.get("error_message") or "",
                "extraction_status": extracted.get("extraction_status") or "",
                "extraction_method": extracted.get("extraction_method") or "",
                "word_count": extracted.get("word_count") or 0,
                "raw_html_path": result.get("raw_html_path") or "",
                "cleaned_text_path": extracted.get("cleaned_text_path") or "",
                "title": extracted.get("title") or row.get("title") or "",
                "source": row.get("source") or "",
                "date": extracted.get("publication_date") or row.get("date") or "",
                "query_family": row.get("query_family") or "",
                "query_id": row.get("query_id") or "",
                "logged_at": utc_now(),
            }
            writer.writerow(log_row)
            handle.flush()
            print(
                "  download={download} extract={extract} words={words}".format(
                    download=status,
                    extract=log_row["extraction_status"] or "-",
                    words=log_row["word_count"],
                ),
                flush=True,
            )
    downloader.close()
    storage.export_csv()
    storage.export_jsonl()
    storage.close()


def run_fulltext_llm(
    cfg: Config,
    candidates: list[dict[str, str]],
    args: argparse.Namespace,
) -> list[dict[str, Any]]:
    articles = load_articles_by_url(cfg.path("db"))
    crawl_log = load_crawl_log(CRAWL_LOG)
    metadata_by_url = {row["url"]: row for row in candidates}
    targets = []
    for row in candidates:
        article = articles.get(row["url"])
        if not article:
            crawl_row = crawl_log.get(row["url"], {})
            final_url = crawl_row.get("final_url") or ""
            if final_url:
                article = articles.get(final_url)
        if not article:
            # If final_url changed during download, match by discovered title/source fallback.
            article = find_article_for_candidate(articles, row)
        if article and article.get("article_text"):
            targets.append((row, article))
    if args.limit:
        targets = targets[: args.limit]

    done = load_llm_checkpoint(LLM_CHECKPOINT)
    LLM_CHECKPOINT.parent.mkdir(parents=True, exist_ok=True)
    with LLM_CHECKPOINT.open("a", encoding="utf-8") as handle:
        for index, (row, article) in enumerate(targets, start=1):
            key = article.get("url") or row["url"]
            if key in done and not should_retry_llm(done[key], args.retry_llm_errors):
                continue
            print(f"[LLM {index}/{len(targets)}] {safe_console(row.get('title', '')[:110])}", flush=True)
            try:
                llm = call_ollama(row, article, args.model, args.timeout_seconds, args.max_text_chars)
            except Exception as exc:
                llm = {
                    "llm_label": "unclear",
                    "llm_score": 50,
                    "llm_confidence": 0.0,
                    "llm_reason": f"LLM call failed: {exc}",
                    "llm_evidence": "",
                    "llm_raw": "",
                    "llm_error": "1",
                }
            payload = {
                "url": key,
                "discovered_url": row["url"],
                "title": article.get("title") or row.get("title") or "",
                "source": article.get("source") or row.get("source") or "",
                "date": article.get("publication_date") or row.get("date") or "",
                "word_count": article.get("word_count") or "",
                "file_path": article.get("cleaned_text_path") or "",
                "metadata_label": row.get("metadata_label") or "",
                "metadata_score": row.get("metadata_score") or "",
                "rule_reason": row.get("rule_reason") or "",
                "query_family": row.get("query_family") or "",
                "query_id": row.get("query_id") or "",
                "query_used": row.get("query_used") or "",
                "llm_model": args.model,
                "scored_at": utc_now(),
                **llm,
            }
            handle.write(json.dumps(payload, ensure_ascii=False) + "\n")
            handle.flush()
            done[key] = payload
    return list(done.values())


def call_ollama(
    metadata: dict[str, str],
    article: dict[str, Any],
    model: str,
    timeout: int,
    max_text_chars: int,
) -> dict[str, Any]:
    prompt = build_prompt(metadata, article, max_text_chars)
    response = requests.post(
        "http://127.0.0.1:11434/api/generate",
        json={
            "model": model,
            "prompt": prompt,
            "stream": False,
            "format": "json",
            "options": {"temperature": 0},
        },
        timeout=timeout,
    )
    response.raise_for_status()
    raw = response.json().get("response") or "{}"
    parsed = parse_json_object(raw)
    label = str(parsed.get("label") or "unclear").strip().lower()
    if label not in {"relevant", "irrelevant", "unclear"}:
        label = "unclear"
    return {
        "llm_label": label,
        "llm_score": clamp_int(parsed.get("score"), default=50),
        "llm_confidence": clamp_float(parsed.get("confidence"), default=0.0),
        "llm_reason": str(parsed.get("reason") or "")[:800],
        "llm_evidence": str(parsed.get("evidence_phrase") or parsed.get("evidence") or "")[:600],
        "llm_raw": raw[:3000],
        "llm_error": "0",
    }


def build_prompt(metadata: dict[str, str], article: dict[str, Any], max_text_chars: int) -> str:
    text = str(article.get("article_text") or "")
    if max_text_chars > 0 and len(text) > max_text_chars:
        text = text[:max_text_chars] + "\n\n[TRUNCATED_FOR_CONTEXT]"
    return f"""
You are classifying Indian news articles for a ghee adulteration research corpus.

Relevant ONLY if the article is about ghee itself being adulterated, fake, suspected adulterated, spurious, unsafe, seized, raided, failed in samples, part of a ghee racket, or investigated by food-safety/FSSAI/FDA enforcement in India.

Also relevant if the article is about ghee adulteration in Tirupati/Tirumala laddu/prasadam, because the target item is ghee.

Irrelevant if:
- it is only a recipe, health/purity tip, beauty, pooja/religious-use, price, or generic lifestyle article
- ghee is only casually mentioned and not the target of adulteration/enforcement
- the article is not about India
- the article is about another food item and not ghee as the adulterated product

Return only JSON:
{{
  "label": "relevant" | "irrelevant" | "unclear",
  "score": 0,
  "confidence": 0.0,
  "reason": "short reason",
  "evidence_phrase": "short phrase from the article supporting the label"
}}

Title: {article.get("title") or metadata.get("title") or ""}
Source: {article.get("source") or metadata.get("source") or ""}
Date: {article.get("publication_date") or metadata.get("date") or ""}
URL: {article.get("url") or metadata.get("url") or ""}
Metadata rule label: {metadata.get("metadata_label", "")}
Metadata rule reason: {metadata.get("rule_reason", "")}
Query family: {metadata.get("query_family", "")}
Query ID: {metadata.get("query_id", "")}
Query used: {metadata.get("query_used", "")}

Article text:
{text}
""".strip()


def write_outputs(llm_rows: list[dict[str, Any]], candidates: list[dict[str, str]]) -> None:
    rows = sorted(
        llm_rows,
        key=lambda row: (
            {"relevant": 0, "unclear": 1, "irrelevant": 2}.get(row.get("llm_label", ""), 9),
            -int(row.get("llm_score") or 0),
            row.get("title", ""),
        ),
    )
    write_csv(OUT_DIR / "ghee_fulltext_llm_scored_articles.csv", rows, output_columns(rows))
    write_review_workbook(OUT_DIR / "ghee_fulltext_llm_review.xlsx", rows)
    write_json(
        OUT_DIR / "ghee_fulltext_llm_summary.json",
        {
            "created_at": utc_now(),
            "candidate_urls": len(candidates),
            "llm_scored_articles": len(rows),
            "llm_label_counts": dict(Counter(row.get("llm_label", "") for row in rows)),
            "llm_checkpoint": str(LLM_CHECKPOINT),
            "crawl_log": str(CRAWL_LOG),
            "review_workbook": str(OUT_DIR / "ghee_fulltext_llm_review.xlsx"),
        },
    )


def write_review_workbook(path: Path, rows: list[dict[str, Any]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Review"
    write_sheet(ws, rows, REVIEW_COLUMNS)
    summary = wb.create_sheet("Summary")
    summary.append(["field", "value"])
    summary.append(["rows", len(rows)])
    summary.append(["label_counts", json.dumps(dict(Counter(row.get("llm_label", "") for row in rows)))])
    summary.append(["checkpoint", str(LLM_CHECKPOINT)])
    summary.append(["crawl_log", str(CRAWL_LOG)])
    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 100
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_sheet(ws, rows: list[dict[str, Any]], columns: list[str]) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    keep_fill = PatternFill("solid", fgColor="FFF2CC")
    relevant_fill = PatternFill("solid", fgColor="D9EAD3")
    unclear_fill = PatternFill("solid", fgColor="FFF2CC")
    irrelevant_fill = PatternFill("solid", fgColor="FCE4D6")
    for col_idx, col in enumerate(columns, start=1):
        cell = ws.cell(1, col_idx, col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[cell.column_letter].width = column_width(col)
    validation = DataValidation(type="list", formula1='"0,1"', allow_blank=True)
    ws.add_data_validation(validation)
    for row_idx, row in enumerate(rows, start=2):
        fill = {
            "relevant": relevant_fill,
            "unclear": unclear_fill,
            "irrelevant": irrelevant_fill,
        }.get(str(row.get("llm_label", "")))
        for col_idx, col in enumerate(columns, start=1):
            value = row.get(col, "")
            cell = ws.cell(row_idx, col_idx, value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if fill:
                cell.fill = fill
            if col == "keep":
                cell.fill = keep_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if col == "url" and value:
                cell.hyperlink = str(value)
                cell.style = "Hyperlink"
        ws.row_dimensions[row_idx].height = 66
    if rows:
        validation.add(f"A2:A{len(rows) + 1}")
    ws.freeze_panes = "F2"
    ws.auto_filter.ref = ws.dimensions


def load_crawl_log(path: Path) -> dict[str, dict[str, str]]:
    if not path.exists():
        return {}
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return {row.get("url", ""): row for row in csv.DictReader(handle) if row.get("url")}


def load_llm_checkpoint(path: Path) -> dict[str, dict[str, Any]]:
    if not path.exists():
        return {}
    rows = {}
    with path.open("r", encoding="utf-8") as handle:
        for line in handle:
            if not line.strip():
                continue
            try:
                row = json.loads(line)
            except json.JSONDecodeError:
                continue
            url = row.get("url", "")
            if url:
                rows[url] = row
    return rows


def load_articles_by_url(db_path: Path) -> dict[str, dict[str, Any]]:
    with sqlite3.connect(db_path) as con:
        con.row_factory = sqlite3.Row
        rows = con.execute("SELECT * FROM articles WHERE article_text IS NOT NULL").fetchall()
    return {dict(row).get("url", ""): dict(row) for row in rows}


def find_article_for_candidate(
    articles: dict[str, dict[str, Any]],
    candidate: dict[str, str],
) -> dict[str, Any] | None:
    title = (candidate.get("title") or "").strip().lower()
    if not title:
        return None
    for article in articles.values():
        if (article.get("title") or "").strip().lower() == title:
            return article
    return None


def should_retry_llm(row: dict[str, Any], retry_errors: bool) -> bool:
    return retry_errors and str(row.get("llm_error", "")) == "1"


def parse_json_object(raw: str) -> dict[str, Any]:
    try:
        parsed = json.loads(raw)
        return parsed if isinstance(parsed, dict) else {}
    except json.JSONDecodeError:
        import re

        match = re.search(r"\{.*\}", raw, flags=re.S)
        if not match:
            return {}
        try:
            parsed = json.loads(match.group(0))
            return parsed if isinstance(parsed, dict) else {}
        except json.JSONDecodeError:
            return {}


def clamp_int(value: Any, default: int) -> int:
    try:
        return max(0, min(100, int(round(float(value)))))
    except (TypeError, ValueError):
        return default


def clamp_float(value: Any, default: float) -> float:
    try:
        return max(0.0, min(1.0, float(value)))
    except (TypeError, ValueError):
        return default


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")


def output_columns(rows: list[dict[str, Any]]) -> list[str]:
    if not rows:
        return REVIEW_COLUMNS
    preferred = REVIEW_COLUMNS + ["discovered_url", "llm_model", "scored_at", "llm_error", "llm_raw"]
    extras = [key for key in rows[0] if key not in preferred]
    return preferred + extras


def column_width(col: str) -> int:
    return {
        "keep": 7,
        "llm_label": 13,
        "llm_score": 10,
        "llm_confidence": 13,
        "title": 70,
        "url": 58,
        "source": 22,
        "date": 13,
        "word_count": 11,
        "llm_reason": 55,
        "llm_evidence": 45,
        "metadata_label": 18,
        "metadata_score": 14,
        "rule_reason": 42,
        "query_family": 14,
        "query_id": 32,
        "query_used": 55,
        "file_path": 40,
    }.get(col, 18)


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def safe_console(text: str) -> str:
    return str(text).encode("ascii", errors="replace").decode("ascii")


if __name__ == "__main__":
    raise SystemExit(main())
