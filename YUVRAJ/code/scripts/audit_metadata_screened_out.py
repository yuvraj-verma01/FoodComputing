from __future__ import annotations

import argparse
import csv
import json
import re
import sqlite3
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from crawler.review_dedupe import load_reviewed_url_keys, split_new_review_rows


DEFAULT_OUTPUT_DIR = ROOT / Path(
    "data/runs/edible_oils_boolean_title_proximity_2026-06-22/"
    "mediacloud/outputs/oil_relevance"
)
DEFAULT_DB = ROOT / Path(
    "data/runs/edible_oils_boolean_title_proximity_2026-06-22/"
    "mediacloud/outputs/articles.db"
)

PRODUCT_TERMS = [
    "edible oil",
    "edible oils",
    "cooking oil",
    "cooking oils",
    "mustard oil",
    "palm oil",
    "soybean oil",
    "soyabean oil",
    "sunflower oil",
    "groundnut oil",
    "coconut oil",
    "rice bran oil",
    "cottonseed oil",
    "sesame oil",
    "vegetable oil",
    "refined oil",
    "loose oil",
    "loose edible oil",
    "rapeseed oil",
]

ADULTERATION_TERMS = [
    "adulterat",
    "fake",
    "spurious",
    "unsafe",
    "substandard",
    "misbrand",
    "contaminat",
    "unfit",
    "non-certified",
    "non certified",
    "failed",
    "quality test",
    "sample failed",
    "samples failed",
    "rancid",
    "reheated",
    "reused",
    "re-use",
    "reuse",
]

ENFORCEMENT_TERMS = [
    "fssai",
    " fda ",
    "fsda",
    "food safety",
    "food safety department",
    "food safety officer",
    "raid",
    "raids",
    "raided",
    "seize",
    "seized",
    "seizure",
    "sample",
    "samples",
    "lab test",
    "quality test",
    "penalty",
    "fine",
    "prosecution",
    "crackdown",
    "inspection",
    "shop sealed",
    "licence suspended",
    "license suspended",
]

INDIA_TERMS = [
    "india",
    "fssai",
    "fsda",
    "food safety department",
    "food safety officer",
    "andhra",
    "arunachal",
    "assam",
    "bihar",
    "chandigarh",
    "chhattisgarh",
    "delhi",
    "goa",
    "gujarat",
    "haryana",
    "himachal",
    "jharkhand",
    "karnataka",
    "kerala",
    "madhya pradesh",
    "maharashtra",
    "manipur",
    "meghalaya",
    "mizoram",
    "nagaland",
    "odisha",
    "punjab",
    "rajasthan",
    "sikkim",
    "tamil nadu",
    "telangana",
    "tripura",
    "uttar pradesh",
    "uttarakhand",
    "west bengal",
    "mumbai",
    "delhi",
    "kanpur",
    "lucknow",
    "ghaziabad",
    "noida",
    "jaipur",
    "indore",
    "bhopal",
    "jabalpur",
    "hyderabad",
    "coimbatore",
    "tiruchi",
    "tiruchirapalli",
    "kochi",
    "thiruvananthapuram",
    "ludhiana",
    "patna",
    "ranchi",
]

INDIAN_SOURCE_HINTS = [
    ".in",
    "timesofindia",
    "hindustantimes",
    "indianexpress",
    "indiatoday",
    "business-standard",
    "businesstoday",
    "financialexpress",
    "freepressjournal",
    "thehindu",
    "thehindubusinessline",
    "newindianexpress",
    "deccanchronicle",
    "livemint",
    "siasat",
    "sentinelassam",
    "tribuneindia",
    "zeenews.india",
    "india.com",
    "ndtv",
    "news18",
    "uniindia",
    "aninews",
    "thehansindia",
    "munsifdaily",
    "deshdoot",
]

HARD_EXCLUDE_TERMS = [
    "ghee",
    "vanaspati",
    "milk",
    "paneer",
    "khoya",
    "mawa",
    "cheese",
    "sweets",
    "diesel",
    "petrol",
    "fuel",
    "crude oil",
    "engine oil",
    "hair oil",
    "essential oil",
    "oil rig",
    "offshore oil",
    "ongc",
    "refinery",
    "soda",
    "brominated",
]

BUSINESS_EXCLUDE_TERMS = [
    "export",
    "exports",
    "import",
    "imports",
    "price",
    "prices",
    "costlier",
    "inflation",
    "futures",
    "derivative",
    "derivatives",
    "commodity",
    "commodities",
    "stock market",
    "stocks",
    "fmcg",
    "sebi",
    "dgft",
    "palm oil ban",
    "export ban",
    "import ban",
    "lift ban",
    "tariff",
]

INTERNATIONAL_EXCLUDE_TERMS = [
    "indonesia",
    "sri lanka",
    "malaysia",
    "kuwait",
    "china",
    "chinese",
    "donald trump",
    "us ",
    " usa",
    "california",
    "global",
    "world",
    "pakistan",
    "bangladesh",
]

AUDIT_COLUMNS = [
    "metadata_decision",
    "crawl_rescue",
    "review_reason",
    "score",
    "title",
    "source",
    "date",
    "url",
    "query_family",
    "query_id",
    "product_hits",
    "adulteration_hits",
    "enforcement_hits",
    "india_hits",
    "exclude_hits",
    "query_used",
]

QUEUE_COLUMNS = [
    "article_id",
    "title",
    "source",
    "date",
    "url",
    "query_family",
    "query_id",
    "query_used",
    "crawl_priority",
    "review_reason",
    "score",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Audit metadata-screened MediaCloud records and create a rescue crawl list."
    )
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--db", type=Path, default=DEFAULT_DB)
    parser.add_argument(
        "--include-reviewed-urls",
        action="store_true",
        help="Keep URLs already human-marked in previous rounds in the rescue crawl queue.",
    )
    return parser.parse_args()


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def write_csv(path: Path, rows: list[dict[str, Any]], columns: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns)
        writer.writeheader()
        for row in rows:
            writer.writerow({col: row.get(col, "") for col in columns})


def load_statuses(db_path: Path) -> dict[str, str]:
    with sqlite3.connect(db_path) as con:
        rows = con.execute("SELECT url, status FROM discovered_urls").fetchall()
    return {url: status for url, status in rows}


def hits(text: str, terms: list[str]) -> list[str]:
    found = []
    padded = f" {text.lower()} "
    for term in terms:
        if term in padded:
            found.append(term.strip())
    return found


def near_signal(text: str) -> bool:
    oil = r"(?:edible|cooking|mustard|palm|soy(?:a|bean)|sunflower|groundnut|coconut|rice bran|cottonseed|sesame|vegetable|refined|loose|rapeseed)\s+oil(?:s)?"
    bad = r"(?:adulterat\w*|fake|spurious|unsafe|substandard|misbrand\w*|contaminat\w*|unfit|failed|non[- ]certified|reheated|reused|seiz\w*|raid\w*|sample\w*|food safety|fssai|fda|fsda)"
    return bool(
        re.search(rf"\b{oil}\b.{{0,90}}\b{bad}\b", text, re.I)
        or re.search(rf"\b{bad}\b.{{0,90}}\b{oil}\b", text, re.I)
    )


def indian_source(source: str) -> bool:
    source_l = source.lower()
    return any(hint in source_l for hint in INDIAN_SOURCE_HINTS)


def audit_row(row: dict[str, str]) -> dict[str, Any]:
    title = row.get("title", "")
    url = row.get("url", "")
    source = row.get("source", "")
    text = f" {title} {url} {source} ".lower().replace("-", " ")

    product = hits(text, PRODUCT_TERMS)
    adulteration = hits(text, ADULTERATION_TERMS)
    enforcement = hits(text, ENFORCEMENT_TERMS)
    india = hits(text, INDIA_TERMS)
    hard_exclude = hits(text, HARD_EXCLUDE_TERMS)
    business_exclude = hits(text, BUSINESS_EXCLUDE_TERMS)
    international_exclude = hits(text, INTERNATIONAL_EXCLUDE_TERMS)
    excludes = hard_exclude + business_exclude + international_exclude

    source_india = indian_source(source)
    proximity = near_signal(text)

    score = 0
    reasons = []
    if product:
        score += 3
    if adulteration:
        score += 4
    if enforcement:
        score += 2
    if proximity:
        score += 4
        reasons.append("oil term appears near adulteration/enforcement term")
    if india or source_india:
        score += 2
    if hard_exclude:
        score -= 6
    if business_exclude:
        score -= 5
    if international_exclude:
        score -= 5

    has_oil = bool(product)
    has_adulteration_of_oil = bool(adulteration and proximity)
    has_enforcement_of_oil = bool(enforcement and proximity and adulteration)
    has_india = bool(india or source_india)
    clean_context = not hard_exclude and not business_exclude and not international_exclude

    if has_oil and has_india and clean_context and (has_adulteration_of_oil or has_enforcement_of_oil):
        decision = "strict_rescue_candidate"
        crawl_rescue = 1
        reasons.append("passes strict metadata test: edible oil + adulteration/enforcement + India signal")
    elif has_oil and has_india and not hard_exclude and proximity and not international_exclude:
        decision = "possible_review"
        crawl_rescue = 0
        reasons.append("borderline oil/enforcement signal but excluded from rescue crawl by conservative screen")
    else:
        decision = "confirm_metadata_drop"
        crawl_rescue = 0
        if not has_oil:
            reasons.append("no edible-oil product signal")
        if not (adulteration or enforcement):
            reasons.append("no adulteration/enforcement signal")
        if not has_india:
            reasons.append("no India/source signal")
        if excludes:
            reasons.append("excluded context: " + "; ".join(sorted(set(excludes))))

    return {
        **row,
        "metadata_decision": decision,
        "crawl_rescue": crawl_rescue,
        "review_reason": " | ".join(reasons),
        "score": score,
        "product_hits": "; ".join(sorted(set(product))),
        "adulteration_hits": "; ".join(sorted(set(adulteration))),
        "enforcement_hits": "; ".join(sorted(set(enforcement))),
        "india_hits": "; ".join(sorted(set(india + (["indian_source"] if source_india else [])))),
        "exclude_hits": "; ".join(sorted(set(excludes))),
    }


def add_table_sheet(wb: Workbook, title: str, rows: list[dict[str, Any]], columns: list[str]) -> None:
    ws = wb.create_sheet(title)
    ws.append(columns)
    for row in rows:
        ws.append([row.get(col, "") for col in columns])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    rescue_fill = PatternFill("solid", fgColor="E2F0D9")
    review_fill = PatternFill("solid", fgColor="FFF2CC")
    drop_fill = PatternFill("solid", fgColor="FCE4D6")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = {
        "A": 24,
        "B": 12,
        "C": 48,
        "D": 8,
        "E": 56,
        "F": 22,
        "G": 14,
        "H": 52,
        "I": 16,
        "J": 22,
        "K": 30,
        "L": 30,
        "M": 30,
        "N": 30,
        "O": 40,
        "P": 70,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    decision_col = columns.index("metadata_decision") + 1 if "metadata_decision" in columns else None
    url_col = columns.index("url") + 1 if "url" in columns else None
    for row_idx in range(2, ws.max_row + 1):
        fill = None
        if decision_col:
            decision = ws.cell(row=row_idx, column=decision_col).value
            if decision == "strict_rescue_candidate":
                fill = rescue_fill
            elif decision == "possible_review":
                fill = review_fill
            else:
                fill = drop_fill
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        if url_col:
            url_cell = ws.cell(row=row_idx, column=url_col)
            if url_cell.value:
                url_cell.hyperlink = str(url_cell.value)
                url_cell.style = "Hyperlink"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def write_workbook(path: Path, rows: list[dict[str, Any]], rescue_rows: list[dict[str, Any]]) -> None:
    wb = Workbook()
    summary = wb.active
    summary.title = "summary"
    counts: dict[str, int] = {}
    for row in rows:
        counts[row["metadata_decision"]] = counts.get(row["metadata_decision"], 0) + 1
    summary.append(["metric", "value"])
    summary.append(["metadata_screened_records_audited", len(rows)])
    summary.append(["strict_rescue_candidates", len(rescue_rows)])
    for key, value in sorted(counts.items()):
        summary.append([key, value])
    summary.append(["criteria", "edible oil + adulteration/enforcement near oil + India/source signal; excludes trade/export/price/international/non-food/dairy contexts"])
    summary.column_dimensions["A"].width = 36
    summary.column_dimensions["B"].width = 120
    for cell in summary[1]:
        cell.font = Font(bold=True)

    add_table_sheet(wb, "rescue_crawl", rescue_rows, AUDIT_COLUMNS)
    add_table_sheet(wb, "all_metadata_audit", rows, AUDIT_COLUMNS)
    wb.save(path)


def main() -> None:
    args = parse_args()
    output_dir = args.output_dir
    metadata_rows = read_csv(output_dir / "metadata_all_articles_review.csv")
    statuses = load_statuses(args.db)
    screened_rows = [row for row in metadata_rows if statuses.get(row.get("url", "")) == "pending"]
    audited = [audit_row(row) for row in screened_rows]
    rescue = [
        {**row, "crawl_priority": "high"}
        for row in audited
        if row["metadata_decision"] == "strict_rescue_candidate"
    ]
    previously_reviewed_rescue: list[dict[str, Any]] = []
    reviewed_url_key_count = 0
    if not args.include_reviewed_urls:
        reviewed_url_keys = load_reviewed_url_keys()
        reviewed_url_key_count = len(reviewed_url_keys)
        rescue, previously_reviewed_rescue = split_new_review_rows(rescue, reviewed_url_keys)
    rescue.sort(key=lambda row: (-int(row["score"]), row.get("date", ""), row.get("title", "")))
    audited.sort(key=lambda row: (row["metadata_decision"] != "strict_rescue_candidate", -int(row["score"]), row.get("title", "")))

    write_csv(output_dir / "metadata_screened_out_audit.csv", audited, AUDIT_COLUMNS)
    write_csv(output_dir / "metadata_rescue_crawl_queue.csv", rescue, QUEUE_COLUMNS)
    write_csv(
        output_dir / "metadata_rescue_previously_reviewed_urls.csv",
        previously_reviewed_rescue,
        QUEUE_COLUMNS,
    )
    write_workbook(output_dir / "metadata_screened_out_audit.xlsx", audited, rescue)

    summary = {
        "created_at": utc_now(),
        "metadata_screened_records_audited": len(audited),
        "strict_rescue_candidates": len(rescue),
        "previously_reviewed_rescue_candidates_excluded": len(previously_reviewed_rescue),
        "reviewed_url_key_count": reviewed_url_key_count,
        "include_reviewed_urls": args.include_reviewed_urls,
        "decision_counts": {
            decision: sum(1 for row in audited if row["metadata_decision"] == decision)
            for decision in sorted({row["metadata_decision"] for row in audited})
        },
        "criteria": [
            "about edible oil: product term in title/url/source",
            "about adulteration of edible oil: adulteration/enforcement term near oil product term",
            "from India: Indian source or India/location/agency signal",
            "excluded: trade/export/import/price/international/non-food/dairy contexts",
        ],
    }
    (output_dir / "metadata_screened_out_audit_summary.json").write_text(
        json.dumps(summary, indent=2),
        encoding="utf-8",
    )
    print(json.dumps(summary, indent=2))
    print(f"Workbook written: {(output_dir / 'metadata_screened_out_audit.xlsx').resolve()}")
    print(f"Rescue crawl queue written: {(output_dir / 'metadata_rescue_crawl_queue.csv').resolve()}")


if __name__ == "__main__":
    main()
