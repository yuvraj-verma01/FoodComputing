"""Extract positive keyword candidates + NOT-term candidates from all 23 unique
Round 2 articles (deduplicated against Round 1 URLs) and build two review Excels.

Outputs (all in reports/edible_oil_adulteration_round_02/round_03_keyword_review/):
  round_03_positive_keyword_review.xlsx   -- mark keep=1/0 for each positive candidate
  round_03_not_term_review.xlsx           -- mark keep=1/0 for each NOT-term candidate
  round_03_positive_keyword_candidates.csv
  round_03_not_term_candidates.csv
  round_03_extraction_summary.json
"""
from __future__ import annotations

import csv
import json
import re
import sys
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "scripts"))

from crawl_sample_keywords import (  # noqa: E402
    COMPOSITE_WEIGHTS,
    METHODS,
    extract_keyword_candidates,
)

# ── Paths ──────────────────────────────────────────────────────────────────────

MASTER_CSV = ROOT / "reports/master_corpus/master_all_articles.csv"
R1_KEYWORD_BANK = ROOT / "reports/edible_oil_adulteration_round_01/keyword_review/round_01_positive_keyword_bank.csv"
R2_RESCUE_CLEAN_CSV = ROOT / "reports/edible_oil_adulteration_round_02/round_03_keyword_review/round_03_from_round_02_rescue_keyword_candidates_clean.csv"
OUTPUT_DIR = ROOT / "reports/edible_oil_adulteration_round_02/round_03_keyword_review"


# ── Helpers ────────────────────────────────────────────────────────────────────

def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def read_csv(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()), extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def thin_border() -> Border:
    side = Side(style="thin", color="BFBFBF")
    return Border(left=side, right=side, top=side, bottom=side)


def normalize_kw(term: str) -> str:
    return re.sub(r"\s+", " ", (term or "").lower().strip())


# ── Load prior keywords (to deduplicate positive candidates) ──────────────────

def load_prior_keyword_norms() -> set[str]:
    """All Round 1 approved + Round 2 rescue candidates already extracted."""
    norms: set[str] = set()
    # Round 1 bank (all kept terms)
    for row in read_csv(R1_KEYWORD_BANK):
        term = row.get("term") or row.get("keyword_or_keyphrase") or ""
        if term:
            norms.add(normalize_kw(term))
    # Round 2 rescue keywords already presented for review
    for row in read_csv(R2_RESCUE_CLEAN_CSV):
        term = row.get("keyword_or_keyphrase") or row.get("term") or ""
        if term:
            norms.add(normalize_kw(term))
    return norms


def load_positive_kept_norms() -> set[str]:
    """Only Round 1 approved-kept terms (review_label=keep_1 or keep_core)."""
    norms: set[str] = set()
    for row in read_csv(R1_KEYWORD_BANK):
        term = row.get("term") or row.get("keyword_or_keyphrase") or ""
        if term:
            norms.add(normalize_kw(term))
    return norms


# ── Build docs list from master CSV rows ─────────────────────────────────────

def rows_to_docs(rows: list[dict[str, str]]) -> list[dict[str, Any]]:
    docs = []
    for i, row in enumerate(rows, start=1):
        text = (row.get("article_text") or "").strip()
        if not text:
            continue
        docs.append({
            "article_number": i,
            "article_id": row.get("article_id") or "",
            "title": row.get("title") or f"Article {i}",
            "url": row.get("url") or "",
            "domain": row.get("domain") or "",
            "source": row.get("source") or "",
            "publication_date": row.get("publication_date") or row.get("date") or "",
            "human_label": row.get("final_human_label") or "",
            "word_count": row.get("word_count") or "",
            "text": text,
        })
    return docs


# ── NOT term risk flag logic ───────────────────────────────────────────────────

def not_term_risk_flag(term: str, positive_kept_norms: set[str]) -> tuple[str, str]:
    """Return (risk_flag, positive_overlap_terms)."""
    norm = normalize_kw(term)
    # Exact match with a kept positive keyword
    if norm in positive_kept_norms:
        return "do_not_not_exact_positive_kept", term

    # Partial overlap (positive keyword is substring of this term or vice versa)
    overlap = []
    for pk in positive_kept_norms:
        if pk in norm or norm in pk:
            overlap.append(pk)
    if overlap:
        risk = "risky_overlap_with_positive_keyword"
        return risk, "; ".join(overlap[:3])

    # Core domain terms that appear in relevant articles
    domain_signals = [
        "oil", "adulterat", "spurious", "fake", "substandard", "misbranded",
        "fssai", "food safety", "seized", "raid", "contaminated", "sample",
        "mustard", "edible", "cooking", "refin", "palmolein",
    ]
    if any(sig in norm for sig in domain_signals):
        return "risky_domain_term_may_drop_relevant", ""

    return "possible_not_candidate", ""


def irrelevant_category(term: str) -> str:
    p = normalize_kw(term)
    if any(t in p for t in ["export", "import", "price", "cost", "rate", "market", "crore", "lakh", "tonne", "billion", "million"]):
        return "price/trade/market"
    if any(t in p for t in ["pakistan", "indonesia", "ukraine", "russia", "bangladesh", "china", "global", "international", "world"]):
        return "international/geography"
    if any(t in p for t in ["recipe", "cook", "kitchen", "diet", "health benefit", "nutrition", "calorie", "cholesterol"]):
        return "culinary/health_education"
    if any(t in p for t in ["stock", "share", "sensex", "bse", "nse", "invest"]):
        return "financial"
    if any(t in p for t in ["election", "minister", "government policy", "parliament", "bjp", "congress"]):
        return "politics/policy"
    return "other_junk"


# ── Positive keyword Excel ─────────────────────────────────────────────────────

POS_COLUMNS = [
    ("keep",                    "keep",                   9),
    ("keyword_or_keyphrase",    "keyword_or_keyphrase",   30),
    ("review_label",            "suggested_label",        16),
    ("composite_score",         "composite_score",        15),
    ("category",                "category",               22),
    ("methods_found",           "methods_found",          28),
    ("method_count",            "method_count",           13),
    ("total_frequency",         "phrase_freq",            13),
    ("document_frequency",      "doc_freq",               11),
    ("review_reason",           "review_reason",          40),
    ("example_article_title",   "example_title",          50),
    ("example_context",         "example_context",        75),
    ("example_url",             "example_url",            55),
]


def build_positive_excel(rows: list[dict[str, Any]], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Keyword Review"

    thin = thin_border()
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    keep_fill = PatternFill("solid", fgColor="FFF2CC")

    ws.row_dimensions[1].height = 32
    for ci, (_key, header, width) in enumerate(POS_COLUMNS, start=1):
        cell = ws.cell(row=1, column=ci, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = width

    ws["A1"].comment = Comment("1 = add to Round 3 query set; 0 = drop", "System")
    dv = DataValidation(type="list", formula1='"0,1"', allow_blank=True,
                        showErrorMessage=True, errorTitle="Use 0 or 1",
                        error="Enter 1 to keep or 0 to drop.")
    ws.add_data_validation(dv)

    label_fills = {
        "keep_core": PatternFill("solid", fgColor="D9EAD3"),
        "manual_review": PatternFill("solid", fgColor="FFF2CC"),
        "drop": PatternFill("solid", fgColor="E7E6E6"),
    }

    for ri, row in enumerate(rows, start=2):
        lbl = row.get("review_label", "")
        row_fill = label_fills.get(lbl, PatternFill("solid", fgColor="FFFFFF"))
        ws.row_dimensions[ri].height = 60
        for ci, (key, _header, _width) in enumerate(POS_COLUMNS, start=1):
            value = "" if key == "keep" else row.get(key, "")
            cell = ws.cell(row=ri, column=ci, value=value)
            cell.border = thin
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.font = Font(size=10)
            cell.fill = keep_fill if key == "keep" else row_fill

    if len(rows) >= 1:
        dv.add(f"A2:A{len(rows)+1}")
        ws.conditional_formatting.add(f"A2:A{len(rows)+1}",
            CellIsRule("equal", ["1"], fill=PatternFill("solid", fgColor="C6EFCE")))
        ws.conditional_formatting.add(f"A2:A{len(rows)+1}",
            CellIsRule("equal", ["0"], fill=PatternFill("solid", fgColor="F4CCCC")))

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions

    # Summary sheet
    ss = wb.create_sheet("Summary")
    ss["A1"] = "Round 3 Positive Keyword Candidates — Round 2 Unique Relevant Articles"
    ss["A1"].font = Font(bold=True, size=13)
    ss["A3"] = "Source articles"
    ss["B3"] = "23 unique Round 2 articles (18 relevant with text)"
    ss["A4"] = "Total candidates"
    ss["B4"] = len(rows)
    ss["A5"] = "Prior-keyword filter"
    ss["B5"] = "Candidates already in Round 1 bank or Round 2 rescue review are excluded"
    ss["A6"] = "Instructions"
    ss["B6"] = "Mark keep=1 to include in Round 3 query set. keep=0 to drop. Blank = not reviewed."
    label_ct = Counter(r.get("review_label","") for r in rows)
    ss["A8"] = "Suggested label counts"
    ss["A8"].font = Font(bold=True)
    for i, (lbl, cnt) in enumerate(sorted(label_ct.items()), start=9):
        ss.cell(i, 1, lbl); ss.cell(i, 2, cnt)
    ss.column_dimensions["A"].width = 28
    ss.column_dimensions["B"].width = 80

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


# ── NOT term Excel ─────────────────────────────────────────────────────────────

NOT_COLUMNS = [
    ("keep",                        "keep_as_NOT",          12),
    ("not_candidate",               "candidate_not_term",   30),
    ("risk_flag",                   "risk_flag",            34),
    ("irrelevant_category",         "category",             22),
    ("irrelevant_composite_score",  "score",                12),
    ("irrelevant_document_frequency","doc_freq",            10),
    ("irrelevant_total_frequency",  "phrase_freq",          12),
    ("positive_exact_kept",         "exact_positive_match", 18),
    ("positive_overlap_terms",      "positive_overlap",     38),
    ("example_irrelevant_title",    "example_title",        55),
    ("example_context",             "example_context",      75),
]

RISK_FILLS = {
    "possible_not_candidate":               PatternFill("solid", fgColor="E2F0D9"),
    "risky_domain_term_may_drop_relevant":  PatternFill("solid", fgColor="FFF2CC"),
    "risky_overlap_with_positive_keyword":  PatternFill("solid", fgColor="FCE4D6"),
    "do_not_not_exact_positive_kept":       PatternFill("solid", fgColor="F4CCCC"),
}


def build_not_term_excel(rows: list[dict[str, Any]], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "NOT Term Review"

    thin = thin_border()
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    keep_fill = PatternFill("solid", fgColor="FFF2CC")

    ws.row_dimensions[1].height = 32
    for ci, (_key, header, width) in enumerate(NOT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=ci, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = width

    ws["A1"].comment = Comment(
        "1 = use as NOT/exclusion term in Round 3 queries\n"
        "0 = do NOT exclude (could appear in relevant articles)\n"
        "CONSERVATIVE: missing NOT term < losing relevant articles",
        "System",
    )
    dv = DataValidation(type="list", formula1='"0,1"', allow_blank=True,
                        showErrorMessage=True, errorTitle="Use 0 or 1",
                        error="Enter 1 to use as NOT term or 0 to reject.")
    ws.add_data_validation(dv)

    for ri, row in enumerate(rows, start=2):
        risk = row.get("risk_flag", "")
        row_fill = RISK_FILLS.get(risk, PatternFill("solid", fgColor="FFFFFF"))
        ws.row_dimensions[ri].height = 60
        for ci, (key, _header, _width) in enumerate(NOT_COLUMNS, start=1):
            value = "" if key == "keep" else row.get(key, "")
            cell = ws.cell(row=ri, column=ci, value=value)
            cell.border = thin
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.font = Font(size=10)
            cell.fill = keep_fill if key == "keep" else row_fill

    if len(rows) >= 1:
        dv.add(f"A2:A{len(rows)+1}")
        ws.conditional_formatting.add(f"A2:A{len(rows)+1}",
            CellIsRule("equal", ["1"], fill=PatternFill("solid", fgColor="C6EFCE")))
        ws.conditional_formatting.add(f"A2:A{len(rows)+1}",
            CellIsRule("equal", ["0"], fill=PatternFill("solid", fgColor="F4CCCC")))

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions

    # Instructions sheet
    inst = wb.create_sheet("Instructions")
    inst["A1"] = "NOT Term Review — Instructions"
    inst["A1"].font = Font(bold=True, size=13)
    lines = [
        ("Goal", "Pick terms that safely EXCLUDE junk from Round 3 queries."),
        ("keep=1", "Term clearly signals out-of-scope content: commodity prices, export bans, international trade, culinary tips, financial/stock news."),
        ("keep=0", "Anything that could appear in a genuine edible-oil adulteration article. When in doubt, reject."),
        ("Risk flag — possible_not_candidate (green)", "Safer to use. Unlikely to appear in relevant articles."),
        ("Risk flag — risky_domain (yellow)", "Contains oil/safety signal — be careful."),
        ("Risk flag — risky_overlap (orange)", "Partially overlaps with a known positive keyword. Very risky to NOT."),
        ("Risk flag — do_not_not_exact (red)", "Exact match with approved positive keyword. Never use as NOT term."),
        ("Rule of thumb", "A missing NOT term is better than excluding a relevant article."),
    ]
    for i, (lbl, text) in enumerate(lines, start=3):
        inst.cell(i, 1, lbl); inst.cell(i, 2, text)
        inst.cell(i, 1).font = Font(bold=True)
        inst.cell(i, 2).alignment = Alignment(vertical="top", wrap_text=True)
    inst.column_dimensions["A"].width = 38
    inst.column_dimensions["B"].width = 90

    # Summary sheet
    ss = wb.create_sheet("Summary")
    ss["A1"] = "NOT Term Candidates — Round 2 Unique Irrelevant Articles"
    ss["A1"].font = Font(bold=True, size=13)
    ss["A3"] = "Source articles"; ss["B3"] = "5 unique irrelevant Round 2 articles with text"
    ss["A4"] = "Total candidates"; ss["B4"] = len(rows)
    risk_ct = Counter(r.get("risk_flag","") for r in rows)
    ss["A6"] = "Risk flag counts"; ss["A6"].font = Font(bold=True)
    for i, (lbl, cnt) in enumerate(sorted(risk_ct.items()), start=7):
        ss.cell(i, 1, lbl); ss.cell(i, 2, cnt)
    ss.column_dimensions["A"].width = 42
    ss.column_dimensions["B"].width = 60

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


# ── Main ───────────────────────────────────────────────────────────────────────

def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    print("Loading master corpus...", flush=True)
    master = read_csv(MASTER_CSV)

    # Unique Round 2 articles only
    r2_unique = [r for r in master if r.get("round_number") == "2"]
    relevant_rows = [r for r in r2_unique if r.get("final_human_label") == "relevant"]
    irrelevant_rows = [r for r in r2_unique if r.get("final_human_label") == "irrelevant"]
    print(f"  Unique Round 2: {len(r2_unique)} total | {len(relevant_rows)} relevant | {len(irrelevant_rows)} irrelevant")

    # ── POSITIVE KEYWORDS ──────────────────────────────────────────────────────
    print("\nExtracting positive keyword candidates from relevant articles...", flush=True)
    prior_norms = load_prior_keyword_norms()
    print(f"  Prior keywords to filter: {len(prior_norms)}")

    rel_docs = rows_to_docs(relevant_rows)
    print(f"  Relevant articles with text: {len(rel_docs)}")
    raw_candidates = extract_keyword_candidates(rel_docs)
    print(f"  Raw candidates before dedup: {len(raw_candidates)}")

    # Filter out prior keywords
    new_candidates = []
    suppressed = []
    for row in raw_candidates:
        norm = normalize_kw(row.get("keyword_or_keyphrase", ""))
        if norm in prior_norms:
            suppressed.append({**row, "suppression_reason": "already_in_prior_keyword_bank"})
        else:
            # Add round metadata
            row["source_round_number"] = "2"
            row["next_round_target"] = "3"
            row["keyword_source_scope"] = "unique_round_02_relevant_articles_only"
            row["keyword_source_article_count"] = len(rel_docs)
            new_candidates.append(row)

    print(f"  New candidates (after prior dedup): {len(new_candidates)}")
    print(f"  Suppressed (already known): {len(suppressed)}")

    # ── NOT TERMS ──────────────────────────────────────────────────────────────
    print("\nExtracting NOT-term candidates from irrelevant articles...", flush=True)
    positive_kept = load_positive_kept_norms()
    irr_docs = rows_to_docs(irrelevant_rows)
    print(f"  Irrelevant articles with text: {len(irr_docs)}")

    not_candidates_raw: list[dict[str, Any]] = []
    if irr_docs:
        raw_not = extract_keyword_candidates(irr_docs)
        for row in raw_not:
            term = row.get("keyword_or_keyphrase", "")
            if not term:
                continue
            risk, overlap = not_term_risk_flag(term, positive_kept)
            example_doc = next(
                (d for d in irr_docs if re.search(re.escape(term), d["text"], re.I)), {}
            )
            not_candidates_raw.append({
                "not_candidate": term,
                "risk_flag": risk,
                "irrelevant_category": irrelevant_category(term),
                "irrelevant_review_label": row.get("review_label", ""),
                "irrelevant_composite_score": row.get("composite_score", ""),
                "irrelevant_document_frequency": row.get("document_frequency", ""),
                "irrelevant_total_frequency": row.get("total_frequency", ""),
                "positive_exact_kept": "1" if risk == "do_not_not_exact_positive_kept" else "0",
                "positive_overlap_terms": overlap,
                "example_irrelevant_title": example_doc.get("title", ""),
                "example_context": row.get("example_context", ""),
            })
        # Sort: safest first
        risk_order = {
            "possible_not_candidate": 0,
            "risky_domain_term_may_drop_relevant": 1,
            "risky_overlap_with_positive_keyword": 2,
            "do_not_not_exact_positive_kept": 3,
        }
        not_candidates_raw.sort(key=lambda r: (
            risk_order.get(r["risk_flag"], 4),
            -float(r["irrelevant_composite_score"] or 0),
        ))
    print(f"  NOT-term candidates: {len(not_candidates_raw)}")

    # ── Write CSVs ─────────────────────────────────────────────────────────────
    pos_csv = OUTPUT_DIR / "round_03_positive_keyword_candidates.csv"
    not_csv = OUTPUT_DIR / "round_03_not_term_candidates.csv"
    sup_csv = OUTPUT_DIR / "round_03_positive_suppressed_prior_keywords.csv"
    write_csv(pos_csv, new_candidates)
    write_csv(not_csv, not_candidates_raw)
    write_csv(sup_csv, suppressed)
    print(f"\nCSVs written.")

    # ── Write Excels ───────────────────────────────────────────────────────────
    pos_xlsx = OUTPUT_DIR / "round_03_positive_keyword_review.xlsx"
    not_xlsx = OUTPUT_DIR / "round_03_not_term_review.xlsx"
    build_positive_excel(new_candidates, pos_xlsx)
    build_not_term_excel(not_candidates_raw, not_xlsx)
    print(f"Positive keyword Excel: {pos_xlsx}")
    print(f"NOT term Excel:         {not_xlsx}")

    # ── Summary ────────────────────────────────────────────────────────────────
    label_counts = Counter(r.get("review_label","") for r in new_candidates)
    risk_counts  = Counter(r.get("risk_flag","") for r in not_candidates_raw)
    summary = {
        "created_at": utc_now(),
        "source": "master_all_articles.csv filtered to round_number=2",
        "unique_r2_articles": len(r2_unique),
        "relevant_with_text": len(rel_docs),
        "irrelevant_with_text": len(irr_docs),
        "prior_keywords_filtered": len(prior_norms),
        "positive_raw_candidates": len(raw_candidates),
        "positive_new_candidates": len(new_candidates),
        "positive_suppressed": len(suppressed),
        "positive_label_counts": dict(sorted(label_counts.items())),
        "not_term_candidates": len(not_candidates_raw),
        "not_term_risk_counts": dict(sorted(risk_counts.items())),
        "outputs": {
            "positive_keyword_csv": str(pos_csv),
            "positive_keyword_xlsx": str(pos_xlsx),
            "not_term_csv": str(not_csv),
            "not_term_xlsx": str(not_xlsx),
        },
    }
    (OUTPUT_DIR / "round_03_extraction_summary.json").write_text(
        json.dumps(summary, indent=2, ensure_ascii=False), encoding="utf-8"
    )
    print("\n=== SUMMARY ===")
    print(json.dumps(summary, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
