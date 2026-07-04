from __future__ import annotations

import argparse
import csv
import json
import sqlite3
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from crawler.oil_relevance import ollama_relevance_check  # noqa: E402


DEFAULT_OUTPUT_DIR = ROOT / Path(
    "data/runs/edible_oils_boolean_title_proximity_2026-06-22/"
    "mediacloud/outputs/oil_relevance"
)
DEFAULT_DB = ROOT / Path(
    "data/runs/edible_oils_boolean_title_proximity_2026-06-22/"
    "mediacloud/outputs/articles.db"
)

AUDIT_COLUMNS = [
    "audit_decision",
    "audit_label",
    "audit_confidence",
    "audit_reason",
    "audit_evidence",
    "final_label",
    "title",
    "source",
    "date",
    "url",
    "oil_role",
    "reason",
    "evidence_phrase",
    "llm_label",
    "llm_reason",
    "edible_oil_terms",
    "adulteration_action_terms",
    "negative_terms",
    "word_count",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Second-pass full-text audit of rejected edible-oil articles."
    )
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--db", type=Path, default=DEFAULT_DB)
    parser.add_argument("--model", default="llama3.1:8b-instruct-q4_K_M")
    parser.add_argument("--timeout", type=int, default=180)
    parser.add_argument("--limit", type=int, default=0)
    return parser.parse_args()


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def load_article_texts(db_path: Path) -> dict[str, dict[str, Any]]:
    with sqlite3.connect(db_path) as con:
        con.row_factory = sqlite3.Row
        rows = con.execute(
            """
            SELECT url, title, article_text, word_count
            FROM articles
            WHERE article_text IS NOT NULL
            """
        ).fetchall()
    return {row["url"]: dict(row) for row in rows}


def read_existing(path: Path) -> dict[str, dict[str, Any]]:
    if not path.exists():
        return {}
    rows = {}
    for line in path.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        row = json.loads(line)
        rows[row["url"]] = row
    return rows


def audit_decision(label: str, confidence: float, row: dict[str, str]) -> str:
    if label == "relevant":
        return "rescue_review"
    if label == "unclear":
        return "rescue_review"
    if row.get("oil_role") == "adulterated_product" and confidence < 0.95:
        return "rescue_review"
    return "confirm_reject"


def write_jsonl(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        for row in rows:
            handle.write(json.dumps(row, ensure_ascii=False) + "\n")


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=AUDIT_COLUMNS)
        writer.writeheader()
        for row in rows:
            writer.writerow({col: row.get(col, "") for col in AUDIT_COLUMNS})


def write_workbook(path: Path, rows: list[dict[str, Any]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "rejected_audit"
    ws.append(AUDIT_COLUMNS)
    for row in rows:
        ws.append([row.get(col, "") for col in AUDIT_COLUMNS])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    rescue_fill = PatternFill("solid", fgColor="FFF2CC")
    reject_fill = PatternFill("solid", fgColor="E2F0D9")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = {
        "A": 18,
        "B": 14,
        "C": 12,
        "D": 42,
        "E": 42,
        "F": 14,
        "G": 52,
        "H": 22,
        "I": 14,
        "J": 48,
        "K": 18,
        "L": 42,
        "M": 42,
        "N": 14,
        "O": 42,
        "P": 28,
        "Q": 32,
        "R": 28,
        "S": 12,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    decision_col = AUDIT_COLUMNS.index("audit_decision") + 1
    url_col = AUDIT_COLUMNS.index("url") + 1
    for row_idx in range(2, ws.max_row + 1):
        decision = ws.cell(row=row_idx, column=decision_col).value
        fill = rescue_fill if decision == "rescue_review" else reject_fill
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        url_cell = ws.cell(row=row_idx, column=url_col)
        if url_cell.value:
            url_cell.hyperlink = str(url_cell.value)
            url_cell.style = "Hyperlink"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    summary = wb.create_sheet("summary", 0)
    counts: dict[str, int] = {}
    for row in rows:
        counts[row["audit_decision"]] = counts.get(row["audit_decision"], 0) + 1
    summary.append(["metric", "value"])
    summary.append(["rejected_articles_audited", len(rows)])
    for key, value in sorted(counts.items()):
        summary.append([key, value])
    summary.append(["note", "rescue_review means do not silently reject; inspect together."])
    summary.column_dimensions["A"].width = 32
    summary.column_dimensions["B"].width = 80
    for cell in summary[1]:
        cell.font = Font(bold=True)

    wb.save(path)


def main() -> None:
    args = parse_args()
    output_dir = args.output_dir
    review_rows = [
        row
        for row in read_csv(output_dir / "all_articles_review.csv")
        if row.get("final_label") == "irrelevant"
    ]
    if args.limit:
        review_rows = review_rows[: args.limit]

    articles = load_article_texts(args.db)
    jsonl_path = output_dir / "rejected_fulltext_audit.jsonl"
    existing = read_existing(jsonl_path)
    results = list(existing.values())
    done_urls = set(existing)

    for index, row in enumerate(review_rows, start=1):
        url = row["url"]
        if url in done_urls:
            continue
        article = articles.get(url, {})
        text = article.get("article_text") or ""
        title = article.get("title") or row.get("title") or ""
        print(f"[audit {index}/{len(review_rows)}] {url}", flush=True)
        try:
            llm = ollama_relevance_check(
                title=title,
                text=text,
                url=url,
                model=args.model,
                timeout_seconds=args.timeout,
            )
        except Exception as exc:
            llm = {
                "llm_label": "unclear",
                "llm_confidence": 0.0,
                "llm_reason": f"audit error: {exc}",
                "evidence_phrase": "",
                "llm_model": args.model,
            }
        label = str(llm.get("llm_label") or "unclear")
        confidence = float(llm.get("llm_confidence") or 0.0)
        result = {
            **row,
            "audit_decision": audit_decision(label, confidence, row),
            "audit_label": label,
            "audit_confidence": confidence,
            "audit_reason": llm.get("llm_reason", ""),
            "audit_evidence": llm.get("evidence_phrase", ""),
            "word_count": article.get("word_count") or "",
            "audit_model": llm.get("llm_model", args.model),
        }
        results.append(result)
        write_jsonl(jsonl_path, results)

    order = {row["url"]: idx for idx, row in enumerate(review_rows)}
    results = [row for row in results if row.get("url") in order]
    results.sort(key=lambda row: order[row["url"]])

    write_jsonl(jsonl_path, results)
    write_csv(output_dir / "rejected_fulltext_audit.csv", results)
    write_workbook(output_dir / "rejected_fulltext_audit.xlsx", results)
    print(f"Audited rejected articles: {len(results)}")
    print(f"Workbook written: {(output_dir / 'rejected_fulltext_audit.xlsx').resolve()}")


if __name__ == "__main__":
    main()
