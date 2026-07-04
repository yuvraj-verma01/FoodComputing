from __future__ import annotations

import argparse
import csv
import json
from collections import OrderedDict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent.parent
DEFAULT_KEYWORD_DIR = (
    ROOT / "reports" / "edible_oil_adulteration_round_01" / "keyword_review"
)
DEFAULT_WORKBOOK = DEFAULT_KEYWORD_DIR / "round_01_keyword_review.xlsx"

REMOVE_TERMS = {
    "samples of the seized",
    "sample of the oil",
    "oil used in india",
    "edible oil worth",
    "adulterated edible",
    "adulterated coconut",
    "adulterated cooking",
    "cooking oils",
    "used cooking oil",
}

COMBINE_ONLY_TERMS = {
    "oil",
    "food safety",
    "seized",
    "raid",
    "raids",
    "raided",
    "fda",
    "fssai",
    "fsda",
    "food safety officer",
    "food safety officers",
    "food safety officials",
    "food safety department",
    "food safety standards",
}

MANUAL_ADDITIONS = {
    "counterfeit": ("fraud/adulteration", "manual_addition", "Approved by user for Round 2 recall."),
    "unsafe": ("fraud/adulteration", "manual_addition", "Approved by user for Round 2 recall."),
    "seizure": ("enforcement/evidence", "manual_addition", "Noun variant of seized approved by user."),
    "crackdown": ("enforcement/evidence", "manual_addition", "Approved by user for enforcement-style queries."),
    "argemone oil": ("adulterant/mixing", "manual_addition", "Known adulterant signal found in relevant articles."),
    "mineral oil": ("adulterant/mixing", "manual_addition", "Known adulterant signal found in relevant articles."),
    "palmolein": ("oil/product", "manual_addition", "Oil/product term approved by user."),
    "cottonseed oil": ("oil/product", "manual_addition", "Oil/product term approved by user."),
    "soyabean oil": ("oil/product", "manual_addition", "Spelling variant approved by user."),
    "loose oil": ("oil/product", "manual_addition", "Product/market form approved by user."),
    "loose edible oil": ("oil/product", "manual_addition", "Product/market form approved by user."),
    "mixed with": ("adulterant/mixing", "manual_addition", "Mixing/adulteration phrase approved by user."),
    "blending": ("adulterant/mixing", "manual_addition", "Blending/adulteration phrase approved by user."),
    "quality test": ("enforcement/evidence", "manual_addition", "Testing/evidence phrase approved by user."),
    "lab test": ("enforcement/evidence", "manual_addition", "Testing/evidence phrase approved by user."),
}

PRODUCT_HINTS = {
    "edible oil",
    "cooking oil",
    "mustard oil",
    "coconut oil",
    "palm oil",
    "rice bran oil",
    "soybean oil",
    "soyabean oil",
    "sunflower oil",
    "groundnut oil",
    "refined oil",
    "cottonseed oil",
    "loose oil",
    "loose edible oil",
    "palmolein",
}

FULL_INCIDENT_PHRASES = {
    "adulterated edible oil",
    "adulterated coconut oil",
    "adulterated cooking oil",
    "fake cooking oil",
}

FRAUD_HINTS = {
    "adulterated",
    "adulteration",
    "adulterated oil",
    "fake",
    "spurious",
    "substandard",
    "substandard quality",
    "counterfeit",
    "unsafe",
    "oil adulterated",
}

ENFORCEMENT_HINTS = {
    "food safety",
    "food safety department",
    "food safety officer",
    "food safety officers",
    "food safety officials",
    "food safety standards",
    "fssai",
    "fda",
    "fsda",
    "seized",
    "seizure",
    "raid",
    "raids",
    "raided",
    "crackdown",
    "quality test",
    "lab test",
    "oil seized",
}

ADULTERANT_MIXING_HINTS = {
    "argemone oil",
    "mineral oil",
    "mixed with",
    "blending",
}

OUTPUT_COLUMNS = [
    "term",
    "category",
    "query_role",
    "use",
    "source",
    "source_round_number",
    "next_round_target",
    "reason",
    "original_review_label",
    "original_category",
    "composite_score",
    "document_frequency",
    "total_frequency",
    "methods_found",
    "example_article_title",
    "example_context",
    "example_url",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Finalize Round 1 positive keyword bank for Round 2.")
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_KEYWORD_DIR)
    return parser.parse_args()


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def normalize_keep(value: Any) -> str:
    text = "" if value is None else str(value).strip()
    return "1" if text in {"1", "1.0"} else "0"


def read_workbook_rows(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Keyword Review"]
    headers = [
        str(ws.cell(1, col).value).strip() if ws.cell(1, col).value is not None else ""
        for col in range(1, ws.max_column + 1)
    ]
    idx = {header: col + 1 for col, header in enumerate(headers)}
    rows = []
    for row_idx in range(2, ws.max_row + 1):
        term = ws.cell(row_idx, idx["keyword_or_keyphrase"]).value
        if not term:
            continue
        rows.append(
            {
                "term": str(term).strip().lower(),
                "keep": normalize_keep(ws.cell(row_idx, idx["keep"]).value),
                "excel_row": row_idx,
                "original_review_label": ws.cell(row_idx, idx["current_label"]).value or "",
                "original_category": ws.cell(row_idx, idx["category"]).value or "",
                "composite_score": ws.cell(row_idx, idx["composite_score"]).value or "",
                "document_frequency": ws.cell(row_idx, idx["document_frequency"]).value or "",
                "total_frequency": ws.cell(row_idx, idx["total_frequency"]).value or "",
                "methods_found": ws.cell(row_idx, idx["methods_found"]).value or "",
                "example_article_title": ws.cell(row_idx, idx["example_article_title"]).value or "",
                "example_context": ws.cell(row_idx, idx["example_context"]).value or "",
                "example_url": ws.cell(row_idx, idx["example_url"]).value or "",
            }
        )
    return rows


def classify_term(term: str) -> str:
    if term in FULL_INCIDENT_PHRASES:
        return "incident/action phrase"
    if term in PRODUCT_HINTS:
        return "oil/product"
    if term in FRAUD_HINTS:
        return "fraud/adulteration"
    if term in ENFORCEMENT_HINTS:
        return "enforcement/evidence"
    if term in ADULTERANT_MIXING_HINTS:
        return "adulterant/mixing"
    return "support/other"


def query_role(term: str, category: str) -> str:
    if term in COMBINE_ONLY_TERMS:
        return "combine_only"
    if category == "oil/product":
        return "product"
    if category == "fraud/adulteration":
        return "fraud_signal"
    if category == "enforcement/evidence":
        return "enforcement_signal"
    if category == "adulterant/mixing":
        return "adulterant_or_mixing_signal"
    if category == "incident/action phrase":
        return "exact_incident_phrase"
    return "support_only"


def use_value(role: str) -> str:
    return "combine_only" if role in {"combine_only", "support_only"} else "query_component"


def build_review_terms(rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    kept = [row for row in rows if row["keep"] == "1"]
    removed = []
    terms: OrderedDict[str, dict[str, Any]] = OrderedDict()
    for row in kept:
        term = row["term"]
        if term in REMOVE_TERMS:
            removed.append({**row, "removal_reason": "agreed_remove_or_do_not_use"})
            continue
        category = classify_term(term)
        role = query_role(term, category)
        terms[term] = {
            "term": term,
            "category": category,
            "query_role": role,
            "use": use_value(role),
            "source": "review_workbook_keep_1",
            "source_round_number": "1",
            "next_round_target": "2",
            "reason": reason_for(term, category, role),
            **{col: row.get(col, "") for col in OUTPUT_COLUMNS if col in row},
        }
    return list(terms.values()), removed


def add_manual_terms(terms: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_term = OrderedDict((row["term"], row) for row in terms)
    for term, (category, source, reason) in MANUAL_ADDITIONS.items():
        role = query_role(term, category)
        row = by_term.get(term, {})
        by_term[term] = {
            "term": term,
            "category": category,
            "query_role": role,
            "use": use_value(role),
            "source": row.get("source", source) if row else source,
            "source_round_number": "1",
            "next_round_target": "2",
            "reason": reason,
            "original_review_label": row.get("original_review_label", ""),
            "original_category": row.get("original_category", ""),
            "composite_score": row.get("composite_score", ""),
            "document_frequency": row.get("document_frequency", ""),
            "total_frequency": row.get("total_frequency", ""),
            "methods_found": row.get("methods_found", ""),
            "example_article_title": row.get("example_article_title", ""),
            "example_context": row.get("example_context", ""),
            "example_url": row.get("example_url", ""),
        }
    return list(by_term.values())


def reason_for(term: str, category: str, role: str) -> str:
    if term in COMBINE_ONLY_TERMS:
        return "Useful only when combined with product/fraud/enforcement terms; do not query alone."
    if category == "oil/product":
        return "Oil/product term for Round 2 Boolean query product group."
    if category == "fraud/adulteration":
        return "Fraud/adulteration signal for Round 2 Boolean query fraud group."
    if category == "enforcement/evidence":
        return "Enforcement/evidence signal for Round 2 Boolean query enforcement group."
    if category == "adulterant/mixing":
        return "Adulterant or mixing signal for Round 2 Boolean query expansion."
    if category == "incident/action phrase":
        return "Exact incident phrase useful for high-precision query variants."
    return "Support term retained from human review; use cautiously."


def sort_terms(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    order = {
        "oil/product": 0,
        "fraud/adulteration": 1,
        "adulterant/mixing": 2,
        "enforcement/evidence": 3,
        "incident/action phrase": 4,
        "support/other": 5,
    }
    role_order = {
        "product": 0,
        "fraud_signal": 1,
        "adulterant_or_mixing_signal": 2,
        "enforcement_signal": 3,
        "exact_incident_phrase": 4,
        "combine_only": 5,
        "support_only": 6,
    }
    return sorted(rows, key=lambda row: (order.get(row["category"], 99), role_order.get(row["query_role"], 99), row["term"]))


def write_csv(path: Path, rows: list[dict[str, Any]], columns: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    args = parse_args()
    rows = read_workbook_rows(args.workbook)
    review_terms, removed = build_review_terms(rows)
    final_terms = sort_terms(add_manual_terms(review_terms))
    output_dir = args.output_dir
    output_dir.mkdir(parents=True, exist_ok=True)

    bank_csv = output_dir / "round_01_positive_keyword_bank.csv"
    bank_json = output_dir / "round_01_positive_keyword_bank.json"
    removed_csv = output_dir / "round_01_positive_keywords_removed.csv"
    summary_json = output_dir / "round_01_positive_keyword_bank_summary.json"

    write_csv(bank_csv, final_terms, OUTPUT_COLUMNS)
    bank_json.write_text(json.dumps(final_terms, indent=2, ensure_ascii=False), encoding="utf-8")
    write_csv(
        removed_csv,
        removed,
        [
            "term",
            "removal_reason",
            "original_review_label",
            "original_category",
            "composite_score",
            "document_frequency",
            "total_frequency",
            "example_article_title",
        ],
    )
    summary = {
        "created_at": utc_now(),
        "workbook": str(args.workbook),
        "review_rows_total": len(rows),
        "review_keep_1_rows": sum(1 for row in rows if row["keep"] == "1"),
        "review_keep_0_or_blank_rows": sum(1 for row in rows if row["keep"] != "1"),
        "removed_agreed_terms": sorted(REMOVE_TERMS),
        "removed_review_kept_rows": len(removed),
        "manual_additions": sorted(MANUAL_ADDITIONS),
        "manual_addition_count": len(MANUAL_ADDITIONS),
        "final_keyword_count": len(final_terms),
        "final_keyword_counts_by_category": count_by(final_terms, "category"),
        "final_keyword_counts_by_query_role": count_by(final_terms, "query_role"),
        "outputs": {
            "positive_keyword_bank_csv": str(bank_csv),
            "positive_keyword_bank_json": str(bank_json),
            "removed_terms_csv": str(removed_csv),
        },
    }
    summary_json.write_text(json.dumps(summary, indent=2, ensure_ascii=False), encoding="utf-8")
    print(json.dumps(summary, indent=2, ensure_ascii=False))


def count_by(rows: list[dict[str, Any]], key: str) -> dict[str, int]:
    counts: OrderedDict[str, int] = OrderedDict()
    for row in rows:
        value = str(row.get(key, ""))
        counts[value] = counts.get(value, 0) + 1
    return dict(counts)


if __name__ == "__main__":
    main()
