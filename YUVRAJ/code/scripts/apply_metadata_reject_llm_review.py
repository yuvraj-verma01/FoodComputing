from __future__ import annotations

import argparse
import csv
import json
import sys
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from crawler.review_dedupe import load_reviewed_url_keys, split_new_review_rows


DEFAULT_OUTPUT_DIR = ROOT / Path(
    "data/runs/edible_oil_adulteration_round_02_2026-06-23/"
    "mediacloud/outputs/oil_relevance"
)
DEFAULT_WORKBOOK = DEFAULT_OUTPUT_DIR / "metadata_reject_llm_review.xlsx"
DEFAULT_SCORED_CSV = DEFAULT_OUTPUT_DIR / "metadata_reject_llm_scored_candidates.csv"

HUMAN_COLUMNS = [
    "human_keep",
    "human_label",
    "human_review_source",
    "already_human_marked",
    "review_priority",
    "llm_score",
    "llm_confidence",
    "llm_reason",
    "title",
    "source",
    "date",
    "url",
    "second_pass_action",
    "second_pass_reason",
    "query_family",
    "query_id",
    "query_used",
]

CRAWL_QUEUE_COLUMNS = [
    "article_id",
    "title",
    "source",
    "date",
    "url",
    "query_family",
    "query_id",
    "query_used",
    "crawl_priority",
    "review_reason",
    "score",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Apply human keep/drop labels from the metadata-reject LLM review workbook."
    )
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--scored-csv", type=Path, default=DEFAULT_SCORED_CSV)
    parser.add_argument("--blank-is-zero", action="store_true", default=True)
    parser.add_argument("--strict-blanks", action="store_true", help="Error if any keep cell is blank.")
    return parser.parse_args()


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def write_csv(path: Path, rows: list[dict[str, Any]], columns: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def read_review_labels(path: Path) -> tuple[dict[str, str], Counter[str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    if "Review" not in wb.sheetnames:
        raise ValueError(f"Workbook has no Review sheet: {path}")
    ws = wb["Review"]
    headers = [str(ws.cell(1, col).value or "").strip() for col in range(1, ws.max_column + 1)]
    if "keep" not in headers or "url" not in headers:
        raise ValueError("Review sheet must contain keep and url columns.")
    keep_col = headers.index("keep") + 1
    url_col = headers.index("url") + 1

    labels: dict[str, str] = {}
    raw_counts: Counter[str] = Counter()
    for row_idx in range(2, ws.max_row + 1):
        url = str(ws.cell(row_idx, url_col).value or "").strip()
        if not url:
            continue
        raw = str(ws.cell(row_idx, keep_col).value or "").strip()
        raw_counts[raw] += 1
        if raw == "":
            labels[url] = "0"
            continue
        if raw not in {"0", "1"}:
            raise ValueError(f"Invalid keep value at row {row_idx}: {raw!r}. Use 0 or 1.")
        labels[url] = raw
    return labels, raw_counts


def merge_labels(
    scored_rows: list[dict[str, str]],
    labels: dict[str, str],
    workbook: Path,
) -> list[dict[str, Any]]:
    merged = []
    missing = sorted({row.get("url", "") for row in scored_rows if row.get("url", "") not in labels})
    if missing:
        raise ValueError(f"{len(missing)} scored rows are missing from workbook labels.")
    for row in scored_rows:
        keep = labels[row.get("url", "")]
        merged.append(
            {
                **row,
                "human_keep": keep,
                "human_label": "rescue_for_crawl" if keep == "1" else "drop",
                "human_review_source": f"{workbook.name}:Review",
            }
        )
    return merged


def crawl_queue_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    queue = []
    for row in rows:
        if row.get("human_keep") != "1":
            continue
        queue.append(
            {
                "article_id": row.get("article_id", ""),
                "title": row.get("title", ""),
                "source": row.get("source", ""),
                "date": row.get("date", ""),
                "url": row.get("url", ""),
                "query_family": row.get("query_family", ""),
                "query_id": row.get("query_id", ""),
                "query_used": row.get("query_used", ""),
                "crawl_priority": row.get("review_priority", "medium"),
                "review_reason": row.get("llm_reason", "") or row.get("second_pass_reason", ""),
                "score": row.get("llm_score", ""),
            }
        )
    queue.sort(key=lambda row: ({"high": 0, "medium": 1, "low": 2}.get(row["crawl_priority"], 9), row["title"]))
    return queue


def style_sheet(ws, columns: list[str]) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    keep_fill = PatternFill("solid", fgColor="E2F0D9")
    drop_fill = PatternFill("solid", fgColor="FCE4D6")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    widths = {
        "A": 10,
        "B": 18,
        "E": 14,
        "F": 10,
        "G": 12,
        "H": 48,
        "I": 68,
        "J": 22,
        "K": 14,
        "L": 60,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    label_col = columns.index("human_keep") + 1 if "human_keep" in columns else None
    url_col = columns.index("url") + 1 if "url" in columns else None
    for row_idx in range(2, ws.max_row + 1):
        fill = None
        if label_col:
            fill = keep_fill if ws.cell(row_idx, label_col).value == "1" else drop_fill
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row_idx, col_idx)
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        if url_col:
            url_cell = ws.cell(row_idx, url_col)
            if url_cell.value:
                url_cell.hyperlink = str(url_cell.value)
                url_cell.style = "Hyperlink"
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def add_sheet(wb: Workbook, title: str, rows: list[dict[str, Any]], columns: list[str]) -> None:
    ws = wb.create_sheet(title)
    ws.append(columns)
    for row in rows:
        ws.append([row.get(col, "") for col in columns])
    style_sheet(ws, columns)


def write_workbook(path: Path, rows: list[dict[str, Any]], summary: dict[str, Any]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "summary"
    ws.append(["metric", "value"])
    for key, value in summary.items():
        ws.append([key, json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list)) else value])
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 100
    for cell in ws[1]:
        cell.font = Font(bold=True)
    add_sheet(wb, "kept_for_crawl", [row for row in rows if row.get("human_keep") == "1"], HUMAN_COLUMNS)
    add_sheet(wb, "dropped", [row for row in rows if row.get("human_keep") == "0"], HUMAN_COLUMNS)
    add_sheet(wb, "all_reviewed", rows, HUMAN_COLUMNS)
    wb.save(path)


def main() -> int:
    args = parse_args()
    labels, raw_counts = read_review_labels(args.workbook)
    if args.strict_blanks and raw_counts.get("", 0):
        raise ValueError(f"Blank keep cells found: {raw_counts['']}.")

    scored_rows = read_csv(args.scored_csv)
    merged_rows = merge_labels(scored_rows, labels, args.workbook)
    _, already_reviewed = split_new_review_rows(merged_rows, load_reviewed_url_keys())
    if already_reviewed:
        raise ValueError(f"Review contains {len(already_reviewed)} URLs already marked in master corpus.")

    queue = crawl_queue_rows(merged_rows)
    keep_counts = Counter(row.get("human_keep", "") for row in merged_rows)
    priority_counts = Counter(row.get("crawl_priority", "") for row in queue)
    summary = {
        "created_at": utc_now(),
        "workbook": str(args.workbook),
        "scored_csv": str(args.scored_csv),
        "blank_keep_cells_treated_as_zero": raw_counts.get("", 0),
        "review_rows": len(merged_rows),
        "keep_counts": dict(sorted(keep_counts.items())),
        "kept_for_crawl": len(queue),
        "dropped_by_human": keep_counts.get("0", 0),
        "already_reviewed_url_matches": len(already_reviewed),
        "crawl_priority_counts": dict(sorted(priority_counts.items())),
    }

    output_dir = args.output_dir
    write_csv(output_dir / "metadata_reject_llm_human_reviewed_candidates.csv", merged_rows, HUMAN_COLUMNS)
    write_csv(
        output_dir / "metadata_reject_llm_human_kept_for_crawl.csv",
        [row for row in merged_rows if row.get("human_keep") == "1"],
        HUMAN_COLUMNS,
    )
    write_csv(
        output_dir / "metadata_reject_llm_human_dropped.csv",
        [row for row in merged_rows if row.get("human_keep") == "0"],
        HUMAN_COLUMNS,
    )
    write_csv(output_dir / "metadata_reject_llm_human_crawl_queue.csv", queue, CRAWL_QUEUE_COLUMNS)
    (output_dir / "metadata_reject_llm_human_review_summary.json").write_text(
        json.dumps(summary, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    write_workbook(output_dir / "metadata_reject_llm_human_reviewed.xlsx", merged_rows, summary)
    print(json.dumps(summary, indent=2, ensure_ascii=False))
    print(f"Crawl queue written: {(output_dir / 'metadata_reject_llm_human_crawl_queue.csv').resolve()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
