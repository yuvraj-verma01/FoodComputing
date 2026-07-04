"""Create a formatted workbook for reviewing discovered Media Cloud URLs."""

from __future__ import annotations

import argparse
import csv
import json
import sys
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from crawler.review_dedupe import load_reviewed_url_keys, split_new_review_rows

DEFAULT_RUN_DIR = ROOT / "data" / "runs" / "edible_oils_boolean_2026-06-21"

HEADER_FILL = "1F4E79"
HEADER_FG = "FFFFFF"
BORDER = "BFBFBF"
KEEP_FILL = "FFF2CC"
KEEP_YES_FILL = "C6EFCE"
KEEP_NO_FILL = "F4CCCC"
STRONG_FILL = "D9EAD3"
OIL_ONLY_FILL = "FFF2CC"
BROAD_FILL = "FCE4D6"
WEAK_FILL = "E7E6E6"
LINK_BLUE = "0563C1"

REVIEW_COLUMNS = [
    ("review_keep", "keep", 8),
    ("title_screen_label", "title_screen_label", 24),
    ("novelty_vs_boolean", "novelty", 22),
    ("query_family", "query_family", 20),
    ("query_id", "query_id", 30),
    ("published_date", "published_date", 14),
    ("domain", "domain", 28),
    ("source", "source", 24),
    ("title_snippet", "title", 70),
    ("title_url_product_terms", "product_terms", 26),
    ("title_url_signal_terms", "signal_terms", 28),
    ("status", "status", 12),
    ("url", "url", 70),
    ("query_used", "query_used", 55),
]


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--run-dir",
        type=Path,
        default=DEFAULT_RUN_DIR,
        help="Run directory containing mediacloud/outputs/discovery_url_review.csv.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="Output XLSX path.",
    )
    parser.add_argument(
        "--include-reviewed-urls",
        action="store_true",
        help="Keep URLs already human-marked in previous rounds in the workbook.",
    )
    args = parser.parse_args()

    run_dir = args.run_dir
    output_dir = run_dir / "mediacloud" / "outputs"
    review_csv = output_dir / "discovery_url_review.csv"
    query_counts_csv = output_dir / "discovery_query_counts.csv"
    summary_json = output_dir / "discovery_summary.json"
    xlsx_path = args.output or (output_dir / "discovery_url_review.xlsx")

    rows = read_rows(review_csv)
    previously_reviewed_rows: list[dict[str, str]] = []
    if not args.include_reviewed_urls:
        rows, previously_reviewed_rows = split_new_review_rows(rows, load_reviewed_url_keys())
    query_families = read_query_families(query_counts_csv)
    for row in rows:
        row["query_family"] = query_families.get(row.get("query_used") or "", "not_in_plan")

    wb = Workbook()
    review_ws = wb.active
    review_ws.title = "URL Review"
    write_review_sheet(review_ws, rows)
    write_summary_sheet(
        wb.create_sheet("Summary"),
        rows,
        summary_json,
        review_csv,
        xlsx_path,
        len(previously_reviewed_rows),
    )
    write_query_sheet(wb.create_sheet("Query Counts"), query_counts_csv)
    write_instructions_sheet(wb.create_sheet("Instructions"))

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)
    print(json.dumps({"output": str(xlsx_path), "rows": len(rows)}, indent=2))
    return 0


def read_rows(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def read_query_families(path: Path) -> dict[str, str]:
    if not path.exists():
        return {}
    mapping = {}
    with path.open(encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            query = row.get("query") or ""
            if row.get("query_family"):
                mapping[query] = row.get("query_family") or "planned_query"
                continue
            product = row.get("product_term") or ""
            fraud = row.get("fraud_term") or ""
            if product and fraud:
                mapping[query] = f"{product} | {fraud}"
            elif product:
                mapping[query] = product
            else:
                mapping[query] = row.get("template_type") or "planned_query"
    return mapping


def write_review_sheet(ws, rows: list[dict[str, str]]) -> None:
    thin = thin_border()
    header_font = Font(name="Calibri", bold=True, color=HEADER_FG, size=11)
    header_fill = PatternFill("solid", fgColor=HEADER_FILL)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, (_key, header, width) in enumerate(REVIEW_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin
        cell.alignment = header_align
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws["A1"].comment = Comment("Enter 1 to keep this URL, or 0 to drop it.", "Codex")
    ws.row_dimensions[1].height = 32

    validation = DataValidation(
        type="list",
        formula1='"0,1"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Use 0 or 1",
        error="Enter 1 to keep or 0 to drop.",
    )
    ws.add_data_validation(validation)

    for row_idx, row in enumerate(rows, start=2):
        row_fill = label_fill(row.get("title_screen_label", ""))
        for col_idx, (key, _header, _width) in enumerate(REVIEW_COLUMNS, start=1):
            value = row.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.font = Font(name="Calibri", size=10)
            if row_fill:
                cell.fill = row_fill
            if key == "review_keep":
                cell.fill = PatternFill("solid", fgColor=KEEP_FILL)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if key == "url" and value:
                cell.hyperlink = value
                cell.style = "Hyperlink"
                cell.font = Font(name="Calibri", size=10, color=LINK_BLUE, underline="single")
        ws.row_dimensions[row_idx].height = 58

    if rows:
        last_row = len(rows) + 1
        validation.add(f"A2:A{last_row}")
        ws.conditional_formatting.add(
            f"A2:A{last_row}",
            CellIsRule(operator="equal", formula=["1"], fill=PatternFill("solid", fgColor=KEEP_YES_FILL)),
        )
        ws.conditional_formatting.add(
            f"A2:A{last_row}",
            CellIsRule(operator="equal", formula=["0"], fill=PatternFill("solid", fgColor=KEEP_NO_FILL)),
        )

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_properties.tabColor = HEADER_FILL
    for col_idx, (key, _header, _width) in enumerate(REVIEW_COLUMNS, start=1):
        if key == "query_used":
            ws.column_dimensions[get_column_letter(col_idx)].hidden = True
            break


def write_summary_sheet(
    ws,
    rows: list[dict[str, str]],
    summary_json: Path,
    review_csv: Path,
    xlsx_path: Path,
    previously_reviewed_rows: int = 0,
) -> None:
    thin = thin_border()
    header_font = Font(name="Calibri", bold=True, color=HEADER_FG, size=11)
    header_fill = PatternFill("solid", fgColor=HEADER_FILL)

    summary = {}
    if summary_json.exists():
        summary = json.loads(summary_json.read_text(encoding="utf-8"))

    ws["A1"] = "Discovery URL Review"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14)

    overview = [
        ("Input CSV", str(review_csv)),
        ("Output XLSX", str(xlsx_path)),
        ("Total URLs", len(rows)),
        ("Previously reviewed URLs omitted", previously_reviewed_rows),
        ("Articles crawled", summary.get("articles_table_rows", 0)),
        ("Date range", f"{summary.get('published_date_min', '')} to {summary.get('published_date_max', '')}"),
        ("Manual keep coding", "1 = keep, 0 = drop, blank = not reviewed"),
    ]
    for idx, (label, value) in enumerate(overview, start=3):
        ws.cell(idx, 1, label).font = Font(name="Calibri", bold=True, size=11)
        ws.cell(idx, 2, value)
        ws.cell(idx, 2).alignment = Alignment(wrap_text=True, vertical="top")

    row_num = 11
    row_num = write_counter_block(
        ws,
        "Title Screen Labels",
        Counter(row.get("title_screen_label", "") for row in rows),
        row_num,
        thin,
        header_font,
        header_fill,
    )
    row_num = write_counter_block(
        ws,
        "Top Domains",
        Counter(row.get("domain", "") for row in rows).most_common(25),
        row_num,
        thin,
        header_font,
        header_fill,
    )
    write_counter_block(
        ws,
        "Query Families",
        Counter(row.get("query_family", "") for row in rows),
        row_num,
        thin,
        header_font,
        header_fill,
    )

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 90
    ws.freeze_panes = "A11"
    ws.sheet_properties.tabColor = "70AD47"


def write_counter_block(ws, title, counts, start_row, thin, header_font, header_fill) -> int:
    if isinstance(counts, Counter):
        items = counts.most_common()
    else:
        items = list(counts)

    ws.cell(start_row, 1, title)
    ws.cell(start_row, 1).font = Font(name="Calibri", bold=True, size=12)
    start_row += 1
    for col_idx, header in enumerate(("Value", "Count"), start=1):
        cell = ws.cell(start_row, col_idx, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin
    start_row += 1
    for value, count in items:
        ws.cell(start_row, 1, value)
        ws.cell(start_row, 2, count)
        ws.cell(start_row, 1).border = thin
        ws.cell(start_row, 2).border = thin
        start_row += 1
    return start_row + 2


def write_query_sheet(ws, query_counts_csv: Path) -> None:
    rows = read_rows(query_counts_csv) if query_counts_csv.exists() else []
    if not rows:
        ws["A1"] = "No query counts file found."
        return

    thin = thin_border()
    headers = list(rows[0].keys())
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(1, col_idx, header)
        cell.font = Font(name="Calibri", bold=True, color=HEADER_FG, size=11)
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.border = thin
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = 22
    ws.column_dimensions["B"].width = 90

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row_idx, col_idx, row.get(header, ""))
            cell.border = thin
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[row_idx].height = 60

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_properties.tabColor = "FFC000"


def write_instructions_sheet(ws) -> None:
    ws["A1"] = "Instructions"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14)
    lines = [
        ("Task", "Review rows on the URL Review sheet."),
        ("keep = 1", "Use 1 when the URL should remain in the crawl queue."),
        ("keep = 0", "Use 0 when the URL should be dropped before crawling."),
        ("Blank", "Leave blank when not reviewed yet."),
        ("Colors", "Green = strongest title match; yellow = oil context only; orange = broad food-safety title; grey = weak/body-only match."),
        ("Hidden query", "The full Boolean query is hidden in the last column; unhide it only if needed."),
    ]
    for row_idx, (label, text) in enumerate(lines, start=3):
        ws.cell(row_idx, 1, label)
        ws.cell(row_idx, 1).font = Font(name="Calibri", bold=True, size=11)
        ws.cell(row_idx, 2, text)
        ws.cell(row_idx, 2).alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 95
    ws.sheet_properties.tabColor = "A9D18E"


def label_fill(label: str):
    return {
        "strong_title_match": PatternFill("solid", fgColor=STRONG_FILL),
        "oil_context_only_in_title": PatternFill("solid", fgColor=OIL_ONLY_FILL),
        "ghee_context_only_in_title": PatternFill("solid", fgColor=OIL_ONLY_FILL),
        "broad_food_safety_title": PatternFill("solid", fgColor=BROAD_FILL),
        "weak_or_body_only_match": PatternFill("solid", fgColor=WEAK_FILL),
    }.get(label)


def thin_border() -> Border:
    side = Side(style="thin", color=BORDER)
    return Border(left=side, right=side, top=side, bottom=side)


if __name__ == "__main__":
    raise SystemExit(main())
