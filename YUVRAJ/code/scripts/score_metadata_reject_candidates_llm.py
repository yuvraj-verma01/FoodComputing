from __future__ import annotations

import argparse
import csv
import json
import sys
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from crawler.review_dedupe import load_reviewed_url_keys, split_new_review_rows, url_keys


DEFAULT_OUTPUT_DIR = ROOT / Path(
    "data/runs/edible_oil_adulteration_round_02_2026-06-23/"
    "mediacloud/outputs/oil_relevance"
)
DEFAULT_WORKBOOK = DEFAULT_OUTPUT_DIR / "metadata_reject_second_pass_audit.xlsx"

OUTPUT_COLUMNS = [
    "keep",
    "already_human_marked",
    "review_priority",
    "llm_candidate_label",
    "llm_score",
    "llm_confidence",
    "llm_reason",
    "llm_metadata_evidence",
    "llm_risk_flags",
    "second_pass_action",
    "second_pass_reason",
    "score",
    "metadata_decision",
    "title",
    "source",
    "date",
    "url",
    "product_hits",
    "adulteration_hits",
    "enforcement_hits",
    "india_hits",
    "exclude_hits",
    "review_reason",
    "query_id",
    "query_family",
    "query_used",
]

READABLE_COLUMNS = [
    ("keep", "keep", 8),
    ("review_priority", "priority", 12),
    ("llm_score", "LLM score", 10),
    ("title", "title", 70),
    ("source", "source", 22),
    ("date", "date", 14),
    ("llm_reason", "LLM reason", 48),
    ("second_pass_reason", "why included", 48),
    ("url", "url", 52),
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="LLM-score metadata-rejected rescue candidates before human review."
    )
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--sheet", default="candidate_rescue")
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--model", default="llama3.1:8b-instruct-q4_K_M")
    parser.add_argument("--timeout-seconds", type=int, default=240)
    parser.add_argument("--limit", type=int, default=0)
    parser.add_argument(
        "--include-reviewed-urls",
        action="store_true",
        help="Keep URLs already human-marked in previous rounds in the LLM scoring input.",
    )
    return parser.parse_args()


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def read_sheet(path: Path, sheet_name: str) -> list[dict[str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Workbook {path} has no sheet {sheet_name!r}; found {wb.sheetnames}")
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value or "").strip() for value in rows[0]]
    out: list[dict[str, str]] = []
    for values in rows[1:]:
        row = {
            headers[index]: "" if value is None else str(value)
            for index, value in enumerate(values)
            if index < len(headers) and headers[index]
        }
        if any(row.values()):
            out.append(row)
    return out


def write_csv(path: Path, rows: list[dict[str, Any]], columns: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def dedupe_input_rows(
    rows: list[dict[str, str]],
    include_reviewed_urls: bool,
) -> tuple[list[dict[str, str]], list[dict[str, str]], list[dict[str, str]], int]:
    seen_keys: set[str] = set()
    unique: list[dict[str, str]] = []
    duplicates_in_input: list[dict[str, str]] = []
    for row in rows:
        keys = url_keys(row.get("url", ""))
        if keys & seen_keys:
            duplicates_in_input.append(row)
            continue
        seen_keys.update(keys)
        unique.append(row)

    reviewed_key_count = 0
    already_reviewed: list[dict[str, str]] = []
    if not include_reviewed_urls:
        reviewed_keys = load_reviewed_url_keys()
        reviewed_key_count = len(reviewed_keys)
        unique, already_reviewed = split_new_review_rows(unique, reviewed_keys)
    return unique, already_reviewed, duplicates_in_input, reviewed_key_count


def build_prompt(row: dict[str, str]) -> str:
    return f"""
You are triaging metadata-only candidates for an Indian edible-oil adulteration news corpus.

Important: you only have metadata, not full article text. Do not claim final article relevance. Decide whether this URL is worth human review/crawling.

Target corpus definition:
- Keep/review if the article is likely about edible oil/cooking oil/mustard oil/refined oil/palm oil/soybean/groundnut/sesame/sunflower/rice bran/cottonseed/coconut/olive oil itself being adulterated, fake, spurious, unsafe, misbranded, tainted, seized, failed in samples, or targeted by food-safety enforcement in India.

Drop if likely:
- petrol, diesel, crude oil, engine oil, lubricating oil, hair oil, hashish/cannabis oil, essential oil, refinery/tanker/ONGC/global oil
- edible-oil price/import/export/business story with no adulteration or safety incident
- reused cooking oil only
- oil is only an adulterant in another food
- general food, sweets, dairy, spice, gutkha, or non-oil food safety story
- generic recipe/health/tips story with no likely adulteration incident

Scoring:
- 90-100: direct edible-oil adulteration/spurious/tainted/mislabelled/seizure/FDA/FSSAI story
- 70-89: generic "oil" but strongly likely edible oil adulteration/enforcement in India
- 50-69: plausible edible-oil safety/regulation story but metadata is uncertain
- 0-49: probably not worth human review

Return only JSON:
{{
  "label": "review" | "drop" | "unclear",
  "score": 0,
  "confidence": 0.0,
  "reason": "short reason",
  "metadata_evidence": "short phrase from title/source/url/metadata",
  "risk_flags": ["short", "flags"]
}}

Metadata:
Title: {row.get("title", "")}
Source: {row.get("source", "")}
Date: {row.get("date", "")}
URL: {row.get("url", "")}
Second-pass action: {row.get("second_pass_action", "")}
Second-pass reason: {row.get("second_pass_reason", "")}
Original metadata score: {row.get("score", "")}
Product hits: {row.get("product_hits", "")}
Adulteration hits: {row.get("adulteration_hits", "")}
Enforcement hits: {row.get("enforcement_hits", "")}
Exclude hits: {row.get("exclude_hits", "")}
Query ID: {row.get("query_id", "")}
Query used: {row.get("query_used", "")}
""".strip()


def parse_json_object(raw: str) -> dict[str, Any]:
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        start = raw.find("{")
        end = raw.rfind("}")
        if start != -1 and end != -1 and end > start:
            return json.loads(raw[start : end + 1])
        raise


def call_ollama(row: dict[str, str], model: str, timeout_seconds: int) -> dict[str, Any]:
    response = requests.post(
        "http://127.0.0.1:11434/api/generate",
        json={
            "model": model,
            "prompt": build_prompt(row),
            "stream": False,
            "format": "json",
            "options": {"temperature": 0},
        },
        timeout=timeout_seconds,
    )
    response.raise_for_status()
    raw = response.json().get("response") or "{}"
    parsed = parse_json_object(raw)
    label = str(parsed.get("label") or "unclear").strip().lower()
    if label not in {"review", "drop", "unclear"}:
        label = "unclear"
    try:
        score = int(round(float(parsed.get("score", 0))))
    except (TypeError, ValueError):
        score = 0
    try:
        confidence = float(parsed.get("confidence", 0.0))
    except (TypeError, ValueError):
        confidence = 0.0
    flags = parsed.get("risk_flags", [])
    if isinstance(flags, list):
        flags_text = "; ".join(str(item) for item in flags)
    else:
        flags_text = str(flags or "")
    return {
        "llm_candidate_label": label,
        "llm_score": max(0, min(100, score)),
        "llm_confidence": max(0.0, min(1.0, confidence)),
        "llm_reason": str(parsed.get("reason") or "")[:600],
        "llm_metadata_evidence": str(parsed.get("metadata_evidence") or "")[:400],
        "llm_risk_flags": flags_text[:400],
        "llm_raw": raw[:2000],
    }


def score_rows(rows: list[dict[str, str]], model: str, timeout_seconds: int) -> list[dict[str, Any]]:
    scored: list[dict[str, Any]] = []
    for index, row in enumerate(rows, start=1):
        title = row.get("title", "")[:90].encode("ascii", errors="replace").decode("ascii")
        print(f"[LLM {index}/{len(rows)}] {title}", flush=True)
        try:
            llm = call_ollama(row, model=model, timeout_seconds=timeout_seconds)
        except Exception as exc:
            llm = {
                "llm_candidate_label": "unclear",
                "llm_score": 50,
                "llm_confidence": 0.0,
                "llm_reason": f"LLM call failed: {exc}",
                "llm_metadata_evidence": "",
                "llm_risk_flags": "llm_error",
                "llm_raw": "",
            }
        scored.append(
            {
                "keep": "",
                "already_human_marked": "0",
                "review_priority": review_priority(llm["llm_score"]),
                **llm,
                **row,
            }
        )
    scored.sort(
        key=lambda row: (
            {"review": 0, "unclear": 1, "drop": 2}.get(row.get("llm_candidate_label", ""), 3),
            -int(row.get("llm_score") or 0),
            row.get("title", ""),
        )
    )
    return scored


def review_priority(score: int) -> str:
    if score >= 80:
        return "high"
    if score >= 60:
        return "medium"
    return "low"


def clean_display_text(value: Any) -> str:
    text = "" if value is None else str(value)
    if any(marker in text for marker in ("â", "Ã", "Â")):
        try:
            text = text.encode("latin1").decode("utf-8")
        except UnicodeError:
            pass
    replacements = {
        "\u0080": "",
        "\u0098": "'",
        "\u0099": "'",
        "\u009c": '"',
        "\u009d": '"',
        "’": "'",
        "‘": "'",
        "“": '"',
        "”": '"',
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return text


def readable_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    out = []
    for row in rows:
        readable = {}
        for key, _header, _width in READABLE_COLUMNS:
            value = row.get(key, "")
            readable[key] = clean_display_text(value) if isinstance(value, str) else value
        out.append(readable)
    return out


def style_sheet(ws, columns: list[str]) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    review_fill = PatternFill("solid", fgColor="E2F0D9")
    unclear_fill = PatternFill("solid", fgColor="FFF2CC")
    drop_fill = PatternFill("solid", fgColor="FCE4D6")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = {
        "A": 8,
        "B": 18,
        "C": 10,
        "D": 12,
        "E": 42,
        "F": 42,
        "G": 24,
        "H": 20,
        "I": 44,
        "J": 8,
        "K": 20,
        "L": 58,
        "M": 22,
        "N": 14,
        "O": 60,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    label_col = columns.index("llm_candidate_label") + 1 if "llm_candidate_label" in columns else None
    url_col = columns.index("url") + 1 if "url" in columns else None
    for row_idx in range(2, ws.max_row + 1):
        label = ws.cell(row=row_idx, column=label_col).value if label_col else ""
        fill = {"review": review_fill, "unclear": unclear_fill, "drop": drop_fill}.get(label)
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        if url_col:
            cell = ws.cell(row=row_idx, column=url_col)
            if cell.value:
                cell.hyperlink = str(cell.value)
                cell.style = "Hyperlink"
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def style_readable_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    keep_fill = PatternFill("solid", fgColor="FFF2CC")
    high_fill = PatternFill("solid", fgColor="E2F0D9")
    medium_fill = PatternFill("solid", fgColor="FFF2CC")
    low_fill = PatternFill("solid", fgColor="FCE4D6")

    for col_idx, (_key, header, width) in enumerate(READABLE_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width
    ws.row_dimensions[1].height = 28

    validation = DataValidation(
        type="list",
        formula1='"0,1"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Use 0 or 1",
        error="Enter 1 to keep/review further, or 0 to drop.",
    )
    ws.add_data_validation(validation)
    if ws.max_row > 1:
        validation.add(f"A2:A{ws.max_row}")

    priority_col = 2
    keep_col = 1
    url_col = 9
    for row_idx in range(2, ws.max_row + 1):
        priority = ws.cell(row=row_idx, column=priority_col).value
        fill = {"high": high_fill, "medium": medium_fill, "low": low_fill}.get(priority)
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.cell(row=row_idx, column=keep_col).fill = keep_fill
        ws.cell(row=row_idx, column=keep_col).alignment = Alignment(horizontal="center", vertical="center")
        url_cell = ws.cell(row=row_idx, column=url_col)
        if url_cell.value:
            url_cell.hyperlink = str(url_cell.value)
            url_cell.style = "Hyperlink"
        ws.row_dimensions[row_idx].height = 72
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def add_sheet(wb: Workbook, title: str, rows: list[dict[str, Any]], columns: list[str]) -> None:
    ws = wb.create_sheet(title)
    ws.append(columns)
    for row in rows:
        ws.append([row.get(col, "") for col in columns])
    style_sheet(ws, columns)


def add_readable_sheet(wb: Workbook, title: str, rows: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet(title)
    columns = [key for key, _header, _width in READABLE_COLUMNS]
    ws.append([header for _key, header, _width in READABLE_COLUMNS])
    for row in readable_rows(rows):
        ws.append([row.get(col, "") for col in columns])
    style_readable_sheet(ws)


def write_workbook(
    path: Path,
    scored_rows: list[dict[str, Any]],
    already_reviewed: list[dict[str, str]],
    duplicates_in_input: list[dict[str, str]],
    summary: dict[str, Any],
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Start Here"
    ws.append(["field", "value"])
    ws.append(["what_to_edit", "Use the Review sheet. Mark keep=1 to rescue/crawl/review further, keep=0 to drop."])
    ws.append(["duplicate_status", "All rows shown for review have already_human_marked=0 in the full details."])
    ws.append(["priority_meaning", "High = strongest metadata signal, Medium = plausible, Low = weaker but not previously reviewed."])
    for key, value in summary.items():
        ws.append([key, json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list)) else value])
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 120
    for cell in ws[1]:
        cell.font = Font(bold=True)

    add_readable_sheet(wb, "Review", scored_rows)
    add_readable_sheet(wb, "High Priority", [r for r in scored_rows if r.get("review_priority") == "high"])
    add_readable_sheet(wb, "Medium Priority", [r for r in scored_rows if r.get("review_priority") == "medium"])
    add_readable_sheet(wb, "Low Priority", [r for r in scored_rows if r.get("review_priority") == "low"])
    add_sheet(wb, "Full Details", scored_rows, OUTPUT_COLUMNS)
    wb["Full Details"].sheet_state = "hidden"
    if already_reviewed:
        add_sheet(wb, "already_reviewed_excluded", already_reviewed, list(already_reviewed[0].keys()))
    if duplicates_in_input:
        add_sheet(wb, "input_duplicates_excluded", duplicates_in_input, list(duplicates_in_input[0].keys()))
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main() -> int:
    args = parse_args()
    rows = read_sheet(args.workbook, args.sheet)
    rows = [
        row
        for row in rows
        if row.get("second_pass_action") in {"rescue_crawl_likely", "manual_review_maybe"}
    ]
    if args.limit:
        rows = rows[: args.limit]

    new_rows, already_reviewed, duplicates_in_input, reviewed_key_count = dedupe_input_rows(
        rows,
        include_reviewed_urls=args.include_reviewed_urls,
    )
    scored_rows = score_rows(new_rows, model=args.model, timeout_seconds=args.timeout_seconds)

    label_counts = Counter(row.get("llm_candidate_label", "") for row in scored_rows)
    second_pass_counts = Counter(row.get("second_pass_action", "") for row in scored_rows)
    summary = {
        "created_at": utc_now(),
        "input_workbook": str(args.workbook),
        "input_sheet": args.sheet,
        "model": args.model,
        "input_candidate_rows": len(rows),
        "input_duplicates_excluded": len(duplicates_in_input),
        "already_reviewed_rows_excluded": len(already_reviewed),
        "reviewed_url_key_count": reviewed_key_count,
        "new_rows_scored": len(scored_rows),
        "llm_label_counts": dict(sorted(label_counts.items())),
        "second_pass_action_counts": dict(sorted(second_pass_counts.items())),
        "note": "LLM scoring is metadata-only triage, not final article relevance.",
    }

    output_dir = args.output_dir
    scored_csv = output_dir / "metadata_reject_llm_scored_candidates.csv"
    review_xlsx = output_dir / "metadata_reject_llm_review.xlsx"
    summary_path = output_dir / "metadata_reject_llm_summary.json"
    write_csv(scored_csv, scored_rows, OUTPUT_COLUMNS + ["llm_raw"])
    write_csv(output_dir / "metadata_reject_llm_already_reviewed_excluded.csv", already_reviewed, list(rows[0].keys()) if rows else [])
    write_csv(output_dir / "metadata_reject_llm_input_duplicates_excluded.csv", duplicates_in_input, list(rows[0].keys()) if rows else [])
    summary_path.write_text(json.dumps(summary, indent=2, ensure_ascii=False), encoding="utf-8")
    write_workbook(review_xlsx, scored_rows, already_reviewed, duplicates_in_input, summary)

    print(json.dumps(summary, indent=2, ensure_ascii=False))
    print(f"Review workbook written: {review_xlsx.resolve()}")
    print(f"Scored CSV written: {scored_csv.resolve()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
