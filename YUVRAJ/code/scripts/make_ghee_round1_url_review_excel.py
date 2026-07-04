"""Create a cleaner ghee Round 1 URL-review workbook.

This workbook is built directly from the ghee Round 1 discovery database. It
does not dedupe against the edible-oil corpus or any prior human-reviewed oil
files. Rows are sorted strongest first for faster 0/1 marking.
"""

from __future__ import annotations

import csv
import re
import sqlite3
import sys
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

RUN_DIR = ROOT / "data" / "runs" / "ghee_adulteration_round_01_2026-06-30"
OUTPUT_DIR = RUN_DIR / "mediacloud" / "outputs"
DB_PATH = OUTPUT_DIR / "articles.db"
QUERY_PLAN = RUN_DIR / "proposed_mediacloud_ghee_round1_seed_queries.csv"
OUT_XLSX = OUTPUT_DIR / "ghee_round1_url_review_SORTED.xlsx"

PRODUCT_TERMS = {
    "ghee",
    "cow ghee",
    "desi ghee",
    "pure ghee",
    "loose ghee",
    "fake ghee",
    "fake cow ghee",
    "adulterated ghee",
    "suspected adulterated ghee",
    "vegetable ghee",
    "ghee racket",
    "ghee racket busted",
}

SIGNAL_TERMS = {
    "adulterated",
    "fake",
    "seized",
    "raid",
    "food safety",
    "fda",
    "fssai",
    "vanaspati",
}

LABEL_ORDER = {
    "strong_title_match": 0,
    "ghee_context_only_in_title": 1,
    "broad_food_safety_title": 2,
    "weak_or_body_only_match": 3,
}

QUERY_FAMILY_BONUS = {
    "title_only": 20,
    "phrase": 16,
    "boolean": 10,
    "proximity": 6,
}

COLUMNS = [
    ("keep", "keep", 7),
    ("open", "open", 9),
    ("title", "title", 76),
    ("url", "url", 64),
    ("strength_label", "strength", 22),
    ("priority_score", "score", 9),
    ("published_date", "date", 13),
    ("source", "source", 24),
    ("domain", "domain", 28),
    ("query_family", "query_type", 14),
    ("query_id", "query_id", 34),
    ("product_terms", "product_terms", 28),
    ("signal_terms", "signal_terms", 24),
    ("query_used", "query_used", 56),
]

HEADER_FILL = "1F4E79"
HEADER_FG = "FFFFFF"
BORDER = "BFBFBF"
KEEP_FILL = "FFF2CC"
KEEP_YES_FILL = "C6EFCE"
KEEP_NO_FILL = "F4CCCC"
STRONG_FILL = "D9EAD3"
GHEE_ONLY_FILL = "FFF2CC"
BROAD_FILL = "FCE4D6"
WEAK_FILL = "E7E6E6"
LINK_BLUE = "0563C1"


def main() -> int:
    query_lookup = load_query_plan(QUERY_PLAN)
    rows = load_rows(DB_PATH, query_lookup)
    rows.sort(key=sort_key)

    wb = Workbook()
    ws = wb.active
    ws.title = "Review"
    write_review_sheet(ws, rows)
    write_summary_sheet(wb.create_sheet("Summary"), rows)
    write_query_sheet(wb.create_sheet("Queries"), query_lookup)

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    print(f"written={OUT_XLSX}")
    print(f"rows={len(rows)}")
    return 0


def load_query_plan(path: Path) -> dict[str, dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return {row["query"]: row for row in csv.DictReader(handle)}


def load_rows(db_path: Path, query_lookup: dict[str, dict[str, str]]) -> list[dict[str, str]]:
    with sqlite3.connect(db_path) as con:
        con.row_factory = sqlite3.Row
        db_rows = con.execute(
            """
            SELECT url, query_used, title_snippet, source, domain, published_date, status
            FROM discovered_urls
            ORDER BY id ASC
            """
        ).fetchall()

    rows = []
    for row in db_rows:
        rec = dict(row)
        query = rec.get("query_used") or ""
        plan = query_lookup.get(query, {})
        screen_text = " ".join(
            [
                rec.get("title_snippet") or "",
                rec.get("url") or "",
                rec.get("source") or "",
                rec.get("domain") or "",
            ]
        )
        product_matches = matching_terms(screen_text, PRODUCT_TERMS)
        signal_matches = matching_terms(screen_text, SIGNAL_TERMS)
        label = title_screen_label(product_matches, signal_matches)
        score = priority_score(label, plan.get("query_family", ""), product_matches, signal_matches)
        rows.append(
            {
                "keep": "",
                "open": "open",
                "title": rec.get("title_snippet") or "",
                "url": rec.get("url") or "",
                "strength_label": label,
                "priority_score": str(score),
                "published_date": rec.get("published_date") or "",
                "source": rec.get("source") or "",
                "domain": rec.get("domain") or "",
                "query_family": plan.get("query_family", "not_in_plan"),
                "query_id": plan.get("query_id", ""),
                "product_terms": "; ".join(product_matches[:8]),
                "signal_terms": "; ".join(signal_matches[:8]),
                "query_used": query,
            }
        )
    return rows


def normalize(text: str | None) -> str:
    text = (text or "").lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return f" {re.sub(r'\\s+', ' ', text).strip()} "


def matching_terms(text: str, terms: set[str]) -> list[str]:
    normalized = normalize(text)
    matches = []
    for term in sorted(terms, key=lambda item: (-len(item), item)):
        normalized_term = normalize(term).strip()
        if f" {normalized_term} " in normalized:
            matches.append(term)
    return matches


def title_screen_label(product_matches: list[str], signal_matches: list[str]) -> str:
    if product_matches and signal_matches:
        return "strong_title_match"
    if product_matches:
        return "ghee_context_only_in_title"
    if signal_matches:
        return "broad_food_safety_title"
    return "weak_or_body_only_match"


def priority_score(
    label: str,
    query_family: str,
    product_matches: list[str],
    signal_matches: list[str],
) -> int:
    base = {
        "strong_title_match": 100,
        "ghee_context_only_in_title": 70,
        "broad_food_safety_title": 45,
        "weak_or_body_only_match": 10,
    }.get(label, 0)
    return base + QUERY_FAMILY_BONUS.get(query_family, 0) + 3 * len(product_matches) + 5 * len(signal_matches)


def sort_key(row: dict[str, str]) -> tuple[int, int, str, str]:
    return (
        LABEL_ORDER.get(row.get("strength_label", ""), 9),
        -int(row.get("priority_score") or 0),
        row.get("query_family", ""),
        row.get("published_date", ""),
    )


def write_review_sheet(ws, rows: list[dict[str, str]]) -> None:
    thin = thin_border()
    for col_idx, (_key, header, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(1, col_idx, header)
        cell.font = Font(name="Calibri", bold=True, color=HEADER_FG, size=11)
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws["A1"].comment = Comment("Mark 1 to keep, 0 to drop.", "Codex")
    ws.row_dimensions[1].height = 30

    validation = DataValidation(
        type="list",
        formula1='"0,1"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Use 0 or 1",
        error="Enter 1 to keep or 0 to drop.",
    )
    ws.add_data_validation(validation)

    for row_idx, row in enumerate(rows, start=2):
        fill = label_fill(row.get("strength_label", ""))
        for col_idx, (key, _header, _width) in enumerate(COLUMNS, start=1):
            value = row.get(key, "")
            cell = ws.cell(row_idx, col_idx, value)
            cell.border = thin
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.font = Font(name="Calibri", size=10)
            if fill:
                cell.fill = fill
            if key == "keep":
                cell.fill = PatternFill("solid", fgColor=KEEP_FILL)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if key == "open" and row.get("url"):
                cell.hyperlink = row["url"]
                cell.value = "open"
                cell.font = Font(name="Calibri", size=10, color=LINK_BLUE, underline="single", bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if key == "url" and row.get("url"):
                cell.hyperlink = row["url"]
                cell.font = Font(name="Calibri", size=10, color=LINK_BLUE, underline="single")
        ws.row_dimensions[row_idx].height = 48

    if rows:
        last_row = len(rows) + 1
        validation.add(f"A2:A{last_row}")
        ws.conditional_formatting.add(
            f"A2:A{last_row}",
            CellIsRule(operator="equal", formula=["1"], fill=PatternFill("solid", fgColor=KEEP_YES_FILL)),
        )
        ws.conditional_formatting.add(
            f"A2:A{last_row}",
            CellIsRule(operator="equal", formula=["0"], fill=PatternFill("solid", fgColor=KEEP_NO_FILL)),
        )

    ws.freeze_panes = "C2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_properties.tabColor = HEADER_FILL
    ws.column_dimensions[get_column_letter(len(COLUMNS))].hidden = True


def write_summary_sheet(ws, rows: list[dict[str, str]]) -> None:
    ws["A1"] = "Ghee Round 1 URL Review"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14)
    items = [
        ("Rows", len(rows)),
        ("Cross-corpus edible-oil dedupe", "Not applied"),
        ("Sort order", "strong_title_match, ghee_context_only, broad_food_safety, weak/body-only"),
        ("Manual marking", "Use keep = 1 or 0 on the Review sheet"),
        ("Workbook", str(OUT_XLSX)),
    ]
    for idx, (label, value) in enumerate(items, start=3):
        ws.cell(idx, 1, label).font = Font(name="Calibri", bold=True)
        ws.cell(idx, 2, value)
        ws.cell(idx, 2).alignment = Alignment(wrap_text=True, vertical="top")

    row_num = 10
    for title, counter in [
        ("Strength labels", Counter(row["strength_label"] for row in rows)),
        ("Query families", Counter(row["query_family"] for row in rows)),
        ("Top domains", Counter(row["domain"] for row in rows)),
    ]:
        ws.cell(row_num, 1, title).font = Font(name="Calibri", bold=True, size=12)
        row_num += 1
        for key, count in counter.most_common(30):
            ws.cell(row_num, 1, key)
            ws.cell(row_num, 2, count)
            row_num += 1
        row_num += 2
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 90


def write_query_sheet(ws, query_lookup: dict[str, dict[str, str]]) -> None:
    headers = ["query_number", "query_id", "query_family", "query"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(1, col_idx, header)
        cell.font = Font(name="Calibri", bold=True, color=HEADER_FG)
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
    for row_idx, row in enumerate(query_lookup.values(), start=2):
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row_idx, col_idx, row.get(header, ""))
    for idx, width in enumerate([12, 42, 18, 90], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def label_fill(label: str):
    return {
        "strong_title_match": PatternFill("solid", fgColor=STRONG_FILL),
        "ghee_context_only_in_title": PatternFill("solid", fgColor=GHEE_ONLY_FILL),
        "broad_food_safety_title": PatternFill("solid", fgColor=BROAD_FILL),
        "weak_or_body_only_match": PatternFill("solid", fgColor=WEAK_FILL),
    }.get(label)


def thin_border() -> Border:
    side = Side(style="thin", color=BORDER)
    return Border(left=side, right=side, top=side, bottom=side)


if __name__ == "__main__":
    raise SystemExit(main())
