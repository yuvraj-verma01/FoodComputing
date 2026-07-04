from __future__ import annotations

import argparse
import csv
import json
import sqlite3
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


DEFAULT_OUTPUT_DIR = Path(
    "data/runs/edible_oils_boolean_title_proximity_2026-06-22/"
    "mediacloud/outputs/oil_relevance"
)
DEFAULT_DB = Path(
    "data/runs/edible_oils_boolean_title_proximity_2026-06-22/"
    "mediacloud/outputs/articles.db"
)

OUTPUT_COLUMNS = [
    "final_keep",
    "final_human_label",
    "human_review_status",
    "human_review_source",
    "model_final_label",
    "model_confidence",
    "title",
    "source",
    "date",
    "url",
    "domain",
    "publication_date",
    "word_count",
    "reason",
    "evidence_phrase",
    "llm_label",
    "llm_confidence",
    "llm_reason",
    "oil_role",
    "edible_oil_terms",
    "adulteration_action_terms",
    "negative_terms",
    "query_family",
    "query_id",
    "article_id",
    "file_path",
    "cleaned_text_path",
    "article_text",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Freeze final human-reviewed corpus outputs.")
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--db", type=Path, default=DEFAULT_DB)
    parser.add_argument(
        "--workbook",
        type=Path,
        default=DEFAULT_OUTPUT_DIR / "human_reviewed_corpus.xlsx",
    )
    return parser.parse_args()


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def norm_keep(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text in {"1", "1.0"}:
        return "1"
    if text in {"0", "0.0"}:
        return "0"
    return ""


def read_sheet(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    try:
        headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    except StopIteration:
        return []
    records = []
    for row in rows:
        records.append({headers[idx]: value for idx, value in enumerate(row) if idx < len(headers)})
    return records


def load_article_text(db_path: Path) -> dict[str, dict[str, Any]]:
    with sqlite3.connect(db_path) as con:
        con.row_factory = sqlite3.Row
        rows = con.execute(
            """
            SELECT url, domain, publication_date, word_count, cleaned_text_path, article_text
            FROM articles
            """
        ).fetchall()
    return {row["url"]: dict(row) for row in rows}


def overlay_needs_review(
    all_rows: list[dict[str, Any]],
    needs_review_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    needs_by_url = {
        str(row.get("url") or "").strip(): row
        for row in needs_review_rows
        if str(row.get("url") or "").strip()
    }
    frozen = []
    for row in all_rows:
        url = str(row.get("url") or "").strip()
        override = needs_by_url.get(url)
        keep = norm_keep(override.get("human_keep") if override else row.get("human_keep"))
        final = dict(row)
        final["final_keep"] = keep
        final["final_human_label"] = "relevant" if keep == "1" else "irrelevant" if keep == "0" else ""
        final["human_review_status"] = "human_reviewed" if keep in {"0", "1"} else "needs_review"
        if override and keep:
            source = str(row.get("human_review_source") or "")
            final["human_review_source"] = f"{source}; human_reviewed_corpus.xlsx:needs_review".strip("; ")
        frozen.append(final)
    return frozen


def attach_text(rows: list[dict[str, Any]], article_text_by_url: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    out = []
    for row in rows:
        url = str(row.get("url") or "")
        article = article_text_by_url.get(url, {})
        merged = dict(row)
        for key in ["domain", "publication_date", "word_count", "cleaned_text_path", "article_text"]:
            if article.get(key) is not None:
                merged[key] = article.get(key)
        out.append({col: merged.get(col, "") for col in OUTPUT_COLUMNS})
    return out


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=OUTPUT_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def write_jsonl(path: Path, rows: list[dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8") as handle:
        for row in rows:
            handle.write(json.dumps(row, ensure_ascii=False) + "\n")


def add_sheet(wb: Workbook, title: str, rows: list[dict[str, Any]], columns: list[str]) -> None:
    ws = wb.create_sheet(title)
    ws.append(columns)
    for row in rows:
        ws.append([row.get(col, "") for col in columns])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    keep_fill = PatternFill("solid", fgColor="E2F0D9")
    drop_fill = PatternFill("solid", fgColor="FCE4D6")
    review_fill = PatternFill("solid", fgColor="FFF2CC")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = {
        "A": 10,
        "B": 18,
        "C": 18,
        "D": 40,
        "E": 16,
        "F": 12,
        "G": 58,
        "H": 22,
        "I": 14,
        "J": 54,
        "K": 22,
        "L": 15,
        "M": 12,
        "N": 45,
        "O": 45,
        "P": 14,
        "Q": 14,
        "R": 42,
        "S": 18,
        "Z": 52,
        "AA": 52,
        "AB": 70,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    keep_col = columns.index("final_keep") + 1
    url_col = columns.index("url") + 1
    for row_idx in range(2, ws.max_row + 1):
        keep = str(ws.cell(row=row_idx, column=keep_col).value or "")
        fill = keep_fill if keep == "1" else drop_fill if keep == "0" else review_fill
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        url_cell = ws.cell(row=row_idx, column=url_col)
        if url_cell.value:
            url_cell.hyperlink = str(url_cell.value)
            url_cell.style = "Hyperlink"
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def write_workbook(path: Path, rows: list[dict[str, Any]], summary: dict[str, Any]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "summary"
    ws.append(["metric", "value"])
    for key, value in summary.items():
        ws.append([key, json.dumps(value, ensure_ascii=False) if isinstance(value, dict) else value])
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 90
    for cell in ws[1]:
        cell.font = Font(bold=True)

    public_columns = [col for col in OUTPUT_COLUMNS if col != "article_text"]
    add_sheet(wb, "final_all", rows, public_columns)
    add_sheet(wb, "final_relevant", [row for row in rows if row["final_keep"] == "1"], public_columns)
    add_sheet(wb, "final_irrelevant", [row for row in rows if row["final_keep"] == "0"], public_columns)
    add_sheet(wb, "unresolved", [row for row in rows if row["final_keep"] == ""], public_columns)
    wb.save(path)


def main() -> None:
    args = parse_args()
    all_rows = read_sheet(args.workbook, "all_articles")
    needs_review_rows = read_sheet(args.workbook, "needs_review")
    overlaid = overlay_needs_review(all_rows, needs_review_rows)
    rows = attach_text(overlaid, load_article_text(args.db))

    counts = Counter(row["final_human_label"] or "needs_review" for row in rows)
    model_counts = Counter(row.get("model_final_label") or "" for row in rows)
    relevant = [row for row in rows if row["final_keep"] == "1"]
    irrelevant = [row for row in rows if row["final_keep"] == "0"]
    unresolved = [row for row in rows if row["final_keep"] == ""]
    relevant_with_text = sum(1 for row in relevant if row.get("article_text"))

    summary = {
        "created_at": utc_now(),
        "source_workbook": str(args.workbook),
        "total_rows": len(rows),
        "final_label_counts": dict(counts),
        "model_label_counts": dict(model_counts),
        "final_relevant_with_article_text": relevant_with_text,
        "unresolved_rows": len(unresolved),
    }

    output_dir = args.output_dir
    write_csv(output_dir / "final_human_reviewed_all_articles.csv", rows)
    write_csv(output_dir / "final_relevant_articles.csv", relevant)
    write_csv(output_dir / "final_irrelevant_articles.csv", irrelevant)
    write_csv(output_dir / "final_unresolved_articles.csv", unresolved)
    write_jsonl(output_dir / "final_relevant_articles.jsonl", relevant)
    (output_dir / "final_review_summary.json").write_text(
        json.dumps(summary, indent=2),
        encoding="utf-8",
    )
    write_workbook(output_dir / "final_human_reviewed_corpus.xlsx", rows, summary)
    print(json.dumps(summary, indent=2))
    print(f"Workbook written: {(output_dir / 'final_human_reviewed_corpus.xlsx').resolve()}")


if __name__ == "__main__":
    main()
