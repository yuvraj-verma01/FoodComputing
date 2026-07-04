"""Metadata rule + local LLM review for ghee Round 1 discovered URLs.

This is the pre-human-review triage stage for ghee. It reads discovered URLs
and MediaCloud titles only; it does not crawl article text and it does not
dedupe against the edible-oil corpus.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sqlite3
import sys
from collections import Counter
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

RUN_DIR = ROOT / "data" / "runs" / "ghee_adulteration_round_01_2026-06-30"
OUTPUT_DIR = RUN_DIR / "mediacloud" / "outputs" / "ghee_relevance"
DB_PATH = RUN_DIR / "mediacloud" / "outputs" / "articles.db"
QUERY_PLAN = RUN_DIR / "proposed_mediacloud_ghee_round1_seed_queries.csv"

GHEE_TERMS = [
    "suspected adulterated ghee",
    "adulterated ghee",
    "fake cow ghee",
    "fake ghee",
    "cow ghee",
    "desi ghee",
    "pure ghee",
    "loose ghee",
    "vegetable ghee",
    "ghee racket busted",
    "ghee racket",
    "ghee",
]

SIGNAL_TERMS = [
    "adulterated",
    "fake",
    "seized",
    "raid",
    "food safety",
    "fssai",
    "fda",
    "vanaspati",
]

INCIDENT_PHRASES = [
    "adulterated ghee",
    "fake ghee",
    "fake cow ghee",
    "suspected adulterated ghee",
    "ghee racket",
    "ghee racket busted",
]

REVIEW_COLUMNS = [
    "keep",
    "llm_label",
    "llm_score",
    "review_priority",
    "metadata_label",
    "title",
    "url",
    "source",
    "date",
    "llm_reason",
    "llm_evidence",
    "rule_reason",
    "ghee_hits",
    "signal_hits",
    "query_family",
    "query_id",
    "query_used",
]


@dataclass
class MetadataDecision:
    metadata_label: str
    metadata_score: int
    review_priority: str
    rule_reason: str
    ghee_hits: list[str]
    signal_hits: list[str]

    def to_row(self) -> dict[str, str]:
        data = asdict(self)
        data["ghee_hits"] = "; ".join(self.ghee_hits)
        data["signal_hits"] = "; ".join(self.signal_hits)
        return data


def main() -> int:
    args = parse_args()
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    checkpoint_path = OUTPUT_DIR / "metadata_llm_checkpoint.jsonl"

    query_lookup = load_query_plan(QUERY_PLAN)
    records = load_discovered(DB_PATH)
    rows = build_metadata_rows(records, query_lookup)

    write_csv(OUTPUT_DIR / "metadata_all_articles_review.csv", rows, list(rows[0].keys()))
    write_summary(OUTPUT_DIR / "metadata_summary.json", rows, extra={"llm_model": args.model})

    llm_input = [row for row in rows if row["llm_queue"] == "1"]
    if args.limit:
        llm_input = llm_input[: args.limit]

    scored = score_with_llm(
        llm_input,
        model=args.model,
        timeout=args.timeout_seconds,
        checkpoint_path=checkpoint_path,
    )
    write_csv(OUTPUT_DIR / "metadata_llm_scored_articles.csv", scored, list(scored[0].keys()) if scored else [])
    write_review_workbook(OUTPUT_DIR / "ghee_round1_llm_review.xlsx", scored, rows)
    write_summary(
        OUTPUT_DIR / "metadata_llm_summary.json",
        scored,
        extra={
            "all_metadata_rows": len(rows),
            "llm_input_rows": len(llm_input),
            "checkpoint_path": str(checkpoint_path),
            "llm_model": args.model,
            "note": "Metadata-only LLM triage; not final article relevance.",
        },
    )
    print(f"metadata rows: {len(rows)}")
    print(f"llm scored rows: {len(scored)}")
    print(f"review workbook: {OUTPUT_DIR / 'ghee_round1_llm_review.xlsx'}")
    return 0


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--model", default="llama3.1:8b-instruct-q4_K_M")
    parser.add_argument("--timeout-seconds", type=int, default=240)
    parser.add_argument("--limit", type=int, default=0)
    return parser.parse_args()


def load_query_plan(path: Path) -> dict[str, dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return {row["query"]: row for row in csv.DictReader(handle)}


def load_discovered(db_path: Path) -> list[dict[str, str]]:
    with sqlite3.connect(db_path) as con:
        con.row_factory = sqlite3.Row
        rows = con.execute(
            """
            SELECT id, url, query_used, title_snippet, source, domain, published_date, status
            FROM discovered_urls
            ORDER BY id ASC
            """
        ).fetchall()
    return [dict(row) for row in rows]


def build_metadata_rows(
    records: list[dict[str, str]],
    query_lookup: dict[str, dict[str, str]],
) -> list[dict[str, str]]:
    rows = []
    for record in records:
        query = record.get("query_used") or ""
        query_row = query_lookup.get(query, {})
        decision = classify_metadata(
            title=record.get("title_snippet") or "",
            url=record.get("url") or "",
            query_family=query_row.get("query_family", ""),
            query_id=query_row.get("query_id", ""),
            query_used=query,
        )
        row = {
            "keep": "",
            "article_id": str(record.get("id") or ""),
            "title": record.get("title_snippet") or "",
            "url": record.get("url") or "",
            "source": record.get("source") or "",
            "domain": record.get("domain") or "",
            "date": record.get("published_date") or "",
            "status": record.get("status") or "",
            "query_family": query_row.get("query_family", "not_in_plan"),
            "query_id": query_row.get("query_id", ""),
            "query_used": query,
            **decision.to_row(),
        }
        row["llm_queue"] = "1" if row["metadata_label"] in {"likely_relevant", "manual_review"} else "0"
        rows.append(row)

    rows.sort(
        key=lambda row: (
            {"likely_relevant": 0, "manual_review": 1, "metadata_drop": 2}.get(row["metadata_label"], 9),
            -int(row["metadata_score"]),
            row["date"],
        )
    )
    return rows


def classify_metadata(title: str, url: str, query_family: str, query_id: str, query_used: str) -> MetadataDecision:
    text = clean_text(" ".join([title, url]))
    query_text = clean_text(" ".join([query_id, query_used]))
    ghee_hits = matching_terms(text, GHEE_TERMS)
    signal_hits = matching_terms(text, SIGNAL_TERMS)
    incident_hits = matching_terms(text, INCIDENT_PHRASES)
    score = 0
    reasons = []

    if incident_hits:
        score += 60
        reasons.append("approved incident phrase in title/url")
    elif ghee_hits:
        score += 35
        reasons.append("ghee product term in title/url")
    elif "ghee" in query_text:
        score += 10
        reasons.append("ghee appears in query but not title/url")

    if signal_hits:
        score += 35
        reasons.append("approved adulteration/enforcement signal in title/url")
    elif any(term in query_text for term in ("adulterated", "fake", "seized", "raid", "fssai", "fda", "food safety")):
        score += 12
        reasons.append("signal appears in query but not title/url")

    score += {"title_only": 18, "phrase": 14, "boolean": 8, "proximity": 5}.get(query_family, 0)

    close = proximity_evidence(text, GHEE_TERMS, SIGNAL_TERMS, window=80)
    if close:
        score += 20
        reasons.append("ghee term appears close to signal term")

    score = min(score, 100)
    if score >= 75 and (ghee_hits or incident_hits) and (signal_hits or incident_hits):
        label = "likely_relevant"
        priority = "high"
    elif score >= 45:
        label = "manual_review"
        priority = "medium"
    else:
        label = "metadata_drop"
        priority = "low"

    return MetadataDecision(
        metadata_label=label,
        metadata_score=score,
        review_priority=priority,
        rule_reason="; ".join(reasons) or "weak metadata evidence",
        ghee_hits=ghee_hits,
        signal_hits=signal_hits,
    )


def score_with_llm(
    rows: list[dict[str, str]],
    model: str,
    timeout: int,
    checkpoint_path: Path,
) -> list[dict[str, Any]]:
    checkpointed = load_checkpoint(checkpoint_path)
    scored = list(checkpointed.values())
    total = len(rows)
    remaining = [row for row in rows if row.get("url", "") not in checkpointed]
    if checkpointed:
        print(
            f"Loaded checkpoint rows: {len(checkpointed)}; remaining: {len(remaining)}",
            flush=True,
        )
    for index, row in enumerate(rows, start=1):
        url = row.get("url", "")
        if url in checkpointed:
            continue
        print(f"[LLM {index}/{total}] {safe_console(row.get('title', '')[:100])}", flush=True)
        try:
            llm = call_ollama(row, model=model, timeout=timeout)
        except Exception as exc:
            llm = {
                "llm_label": "unclear",
                "llm_score": 50,
                "llm_confidence": 0.0,
                "llm_reason": f"LLM call failed: {exc}",
                "llm_evidence": "",
                "llm_risk_flags": "llm_error",
                "llm_raw": "",
            }
        scored_row = {"keep": "", **llm, **row}
        append_checkpoint(checkpoint_path, scored_row)
        scored.append(scored_row)
    scored.sort(
        key=lambda row: (
            {"review": 0, "unclear": 1, "drop": 2}.get(str(row.get("llm_label", "")), 9),
            -int(row.get("llm_score") or 0),
            {"high": 0, "medium": 1, "low": 2}.get(row.get("review_priority", ""), 9),
        )
    )
    return scored


def load_checkpoint(path: Path) -> dict[str, dict[str, Any]]:
    if not path.exists():
        return {}
    rows: dict[str, dict[str, Any]] = {}
    with path.open("r", encoding="utf-8") as handle:
        for line in handle:
            if not line.strip():
                continue
            try:
                row = json.loads(line)
            except json.JSONDecodeError:
                continue
            url = row.get("url", "")
            if url:
                rows[url] = row
    return rows


def append_checkpoint(path: Path, row: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a", encoding="utf-8") as handle:
        handle.write(json.dumps(row, ensure_ascii=False) + "\n")


def safe_console(text: str) -> str:
    return str(text).encode("ascii", errors="replace").decode("ascii")


def call_ollama(row: dict[str, str], model: str, timeout: int) -> dict[str, Any]:
    response = requests.post(
        "http://127.0.0.1:11434/api/generate",
        json={
            "model": model,
            "prompt": build_prompt(row),
            "stream": False,
            "format": "json",
            "options": {"temperature": 0},
        },
        timeout=timeout,
    )
    response.raise_for_status()
    raw = response.json().get("response") or "{}"
    parsed = parse_json_object(raw)
    label = str(parsed.get("label") or "unclear").lower().strip()
    if label not in {"review", "drop", "unclear"}:
        label = "unclear"
    score = safe_int(parsed.get("score"), default=50)
    confidence = safe_float(parsed.get("confidence"), default=0.0)
    flags = parsed.get("risk_flags") or []
    if isinstance(flags, list):
        flags_text = "; ".join(str(flag) for flag in flags)
    else:
        flags_text = str(flags)
    return {
        "llm_label": label,
        "llm_score": max(0, min(100, score)),
        "llm_confidence": max(0.0, min(1.0, confidence)),
        "llm_reason": str(parsed.get("reason") or "")[:600],
        "llm_evidence": str(parsed.get("metadata_evidence") or parsed.get("evidence") or "")[:400],
        "llm_risk_flags": flags_text[:400],
        "llm_raw": raw[:2000],
    }


def build_prompt(row: dict[str, str]) -> str:
    return f"""
You are triaging metadata-only candidates for an Indian ghee adulteration news corpus.

Important: you only have title/source/date/url/query metadata, not full article text.
Classify whether this URL is worth human review for the target corpus.

Target corpus:
- Keep/review if likely about ghee itself being adulterated, fake, suspected adulterated, seized, raided, part of a ghee racket, or investigated by food-safety/FSSAI/FDA enforcement in India.
- Keep/review if likely about ghee adulteration in Tirupati/Tirumala laddu/prasadam, because the target item is ghee.
- Keep/review if vanaspati or vegetable ghee appears as ghee adulteration context.

Drop if likely:
- generic ghee health, purity tips, recipe, lifestyle, beauty, or religious-use story without likely adulteration/enforcement incident
- general politics or temple story where ghee/adulteration is not likely the article focus
- non-India story
- general food safety story without likely ghee target

Scoring:
- 90-100: direct ghee adulteration/fake/seizure/raid/FSSAI/FDA/racket story
- 70-89: likely ghee adulteration/enforcement but metadata has some ambiguity
- 50-69: plausible but uncertain, worth manual review
- 0-49: probably drop

Return only JSON:
{{
  "label": "review" | "drop" | "unclear",
  "score": 0,
  "confidence": 0.0,
  "reason": "short reason",
  "metadata_evidence": "short phrase from title/source/url/query",
  "risk_flags": ["short", "flags"]
}}

Title: {row.get("title", "")}
Source: {row.get("source", "")}
Date: {row.get("date", "")}
URL: {row.get("url", "")}
Metadata rule label: {row.get("metadata_label", "")}
Rule reason: {row.get("rule_reason", "")}
Ghee hits: {row.get("ghee_hits", "")}
Signal hits: {row.get("signal_hits", "")}
Query family: {row.get("query_family", "")}
Query ID: {row.get("query_id", "")}
Query used: {row.get("query_used", "")}
""".strip()


def write_review_workbook(path: Path, scored_rows: list[dict[str, Any]], all_rows: list[dict[str, str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Review"
    write_sheet(ws, scored_rows, REVIEW_COLUMNS, readable=True)
    write_sheet(wb.create_sheet("High Priority"), [r for r in scored_rows if r.get("review_priority") == "high"], REVIEW_COLUMNS, readable=True)
    write_sheet(wb.create_sheet("Medium Priority"), [r for r in scored_rows if r.get("review_priority") == "medium"], REVIEW_COLUMNS, readable=True)
    write_sheet(wb.create_sheet("Metadata All"), all_rows, list(all_rows[0].keys()) if all_rows else [], readable=False)
    write_summary_sheet(wb.create_sheet("Summary"), scored_rows, all_rows)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_sheet(ws, rows: list[dict[str, Any]], columns: list[str], readable: bool) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    keep_fill = PatternFill("solid", fgColor="FFF2CC")
    review_fill = PatternFill("solid", fgColor="D9EAD3")
    unclear_fill = PatternFill("solid", fgColor="FFF2CC")
    drop_fill = PatternFill("solid", fgColor="FCE4D6")
    for col_idx, col in enumerate(columns, start=1):
        cell = ws.cell(1, col_idx, col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[cell.column_letter].width = column_width(col)
    validation = DataValidation(type="list", formula1='"0,1"', allow_blank=True)
    ws.add_data_validation(validation)
    for row_idx, row in enumerate(rows, start=2):
        label = row.get("llm_label") or row.get("metadata_label")
        fill = {"review": review_fill, "likely_relevant": review_fill, "unclear": unclear_fill, "manual_review": unclear_fill, "drop": drop_fill, "metadata_drop": drop_fill}.get(str(label), None)
        for col_idx, col in enumerate(columns, start=1):
            value = row.get(col, "")
            cell = ws.cell(row_idx, col_idx, value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if fill:
                cell.fill = fill
            if col == "keep":
                cell.fill = keep_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if col == "url" and value:
                cell.hyperlink = str(value)
                cell.style = "Hyperlink"
        ws.row_dimensions[row_idx].height = 58 if readable else 42
    if rows and "keep" in columns:
        validation.add(f"A2:A{len(rows) + 1}")
    ws.freeze_panes = "D2" if readable else "A2"
    ws.auto_filter.ref = ws.dimensions


def write_summary_sheet(ws, scored_rows: list[dict[str, Any]], all_rows: list[dict[str, str]]) -> None:
    ws["A1"] = "Ghee Round 1 Metadata + LLM Review"
    ws["A1"].font = Font(bold=True, size=14)
    summary_items = [
        ("all discovered URLs", len(all_rows)),
        ("LLM scored rows", len(scored_rows)),
        ("cross-corpus oil dedupe", "not applied"),
        ("article crawl", "not performed"),
    ]
    for idx, (key, value) in enumerate(summary_items, start=3):
        ws.cell(idx, 1, key).font = Font(bold=True)
        ws.cell(idx, 2, value)
    row_num = 9
    for title, counter in [
        ("metadata labels", Counter(row.get("metadata_label", "") for row in all_rows)),
        ("llm labels", Counter(row.get("llm_label", "") for row in scored_rows)),
        ("review priorities", Counter(row.get("review_priority", "") for row in scored_rows)),
    ]:
        ws.cell(row_num, 1, title).font = Font(bold=True)
        row_num += 1
        for key, count in counter.most_common():
            ws.cell(row_num, 1, key)
            ws.cell(row_num, 2, count)
            row_num += 1
        row_num += 2
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 80


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def write_summary(path: Path, rows: list[dict[str, Any]], extra: dict[str, Any]) -> None:
    summary = {
        "created_at": datetime.now(timezone.utc).isoformat(),
        "rows": len(rows),
        "metadata_label_counts": dict(Counter(row.get("metadata_label", "") for row in rows)),
        "llm_label_counts": dict(Counter(row.get("llm_label", "") for row in rows)),
        "review_priority_counts": dict(Counter(row.get("review_priority", "") for row in rows)),
        **extra,
    }
    path.write_text(json.dumps(summary, indent=2, ensure_ascii=False), encoding="utf-8")


def clean_text(text: str) -> str:
    text = re.sub(r"[^a-z0-9]+", " ", (text or "").lower())
    return re.sub(r"\s+", " ", text).strip()


def matching_terms(text: str, terms: list[str]) -> list[str]:
    out = []
    padded = f" {clean_text(text)} "
    for term in terms:
        if f" {clean_text(term)} " in padded:
            out.append(term)
    return out


def proximity_evidence(text: str, left_terms: list[str], right_terms: list[str], window: int = 80) -> bool:
    lowered = clean_text(text)
    for left in left_terms:
        left_norm = clean_text(left)
        for match in re.finditer(re.escape(left_norm), lowered):
            snippet = lowered[max(0, match.start() - window) : match.end() + window]
            for right in right_terms:
                if clean_text(right) in snippet:
                    return True
    return False


def parse_json_object(raw: str) -> dict[str, Any]:
    try:
        parsed = json.loads(raw)
        return parsed if isinstance(parsed, dict) else {}
    except json.JSONDecodeError:
        match = re.search(r"\{.*\}", raw, flags=re.S)
        if not match:
            return {}
        try:
            parsed = json.loads(match.group(0))
            return parsed if isinstance(parsed, dict) else {}
        except json.JSONDecodeError:
            return {}


def safe_int(value: Any, default: int) -> int:
    try:
        return int(round(float(value)))
    except (TypeError, ValueError):
        return default


def safe_float(value: Any, default: float) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def column_width(col: str) -> int:
    return {
        "keep": 7,
        "llm_label": 12,
        "llm_score": 10,
        "review_priority": 15,
        "metadata_label": 18,
        "title": 72,
        "url": 58,
        "source": 24,
        "date": 13,
        "llm_reason": 52,
        "llm_evidence": 42,
        "rule_reason": 44,
        "ghee_hits": 28,
        "signal_hits": 24,
        "query_family": 14,
        "query_id": 34,
        "query_used": 58,
    }.get(col, 20)


if __name__ == "__main__":
    raise SystemExit(main())
