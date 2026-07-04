"""Build the single Round 3 keyword review workbook.

Merges ALL candidate sources into one file with two sheets:
  Sheet 1 — Positive Keywords  (9 rescue + new unique Round-2 extractions, deduped vs Round-1 bank)
  Sheet 2 — NOT Terms          (candidates from unique irrelevant Round-2 articles)
  Sheet 3 — Instructions

Also deletes superseded intermediate xlsx files from earlier partial runs.
"""
from __future__ import annotations

import csv
import json
import re
import sys
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent.parent

# ── Input paths ────────────────────────────────────────────────────────────────
KW_DIR   = ROOT / "reports/edible_oil_adulteration_round_02/round_03_keyword_review"
R1_BANK  = ROOT / "reports/edible_oil_adulteration_round_01/keyword_review/round_01_positive_keyword_bank.csv"
R1_R2_SEEDS = [
    ROOT / "reports/edible_oil_adulteration_round_01/keyword_review/round_01_positive_keyword_bank.csv",
]

RESCUE_CSV  = KW_DIR / "round_03_from_round_02_rescue_keyword_candidates_clean.csv"
NEW160_CSV  = KW_DIR / "round_03_positive_keyword_candidates.csv"
NOT_CSV     = KW_DIR / "round_03_not_term_candidates.csv"

OUTPUT_XLSX = KW_DIR / "round_03_keyword_review.xlsx"

STALE_FILES = [
    KW_DIR / "round_03_from_round_02_rescue_keyword_review.xlsx",
    KW_DIR / "round_03_from_round_02_rescue_keyword_review_clean.xlsx",
    KW_DIR / "round_03_positive_keyword_review.xlsx",
    KW_DIR / "round_03_not_term_review.xlsx",
]


# ── Helpers ────────────────────────────────────────────────────────────────────

def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def read_csv(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    if not rows:
        return
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()), extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def norm(term: str) -> str:
    return re.sub(r"\s+", " ", (term or "").lower().strip())


def thin_border() -> Border:
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def add_dv(ws, n_rows: int) -> None:
    dv = DataValidation(type="list", formula1='"0,1"', allow_blank=True,
                        showErrorMessage=True, errorTitle="Use 0 or 1",
                        error="Enter 1 to keep or 0 to drop.")
    ws.add_data_validation(dv)
    if n_rows:
        dv.add(f"A2:A{n_rows + 1}")
        ws.conditional_formatting.add(f"A2:A{n_rows+1}",
            CellIsRule("equal", ["1"], fill=PatternFill("solid", fgColor="C6EFCE")))
        ws.conditional_formatting.add(f"A2:A{n_rows+1}",
            CellIsRule("equal", ["0"], fill=PatternFill("solid", fgColor="F4CCCC")))


# ── Load & deduplicate positive candidates ─────────────────────────────────────

def load_r1_bank_norms() -> set[str]:
    norms: set[str] = set()
    for row in read_csv(R1_BANK):
        t = row.get("term") or row.get("keyword_or_keyphrase") or ""
        if t:
            norms.add(norm(t))
    return norms


def merge_positive_candidates(r1_norms: set[str]) -> list[dict[str, Any]]:
    seen_norms: set[str] = set(r1_norms)
    merged: list[dict[str, Any]] = []

    # ── 1. Rescue candidates (9, manually curated) ────────────────────────────
    for row in read_csv(RESCUE_CSV):
        kw = row.get("keyword_or_keyphrase") or row.get("term") or ""
        n = norm(kw)
        if not n or n in seen_norms:
            continue
        seen_norms.add(n)
        merged.append({
            "keep": "",
            "keyword_or_keyphrase": kw,
            "review_label": row.get("review_label") or row.get("original_review_label") or "manual_review",
            "source": "rescue_manual",
            "category": row.get("category") or "",
            "composite_score": row.get("composite_score") or "",
            "methods_found": row.get("methods_found") or "manual_article_read",
            "method_count": row.get("method_count") or "1",
            "total_frequency": row.get("total_frequency") or "",
            "document_frequency": row.get("document_frequency") or "",
            "review_reason": row.get("review_reason") or row.get("reason") or "",
            "example_article_title": row.get("example_article_title") or "",
            "example_context": row.get("example_context") or "",
            "example_url": row.get("example_url") or "",
            "source_round_number": "2",
            "next_round_target": "3",
        })

    # ── 2. New extraction from unique Round 2 relevant articles (160) ─────────
    for row in read_csv(NEW160_CSV):
        kw = row.get("keyword_or_keyphrase") or ""
        n = norm(kw)
        if not n or n in seen_norms:
            continue
        seen_norms.add(n)
        merged.append({
            "keep": "",
            "keyword_or_keyphrase": kw,
            "review_label": row.get("review_label") or "manual_review",
            "source": "extraction_unique_r2_relevant",
            "category": row.get("category") or "",
            "composite_score": row.get("composite_score") or "",
            "methods_found": row.get("methods_found") or "",
            "method_count": row.get("method_count") or "",
            "total_frequency": row.get("total_frequency") or "",
            "document_frequency": row.get("document_frequency") or "",
            "review_reason": row.get("review_reason") or "",
            "example_article_title": row.get("example_article_title") or "",
            "example_context": row.get("example_context") or "",
            "example_url": row.get("example_url") or "",
            "source_round_number": "2",
            "next_round_target": "3",
        })

    # Sort: keep_core first, then manual_review, then drop; within each by score desc
    label_order = {"keep_core": 0, "manual_review": 1, "drop": 2}
    merged.sort(key=lambda r: (
        label_order.get(r.get("review_label", ""), 3),
        -float(r.get("composite_score") or 0),
    ))
    return merged


# ── Positive keyword sheet ─────────────────────────────────────────────────────

POS_COLS = [
    ("keep",                  "keep",           9),
    ("keyword_or_keyphrase",  "keyword",        30),
    ("review_label",          "suggested",      16),
    ("source",                "source",         22),
    ("category",              "category",       22),
    ("composite_score",       "score",          10),
    ("methods_found",         "methods",        28),
    ("total_frequency",       "freq",           8),
    ("document_frequency",    "doc_freq",       9),
    ("review_reason",         "reason",         42),
    ("example_article_title", "example_title",  50),
    ("example_context",       "example_context",75),
    ("example_url",           "url",            55),
]

LABEL_FILLS = {
    "keep_core":     PatternFill("solid", fgColor="D9EAD3"),
    "manual_review": PatternFill("solid", fgColor="FFF2CC"),
    "drop":          PatternFill("solid", fgColor="E7E6E6"),
}


def write_positive_sheet(ws, rows: list[dict[str, Any]]) -> None:
    thin = thin_border()
    hfill = PatternFill("solid", fgColor="1F4E78")
    hfont = Font(color="FFFFFF", bold=True, size=10)
    keep_fill = PatternFill("solid", fgColor="FFF2CC")

    ws.row_dimensions[1].height = 30
    for ci, (_k, header, width) in enumerate(POS_COLS, start=1):
        c = ws.cell(row=1, column=ci, value=header)
        c.fill = hfill; c.font = hfont; c.border = thin
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = width

    ws["A1"].comment = Comment(
        "1 = include in Round 3 query set\n0 = drop\nBlank = not reviewed yet", "System")
    add_dv(ws, len(rows))

    for ri, row in enumerate(rows, start=2):
        lbl = row.get("review_label", "")
        rfill = LABEL_FILLS.get(lbl, PatternFill("solid", fgColor="FFFFFF"))
        ws.row_dimensions[ri].height = 58
        for ci, (key, _, _w) in enumerate(POS_COLS, start=1):
            val = "" if key == "keep" else row.get(key, "")
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = thin
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.font = Font(size=10)
            c.fill = keep_fill if key == "keep" else rfill

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions


# ── NOT term sheet ─────────────────────────────────────────────────────────────

NOT_COLS = [
    ("keep",                         "keep_as_NOT",    12),
    ("not_candidate",                "candidate",      30),
    ("risk_flag",                    "risk_flag",      34),
    ("irrelevant_category",          "category",       22),
    ("irrelevant_composite_score",   "score",          10),
    ("irrelevant_document_frequency","doc_freq",       10),
    ("irrelevant_total_frequency",   "freq",           8),
    ("positive_exact_kept",          "exact_match",    13),
    ("positive_overlap_terms",       "overlap_with",   38),
    ("example_irrelevant_title",     "example_title",  55),
    ("example_context",              "example_context",75),
]

RISK_FILLS = {
    "possible_not_candidate":              PatternFill("solid", fgColor="E2F0D9"),
    "risky_domain_term_may_drop_relevant": PatternFill("solid", fgColor="FFF2CC"),
    "risky_overlap_with_positive_keyword": PatternFill("solid", fgColor="FCE4D6"),
    "do_not_not_exact_positive_kept":      PatternFill("solid", fgColor="F4CCCC"),
}


def write_not_sheet(ws, rows: list[dict[str, str]]) -> None:
    thin = thin_border()
    hfill = PatternFill("solid", fgColor="1F4E78")
    hfont = Font(color="FFFFFF", bold=True, size=10)
    keep_fill = PatternFill("solid", fgColor="FFF2CC")

    ws.row_dimensions[1].height = 30
    for ci, (_k, header, width) in enumerate(NOT_COLS, start=1):
        c = ws.cell(row=1, column=ci, value=header)
        c.fill = hfill; c.font = hfont; c.border = thin
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = width

    ws["A1"].comment = Comment(
        "1 = exclude this term from Round 3 queries (NOT operator)\n"
        "0 = do NOT exclude (appears in relevant articles too)\n"
        "RED rows = exact positive keyword match — NEVER mark 1", "System")
    add_dv(ws, len(rows))

    for ri, row in enumerate(rows, start=2):
        risk = row.get("risk_flag", "")
        rfill = RISK_FILLS.get(risk, PatternFill("solid", fgColor="FFFFFF"))
        ws.row_dimensions[ri].height = 55
        for ci, (key, _, _w) in enumerate(NOT_COLS, start=1):
            val = "" if key == "keep" else row.get(key, "")
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = thin
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.font = Font(size=10)
            c.fill = keep_fill if key == "keep" else rfill

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions


# ── Instructions sheet ─────────────────────────────────────────────────────────

def write_instructions(ws) -> None:
    ws["A1"] = "Round 3 Keyword Review — Instructions"
    ws["A1"].font = Font(bold=True, size=14)

    pos_lines = [
        ("POSITIVE KEYWORDS (Sheet 1)", ""),
        ("keep = 1", "Include this keyword in Round 3 MediaCloud queries."),
        ("keep = 0", "Drop — too noisy, too broad, or already covered."),
        ("keep_core (green rows)", "High-confidence keeper. Still confirm with keep=1."),
        ("manual_review (yellow rows)", "Needs your judgment. Read example_context."),
        ("drop (grey rows)", "Auto-suggested drop. Override with 1 if you disagree."),
        ("source = rescue_manual", "Manually found from rescue articles. Higher quality signal."),
        ("source = extraction_unique_r2", "Statistically extracted from unique Round 2 articles."),
    ]
    not_lines = [
        ("NOT TERMS (Sheet 2)", ""),
        ("keep = 1", "Add as NOT term — exclude articles mentioning this."),
        ("keep = 0", "Do not exclude — this term can appear in relevant articles."),
        ("GREEN rows", "possible_not_candidate — relatively safe to exclude."),
        ("YELLOW rows", "risky_domain — contains oil/safety signal; be careful."),
        ("ORANGE rows", "Overlaps with a positive keyword — very risky; usually mark 0."),
        ("RED rows", "Exact positive keyword match — NEVER mark 1."),
        ("Rule of thumb", "A missing NOT term is much better than losing a relevant article."),
    ]

    row = 3
    for label, text in pos_lines + [("", "")] + not_lines:
        ws.cell(row, 1, label).font = Font(bold=True if label else False)
        ws.cell(row, 2, text).alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 90
    ws.row_dimensions[1].height = 22


# ── Main ───────────────────────────────────────────────────────────────────────

def main() -> None:
    r1_norms = load_r1_bank_norms()
    print(f"Round 1 bank keywords (dedup reference): {len(r1_norms)}", flush=True)

    positives = merge_positive_candidates(r1_norms)
    not_rows  = read_csv(NOT_CSV)

    # Extra dedup pass: remove any positive candidate that slipped through vs R1 bank
    positives = [r for r in positives if norm(r.get("keyword_or_keyphrase","")) not in r1_norms]

    print(f"Positive candidates merged: {len(positives)}", flush=True)
    print(f"  rescue_manual:              {sum(1 for r in positives if r['source']=='rescue_manual')}")
    print(f"  extraction_unique_r2:       {sum(1 for r in positives if r['source']=='extraction_unique_r2_relevant')}")
    print(f"NOT-term candidates:         {len(not_rows)}", flush=True)

    wb = Workbook()
    ws_pos = wb.active
    ws_pos.title = "Positive Keywords"
    write_positive_sheet(ws_pos, positives)

    ws_not = wb.create_sheet("NOT Terms")
    write_not_sheet(ws_not, not_rows)

    ws_inst = wb.create_sheet("Instructions")
    write_instructions(ws_inst)

    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_XLSX)
    print(f"\nSaved: {OUTPUT_XLSX}", flush=True)

    # Write merged CSV for audit
    merged_csv = KW_DIR / "round_03_positive_keyword_candidates_merged.csv"
    write_csv(merged_csv, positives)

    # Remove stale xlsx files
    removed = []
    for p in STALE_FILES:
        if p.exists():
            p.unlink()
            removed.append(p.name)
    if removed:
        print(f"Removed stale files: {removed}", flush=True)

    summary = {
        "created_at": utc_now(),
        "r1_bank_dedup_reference_count": len(r1_norms),
        "positive_candidates_total": len(positives),
        "positive_source_counts": dict(Counter(r["source"] for r in positives)),
        "positive_label_counts": dict(Counter(r.get("review_label","") for r in positives)),
        "not_term_candidates_total": len(not_rows),
        "not_term_risk_counts": dict(Counter(r.get("risk_flag","") for r in not_rows)),
        "output_xlsx": str(OUTPUT_XLSX),
        "stale_files_removed": removed,
    }
    (KW_DIR / "round_03_keyword_review_summary.json").write_text(
        json.dumps(summary, indent=2, ensure_ascii=False), encoding="utf-8"
    )
    print("\n=== SUMMARY ===")
    print(json.dumps(summary, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
