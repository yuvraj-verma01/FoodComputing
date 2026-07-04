"""Quick audit of full corpus counts across all rounds."""
import csv
import sys
import openpyxl
from collections import Counter
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parent.parent
runs = ROOT / "data" / "runs"
reports = ROOT / "reports"
seeds = ROOT / "data" / "seeds" / "seed_urls.csv"


def read_csv(p):
    if not p.exists():
        return []
    with p.open(encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


# --- Round 0 (seeds) ---
seed_rows = read_csv(seeds)

# --- Round 1 ---
r1_all = read_csv(reports / "edible_oil_adulteration_round_01" / "round_01_all_articles.csv")
r1_rel = read_csv(reports / "edible_oil_adulteration_round_01" / "round_01_relevant_articles.csv")

# --- Round 2 ---
r2_run = runs / "edible_oil_adulteration_round_02_2026-06-23" / "mediacloud" / "outputs"
r2_disc = read_csv(r2_run / "discovery_url_review.csv")
r2_all  = read_csv(r2_run / "oil_relevance" / "all_articles_review.csv")
r2_final = read_csv(r2_run / "oil_relevance" / "final_human_reviewed_all_articles.csv")

# --- Round 3 ---
r3_run = runs / "edible_oil_adulteration_round_03_2026-06-23" / "mediacloud" / "outputs"
r3_disc = read_csv(r3_run / "discovery_url_review.csv")
r3_meta = read_csv(r3_run / "oil_relevance" / "metadata_all_articles_review.csv")
r3_all  = read_csv(r3_run / "oil_relevance" / "all_articles_review.csv")

# Round 3 human review from xlsx
xlsx = r3_run / "oil_relevance" / "round_03_article_review.xlsx"
wb = openpyxl.load_workbook(xlsx, data_only=True)
ws = wb["Round 3 Review"]
headers = [c.value for c in ws[1]]
r3_review = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if any(v is not None for v in row):
        r3_review.append(dict(zip(headers, row)))
r3_kept = [r for r in r3_review if str(r.get("keep", "")) == "1"]
r3_dropped = [r for r in r3_review if str(r.get("keep", "")) == "0"]

# --- Master corpus (pre-R3) ---
master_all = read_csv(reports / "master_corpus" / "master_all_articles.csv")
master_rel = read_csv(reports / "master_corpus" / "master_relevant_articles.csv")

print("=" * 65)
print("  CORPUS AUDIT — Rounds 0 through 3")
print("=" * 65)

print("\n--- 1. URLs DISCOVERED (MediaCloud) ---")
# Round 1 doesn't have a separate discovery CSV; all_articles = discovered+classified
print(f"  Round 0 (seed articles)          :  {len(seed_rows):>5}")
print(f"  Round 1 (seed queries → MC)      :  {len(r1_all):>5}  (all went through classifier)")
print(f"  Round 2 (MC discovery)           :  {len(r2_disc):>5}")
print(f"  Round 3 (MC discovery, deduped)  :  {len(r3_meta):>5}  (after −122 already-reviewed)")
total_disc = len(seed_rows) + len(r1_all) + len(r2_disc) + len(r3_meta)
print(f"  TOTAL                            :  {total_disc:>5}")

print("\n--- 2. ARTICLES CRAWLED & CLASSIFIED ---")
r2_crawled = len(r2_all)
r3_crawled = len(r3_all)
print(f"  Round 1 crawled+classified       :  {len(r1_all):>5}")
print(f"  Round 2 crawled+classified       :  {r2_crawled:>5}")
print(f"  Round 3 crawled+classified       :  {r3_crawled:>5}")
total_crawled = len(r1_all) + r2_crawled + r3_crawled
print(f"  TOTAL                            :  {total_crawled:>5}")
print(f"  (Round 3: {len(r3_meta)-r3_crawled} of {len(r3_meta)} new URLs never crawled — metadata dropped)")

print("\n--- 3. HUMAN REVIEWED (marked 1/0) ---")
# Master pre-R3
master_human = len([r for r in master_all if r.get("human_review_status") in
                    ("human_reviewed", "conflict_resolved", "metadata_rescue_fulltext_classified")])
master_kept_pre = len([r for r in master_all if str(r.get("final_keep")) == "1"])
master_dropped_pre = len([r for r in master_all if str(r.get("final_keep")) == "0"])
print(f"  Rounds 1+2 in master corpus      :  {len(master_all):>5}  ({master_kept_pre} kept, {master_dropped_pre} dropped)")
print(f"  Round 3 human review (xlsx)      :  {len(r3_review):>5}  ({len(r3_kept)} kept, {len(r3_dropped)} dropped)")
total_reviewed = len(master_all) + len(r3_review)
total_kept = master_kept_pre + len(r3_kept)
print(f"  TOTAL reviewed                   :  {total_reviewed:>5}")
print(f"  TOTAL kept (relevant)            :  {total_kept:>5}")

print("\n--- 4. SEEDS (Round 0) to add ---")
for r in seed_rows:
    title = r.get("title_snippet", "")[:65]
    print(f"    {title}")

print("\n--- MASTER CORPUS (pre-R3 merge) ---")
by_round = Counter(r.get("round_number") for r in master_all)
print(f"  master_all_articles.csv          :  {len(master_all):>5}  by round: {dict(by_round)}")
print(f"  master_relevant_articles.csv     :  {len(master_rel):>5}")

print("\n--- AFTER R3 + SEED MERGE (projected) ---")
proj_total = len(master_all) + len(r3_review) + len(seed_rows)
proj_rel   = master_kept_pre + len(r3_kept) + len(seed_rows)
print(f"  Projected total                  :  {proj_total:>5}")
print(f"  Projected relevant               :  {proj_rel:>5}")
