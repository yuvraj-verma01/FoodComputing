"""Create a formatted keyword review workbook.

The workbook is for manual review of sample-derived keyword candidates.
Column A is intentionally blank and named ``keep``:
  1 = keep for seed-query construction
  0 = drop
"""

from __future__ import annotations

import argparse
import csv
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parent.parent
DEFAULT_RUN_DIR = ROOT / "data" / "runs" / "edible_oils_from_sample_2026-06-21"

HEADER_FILL = "1F4E79"
HEADER_FG = "FFFFFF"
KEEP_COL_FILL = "FFF2CC"
KEEP_CORE_FILL = "D9EAD3"
MANUAL_REVIEW_FILL = "FFF2CC"
DROP_FILL = "E7E6E6"
BORDER = "BFBFBF"
LINK_BLUE = "0563C1"

REVIEW_COLUMNS = [
    ("keep", "keep", 9),
    ("keyword_or_keyphrase", "keyword_or_keyphrase", 28),
    ("review_label", "current_label", 16),
    ("composite_score", "composite_score", 15),
    ("category", "category", 22),
    ("methods_found", "methods_found", 28),
    ("method_count", "method_count", 13),
    ("total_frequency", "total_frequency", 15),
    ("document_frequency", "document_frequency", 18),
    ("method_agreement_score", "method_agreement_score", 22),
    ("tfidf_score_normalized", "tfidf_score_normalized", 24),
    ("yake_score_normalized", "yake_score_normalized", 24),
    ("frequency_score_normalized", "frequency_score_normalized", 27),
    ("document_frequency_score", "document_frequency_score", 25),
    ("review_reason", "review_reason", 42),
    ("example_article_number", "example_article_number", 20),
    ("example_article_title", "example_article_title", 46),
    ("example_relevance_label", "example_relevance_label", 22),
    ("example_query_used", "example_query_used", 42),
    ("example_domain", "example_domain", 24),
    ("example_context", "example_context", 70),
    ("example_url", "example_url", 55),
]


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--input",
        default=str(DEFAULT_RUN_DIR / "sample_keyword_candidates_clean.csv"),
        help="Clean keyword candidate CSV.",
    )
    parser.add_argument(
        "--output",
        default=str(DEFAULT_RUN_DIR / "keyword_review.xlsx"),
        help="Output XLSX workbook.",
    )
    args = parser.parse_args()

    input_path = Path(args.input)
    output_path = Path(args.output)
    rows = read_rows(input_path)
    wb = Workbook()
    review_ws = wb.active
    review_ws.title = "Keyword Review"
    write_review_sheet(review_ws, rows)
    write_summary_sheet(wb.create_sheet("Summary"), rows, input_path, output_path)
    write_instructions_sheet(wb.create_sheet("Instructions"))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Wrote {output_path}")
    print(f"Rows: {len(rows)}")
    return 0


def read_rows(path: Path) -> list[dict]:
    with path.open(encoding="utf-8-sig", newline="") as fh:
        return list(csv.DictReader(fh))


def write_review_sheet(ws, rows: list[dict]) -> None:
    thin = thin_border()
    header_font = Font(name="Calibri", bold=True, color=HEADER_FG, size=11)
    header_fill = PatternFill("solid", fgColor=HEADER_FILL)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, (_key, header, width) in enumerate(REVIEW_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin
        cell.alignment = header_align
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws["A1"].comment = Comment("Enter 1 to keep this keyword, or 0 to drop it.", "Codex")
    ws.row_dimensions[1].height = 32

    data_validation = DataValidation(
        type="list",
        formula1='"0,1"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Use 0 or 1",
        error="Enter 1 to keep or 0 to drop.",
    )
    ws.add_data_validation(data_validation)

    for row_idx, row in enumerate(rows, start=2):
        label = row.get("review_label", "")
        fill = label_fill(label)
        for col_idx, (key, _header, _width) in enumerate(REVIEW_COLUMNS, start=1):
            value = "" if key == "keep" else row.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.font = Font(name="Calibri", size=10)
            if fill:
                cell.fill = fill
            if key == "keep":
                cell.fill = PatternFill("solid", fgColor=KEEP_COL_FILL)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if key == "example_url" and value:
                cell.hyperlink = value
                cell.style = "Hyperlink"
                cell.font = Font(name="Calibri", size=10, color=LINK_BLUE, underline="single")
        ws.row_dimensions[row_idx].height = 52

    if rows:
        data_validation.add(f"A2:A{len(rows) + 1}")
        ws.conditional_formatting.add(
            f"A2:A{len(rows) + 1}",
            CellIsRule(operator="equal", formula=["1"], fill=PatternFill("solid", fgColor="C6EFCE")),
        )
        ws.conditional_formatting.add(
            f"A2:A{len(rows) + 1}",
            CellIsRule(operator="equal", formula=["0"], fill=PatternFill("solid", fgColor="F4CCCC")),
        )

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_properties.tabColor = HEADER_FILL


def write_summary_sheet(ws, rows: list[dict], input_path: Path, output_path: Path) -> None:
    thin = thin_border()
    title_font = Font(name="Calibri", bold=True, size=13)
    header_font = Font(name="Calibri", bold=True, color=HEADER_FG, size=11)
    header_fill = PatternFill("solid", fgColor=HEADER_FILL)

    ws["A1"] = "Keyword Review Workbook"
    ws["A1"].font = title_font
    ws["A3"] = "Input CSV"
    ws["B3"] = str(input_path)
    ws["A4"] = "Output XLSX"
    ws["B4"] = str(output_path)
    ws["A5"] = "Manual keep coding"
    ws["B5"] = "1 = keep, 0 = drop, blank = not reviewed"

    blocks = [
        ("Current Label Counts", Counter(r.get("review_label", "") for r in rows)),
        ("Category Counts", Counter(r.get("category", "") for r in rows)),
    ]
    start_row = 7
    for title, counts in blocks:
        ws.cell(start_row, 1, title)
        ws.cell(start_row, 1).font = Font(name="Calibri", bold=True, size=12)
        start_row += 1
        ws.cell(start_row, 1, "Value")
        ws.cell(start_row, 2, "Count")
        for col in (1, 2):
            cell = ws.cell(start_row, col)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin
        start_row += 1
        for value, count in sorted(counts.items()):
            ws.cell(start_row, 1, value)
            ws.cell(start_row, 2, count)
            ws.cell(start_row, 1).border = thin
            ws.cell(start_row, 2).border = thin
            start_row += 1
        start_row += 2

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 95
    ws.freeze_panes = "A7"
    ws.sheet_properties.tabColor = "70AD47"


def write_instructions_sheet(ws) -> None:
    lines = [
        ("Task", "Review every keyword row on the Keyword Review sheet."),
        ("keep = 1", "Use 1 when the keyword should be kept for seed-query construction."),
        ("keep = 0", "Use 0 when the keyword should be dropped."),
        ("Blank", "Leave blank if you have not reviewed it yet."),
        ("Current label", "current_label is the script's suggested label: keep_core, manual_review, or drop."),
        ("Important", "The manual keep column is the authority for the next step."),
    ]
    ws["A1"] = "Instructions"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14)
    for idx, (label, text) in enumerate(lines, start=3):
        ws.cell(idx, 1, label)
        ws.cell(idx, 2, text)
        ws.cell(idx, 1).font = Font(name="Calibri", bold=True, size=11)
        ws.cell(idx, 1).alignment = Alignment(vertical="top")
        ws.cell(idx, 2).alignment = Alignment(vertical="top", wrap_text=True)
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 90
    ws.sheet_properties.tabColor = "FFC000"


def label_fill(label: str):
    if label == "keep_core":
        return PatternFill("solid", fgColor=KEEP_CORE_FILL)
    if label == "manual_review":
        return PatternFill("solid", fgColor=MANUAL_REVIEW_FILL)
    if label == "drop":
        return PatternFill("solid", fgColor=DROP_FILL)
    return None


def thin_border() -> Border:
    side = Side(style="thin", color=BORDER)
    return Border(left=side, right=side, top=side, bottom=side)


if __name__ == "__main__":
    raise SystemExit(main())
