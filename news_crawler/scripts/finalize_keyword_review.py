"""Finalize reviewed sample keywords and propose Media Cloud seed queries.

Inputs:
  - keyword_review.xlsx filled with keep=1/0 decisions

Outputs:
  - final_keyword_bank.csv/json
  - proposed_mediacloud_seed_queries.csv/json

This script does not call Media Cloud. It only turns the reviewed sample-derived
keywords into a compact, high-precision seed query proposal.
"""

from __future__ import annotations

import argparse
import csv
import json
from collections import OrderedDict
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent.parent
DEFAULT_RUN_DIR = ROOT / "data" / "runs" / "edible_oils_from_sample_2026-06-21"
DEFAULT_DATE_START = "2021-01-01"
DEFAULT_DATE_END = "2026-06-21"

REMOVE_TERMS = {
    "edible",
    "oils",
    "vegetable",
    "coconut",
    "food products",
    "officials",
    "police",
    "safety department",
    "safety officer",
    "fsda team",
    "fsda team raided",
    "fsda team seized",
    "litres oil",
    "litres oil kilograms",
    "seizing litres oil",
    "success seizing",
    "success seizing litres",
    "oil tins",
    "edible oil worth",
    "litres of edible oil",
    "litres edible oil",
    "pradesh food safety",
    "jaisalmer food safety",
    "jaisalmer food safety officer",
}

ADD_TERMS = {
    "product": [
        ("cooking oil", "manual_addition", "Common synonym found in sample title/text; useful for public-facing news queries."),
        ("olive oil", "user_keep", "Explicitly kept by user despite low sample signal."),
    ],
    "fraud": [
        ("fake", "manual_addition", "High-value fraud term; useful for phrases like fake mustard oil."),
    ],
    "enforcement": [
        ("food safety officer", "manual_addition", "Better complete phrase than fragment 'safety officer'."),
        ("seizure", "manual_addition", "Headline/query variant of seized."),
    ],
}

TERM_CATEGORY = {
    "product": {
        "edible oil",
        "cooking oil",
        "mustard oil",
        "coconut oil",
        "soybean oil",
        "vegetable oil",
        "groundnut oil",
        "palm oil",
        "olive oil",
        "ghee and oil",
        "adulterated ghee and oil",
    },
    "fraud": {
        "adulterated",
        "adulteration",
        "food adulteration",
        "adulterated food",
        "adulterated food products",
        "adulterated food items",
        "adulterated foods",
        "adulterated food business",
        "substandard",
        "fake",
    },
    "enforcement": {
        "seized",
        "seizure",
        "raid",
        "raids",
        "fssai",
        "fsda",
        "food safety",
        "food safety department",
        "food safety officer",
        "food safety officials",
        "food safety officers",
        "food safety and standards",
        "food safety designated officers",
        "commissioner of food safety",
        "food safety commission",
        "crackdown",
        "ban",
        "banned",
    },
    "support_only": {
        "oil",
        "samples",
        "seizing",
        "raided",
        "major crackdown",
        "oil manufacturing",
        "oil manufacturing unit",
        "edible oils",
    },
}


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--workbook",
        default=str(DEFAULT_RUN_DIR / "keyword_review.xlsx"),
        help="Filled keyword review workbook.",
    )
    parser.add_argument(
        "--run-dir",
        default=str(DEFAULT_RUN_DIR),
        help="Run directory for final outputs.",
    )
    parser.add_argument(
        "--date-start",
        default=DEFAULT_DATE_START,
        help="Start date for the proposed Media Cloud search window.",
    )
    parser.add_argument(
        "--date-end",
        default=DEFAULT_DATE_END,
        help="End date for the proposed Media Cloud search window.",
    )
    args = parser.parse_args()

    workbook_path = Path(args.workbook)
    run_dir = Path(args.run_dir)
    reviewed_rows = read_reviewed_rows(workbook_path)
    final_terms = build_final_terms(reviewed_rows)
    queries = build_seed_queries(date_start=args.date_start, date_end=args.date_end)

    term_csv = run_dir / "final_keyword_bank.csv"
    term_json = run_dir / "final_keyword_bank.json"
    query_csv = run_dir / "proposed_mediacloud_seed_queries.csv"
    query_json = run_dir / "proposed_mediacloud_seed_queries.json"
    summary_path = run_dir / "keyword_finalization_summary.json"

    write_csv(term_csv, final_terms)
    term_json.write_text(json.dumps(final_terms, indent=2, ensure_ascii=False), encoding="utf-8")
    write_csv(query_csv, queries)
    query_json.write_text(json.dumps(queries, indent=2, ensure_ascii=False), encoding="utf-8")

    summary = {
        "created_at": datetime.now(timezone.utc).isoformat(),
        "workbook": str(workbook_path.resolve()),
        "reviewed_keep_1_rows": sum(1 for r in reviewed_rows if r["keep"] == "1"),
        "reviewed_keep_0_rows": sum(1 for r in reviewed_rows if r["keep"] == "0"),
        "removed_after_review": sorted(REMOVE_TERMS),
        "manual_additions": [
            {"term": term, "category": category, "reason": reason}
            for category, additions in ADD_TERMS.items()
            for term, _source, reason in additions
        ],
        "final_term_count": len(final_terms),
        "final_term_counts_by_category": count_by(final_terms, "category"),
        "proposed_query_count": len(queries),
        "proposed_date_range": {
            "start": args.date_start,
            "end": args.date_end,
            "note": "Last-five-years window for this run, ending on 2026-06-21.",
        },
        "query_generation_strategy": (
            "Compact high-precision subset from reviewed terms; not all possible permutations."
        ),
        "outputs": {
            "final_keyword_bank_csv": str(term_csv),
            "final_keyword_bank_json": str(term_json),
            "proposed_seed_queries_csv": str(query_csv),
            "proposed_seed_queries_json": str(query_json),
        },
    }
    summary_path.write_text(json.dumps(summary, indent=2, ensure_ascii=False), encoding="utf-8")

    print(json.dumps(summary, indent=2, ensure_ascii=False))
    return 0


def read_reviewed_rows(path: Path) -> list[dict]:
    ws = load_workbook(path, data_only=True)["Keyword Review"]
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    idx = {header: i + 1 for i, header in enumerate(headers)}
    rows = []
    for row_num in range(2, ws.max_row + 1):
        term = ws.cell(row_num, idx["keyword_or_keyphrase"]).value
        if not term:
            continue
        keep = normalize_keep(ws.cell(row_num, idx["keep"]).value)
        rows.append(
            {
                "term": str(term).strip().lower(),
                "keep": keep,
                "excel_row": row_num,
                "current_label": ws.cell(row_num, idx["current_label"]).value or "",
                "category_original": ws.cell(row_num, idx["category"]).value or "",
                "composite_score": ws.cell(row_num, idx["composite_score"]).value or "",
                "total_frequency": ws.cell(row_num, idx["total_frequency"]).value or "",
                "document_frequency": ws.cell(row_num, idx["document_frequency"]).value or "",
                "example_article_title": ws.cell(row_num, idx["example_article_title"]).value or "",
            }
        )
    return rows


def normalize_keep(value) -> str:
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    if value in (1, "1"):
        return "1"
    if value in (0, "0"):
        return "0"
    return ""


def build_final_terms(reviewed_rows: list[dict]) -> list[dict]:
    terms: OrderedDict[str, dict] = OrderedDict()
    kept_rows = [row for row in reviewed_rows if row["keep"] == "1"]

    for row in kept_rows:
        term = row["term"]
        if term in REMOVE_TERMS:
            continue
        category = classify_term(term)
        if category == "unassigned":
            continue
        terms[term] = {
            "term": term,
            "category": category,
            "source": "review_workbook",
            "use": term_use(term, category),
            "reason": final_reason(term, category),
            "review_excel_row": row["excel_row"],
            "original_label": row["current_label"],
            "composite_score": row["composite_score"],
            "total_frequency": row["total_frequency"],
            "document_frequency": row["document_frequency"],
            "example_article_title": row["example_article_title"],
        }

    for category, additions in ADD_TERMS.items():
        for term, source, reason in additions:
            terms[term] = {
                "term": term,
                "category": category,
                "source": source,
                "use": term_use(term, category),
                "reason": reason,
                "review_excel_row": "",
                "original_label": "",
                "composite_score": "",
                "total_frequency": "",
                "document_frequency": "",
                "example_article_title": "",
            }

    return sorted(terms.values(), key=lambda r: (category_order(r["category"]), r["term"]))


def classify_term(term: str) -> str:
    for category in ("product", "fraud", "enforcement", "support_only"):
        if term in TERM_CATEGORY[category]:
            return category
    return "unassigned"


def category_order(category: str) -> int:
    return {"product": 0, "fraud": 1, "enforcement": 2, "support_only": 3}.get(category, 9)


def term_use(term: str, category: str) -> str:
    if category == "support_only":
        return "do_not_query_alone"
    if term in {"food safety", "food adulteration", "adulterated food", "oil", "samples"}:
        return "combine_only"
    if "ghee" in term:
        return "mixed_oil_ghee_only"
    return "query_component"


def final_reason(term: str, category: str) -> str:
    if category == "product":
        return "Product/oil term for combination with fraud or enforcement signals."
    if category == "fraud":
        return "Fraud/adulteration signal for combination with oil-product terms."
    if category == "enforcement":
        return "Enforcement/evidence signal for high-precision news queries."
    return "Kept only as a helper term; avoid standalone Media Cloud queries."


def build_seed_queries(date_start: str, date_end: str) -> list[dict]:
    rows = []

    def add(
        query: str,
        template_type: str,
        product_term: str = "",
        fraud_term: str = "",
        enforcement_term: str = "",
        source_terms: list[str] | None = None,
        manual_reason: str = "",
        breadth: str = "medium",
        geography_term: str = "India",
    ) -> None:
        if source_terms is None:
            source_terms = [
                term for term in (product_term, fraud_term, enforcement_term, geography_term) if term
            ]
        rows.append(
            {
                "query": query,
                "template_type": template_type,
                "product_term": product_term,
                "fraud_term": fraud_term,
                "enforcement_term": enforcement_term,
                "geography_term": geography_term,
                "date_start": date_start,
                "date_end": date_end,
                "source_terms": "; ".join(source_terms),
                "manual_reason": manual_reason,
                "breadth": breadth,
                "status": "proposed_review",
                "permutation_scope": "selected_subset_not_all_permutations",
            }
        )

    add(
        '"edible oil" adulteration India',
        "product_plus_fraud",
        product_term="edible oil",
        fraud_term="adulteration",
        manual_reason="Core product plus core fraud signal.",
    )
    add(
        '"edible oil" adulterated India',
        "product_plus_fraud",
        product_term="edible oil",
        fraud_term="adulterated",
        manual_reason="Captures headline/body variants using adjective form.",
    )
    add(
        '"edible oil" seized India',
        "product_plus_enforcement",
        product_term="edible oil",
        enforcement_term="seized",
        manual_reason="High-precision enforcement query from sample pattern.",
    )
    add(
        '"substandard edible oil" India',
        "fraud_phrase_plus_product",
        product_term="edible oil",
        fraud_term="substandard",
        manual_reason="Directly matches substandard-oil sample language.",
        breadth="narrow",
    )
    add(
        '"fake edible oil" India',
        "fraud_phrase_plus_product",
        product_term="edible oil",
        fraud_term="fake",
        manual_reason="Fraud synonym added for recall.",
        breadth="narrow",
    )
    add(
        'FSSAI "edible oil" India',
        "authority_plus_product",
        product_term="edible oil",
        enforcement_term="fssai",
        manual_reason="Authority plus product term.",
    )
    add(
        'FSDA "edible oil" seized',
        "authority_plus_product_plus_enforcement",
        product_term="edible oil",
        enforcement_term="fsda; seized",
        source_terms=["fsda", "edible oil", "seized"],
        manual_reason="State enforcement acronym plus product/action.",
        breadth="narrow",
    )
    add(
        '"food safety department" "edible oil" India',
        "agency_phrase_plus_product",
        product_term="edible oil",
        enforcement_term="food safety department",
        manual_reason="Department phrase combined with product to avoid broad drift.",
    )
    add(
        '"food safety officer" "edible oil" India',
        "agency_phrase_plus_product",
        product_term="edible oil",
        enforcement_term="food safety officer",
        manual_reason="Officer phrase added back as complete enforcement term.",
    )
    add(
        '"edible oil" raid India',
        "product_plus_enforcement",
        product_term="edible oil",
        enforcement_term="raid",
        manual_reason="Common enforcement event phrasing.",
    )
    add(
        '"edible oil" seizure India',
        "product_plus_enforcement",
        product_term="edible oil",
        enforcement_term="seizure",
        manual_reason="Noun variant of seized.",
    )

    add(
        '"cooking oil" adulteration India',
        "product_plus_fraud",
        product_term="cooking oil",
        fraud_term="adulteration",
        manual_reason="Public-facing synonym for edible oil.",
    )
    add(
        '"adulterated cooking oil" seized India',
        "product_plus_fraud_plus_enforcement",
        product_term="cooking oil",
        fraud_term="adulterated",
        enforcement_term="seized",
        manual_reason="Specific incident-style query.",
        breadth="narrow",
    )
    add(
        '"mustard oil" adulteration India',
        "product_plus_fraud",
        product_term="mustard oil",
        fraud_term="adulteration",
        manual_reason="Specific oil type from sample.",
    )
    add(
        '"fake mustard oil" India',
        "fraud_phrase_plus_product",
        product_term="mustard oil",
        fraud_term="fake",
        manual_reason="High-value fraud phrase for mustard oil.",
        breadth="narrow",
    )
    add(
        '"mustard oil" seized India',
        "product_plus_enforcement",
        product_term="mustard oil",
        enforcement_term="seized",
        manual_reason="Specific oil plus enforcement action.",
    )
    add(
        '"coconut oil" adulteration India',
        "product_plus_fraud",
        product_term="coconut oil",
        fraud_term="adulteration",
        manual_reason="Specific oil type from sample.",
    )
    add(
        '"substandard coconut oil" India',
        "fraud_phrase_plus_product",
        product_term="coconut oil",
        fraud_term="substandard",
        manual_reason="Directly captures ban/substandard style stories.",
        breadth="narrow",
    )
    add(
        '"coconut oil" banned India',
        "product_plus_enforcement",
        product_term="coconut oil",
        enforcement_term="banned",
        manual_reason="Specific enforcement outcome from sample.",
        breadth="narrow",
    )
    add(
        '"soybean oil" adulteration India',
        "product_plus_fraud",
        product_term="soybean oil",
        fraud_term="adulteration",
        manual_reason="Specific oil type from sample.",
    )
    add(
        '"vegetable oil" adulteration India',
        "product_plus_fraud",
        product_term="vegetable oil",
        fraud_term="adulteration",
        manual_reason="Specific product phrase from sample.",
    )
    add(
        '"groundnut oil" adulteration India',
        "product_plus_fraud",
        product_term="groundnut oil",
        fraud_term="adulteration",
        manual_reason="Specific oil type from sample.",
    )
    add(
        '"palm oil" adulteration India',
        "product_plus_fraud",
        product_term="palm oil",
        fraud_term="adulteration",
        manual_reason="Specific oil type from sample.",
    )
    add(
        '"olive oil" adulteration India',
        "product_plus_fraud",
        product_term="olive oil",
        fraud_term="adulteration",
        manual_reason="User explicitly kept olive oil.",
    )

    add(
        '"adulterated oil" seized India',
        "support_product_plus_fraud_plus_enforcement",
        product_term="oil",
        fraud_term="adulterated",
        enforcement_term="seized",
        manual_reason="Broad but high-precision incident phrase.",
    )
    add(
        '"food adulteration" "edible oil" India',
        "broad_fraud_plus_product_anchor",
        product_term="edible oil",
        fraud_term="food adulteration",
        manual_reason="Uses broad food-adulteration term only with oil anchor.",
    )
    add(
        '"oil" "food safety" "seized" India',
        "support_product_plus_enforcement",
        product_term="oil",
        enforcement_term="food safety; seized",
        source_terms=["oil", "food safety", "seized"],
        manual_reason="Support term oil anchored by enforcement and India.",
        breadth="broad",
    )
    add(
        '"ghee and oil" adulteration India',
        "mixed_oil_ghee_plus_fraud",
        product_term="ghee and oil",
        fraud_term="adulteration",
        manual_reason="Mixed oil+ghee sample pattern without pure ghee-only drift.",
    )
    add(
        '"adulterated ghee and oil" seized India',
        "mixed_oil_ghee_plus_enforcement",
        product_term="adulterated ghee and oil",
        enforcement_term="seized",
        manual_reason="Specific mixed incident phrase from sample.",
        breadth="narrow",
    )

    return rows


def count_by(rows: list[dict], key: str) -> dict:
    counts: OrderedDict[str, int] = OrderedDict()
    for row in rows:
        value = row[key]
        counts[value] = counts.get(value, 0) + 1
    return dict(counts)


def write_csv(path: Path, rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


if __name__ == "__main__":
    raise SystemExit(main())
