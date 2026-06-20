"""
Generate a formatted Excel workbook from articles.csv.

Sheets:
  1. Relevant       - articles with relevance_label == 'relevant'
  2. All Articles   - every non-duplicate article
  3. Summary        - counts by source, domain, state, oil type

Run: python scripts/make_excel.py
Output: data/outputs/articles_report.xlsx
"""

import csv
import json
import re
from collections import Counter
from pathlib import Path

import openpyxl
from openpyxl.styles import (Alignment, Border, Font, GradientFill,
                              PatternFill, Side)
from openpyxl.utils import get_column_letter

# ── Paths ─────────────────────────────────────────────────────────────────────
ROOT = Path(__file__).resolve().parent.parent
CSV_IN = ROOT / "data" / "outputs" / "articles.csv"
XLSX_OUT = ROOT / "data" / "outputs" / "articles_report.xlsx"

# ── Colour palette ─────────────────────────────────────────────────────────────
C_HEADER_BG   = "1F4E79"   # dark navy
C_HEADER_FG   = "FFFFFF"
C_RELEVANT    = "C6EFCE"   # light green
C_MAYBE       = "FFEB9C"   # light yellow
C_IRRELEVANT  = "FFC7CE"   # light red
C_ALT_ROW     = "EEF2F7"   # very light blue (alternating rows)
C_BORDER      = "BFBFBF"

# ── Columns to include in article sheets ──────────────────────────────────────
ARTICLE_COLS = [
    ("title",                "Title",              55),
    ("publication_date",     "Date",               12),
    ("source",               "Source",             22),
    ("domain",               "Domain",             28),
    ("relevance_label",      "Relevance",          12),
    ("relevance_score",      "Score",               7),
    ("food_terms_found",     "Oil Terms",          28),
    ("adulteration_terms_found", "Adulteration Terms", 28),
    ("location_terms_found", "Location Terms",     28),
    ("action_terms_found",   "Action Terms",       28),
    ("query_used",           "Query Used",         30),
    ("discovery_method",     "Source Method",      16),
    ("word_count",           "Words",               8),
    ("is_duplicate",         "Duplicate?",         10),
    ("url",                  "URL",                55),
]


def _thin_border():
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)


def _header_fill():
    return PatternFill("solid", fgColor=C_HEADER_BG)


def _rel_fill(label):
    if label == "relevant":
        return PatternFill("solid", fgColor=C_RELEVANT)
    if label == "maybe_relevant":
        return PatternFill("solid", fgColor=C_MAYBE)
    if label == "irrelevant":
        return PatternFill("solid", fgColor=C_IRRELEVANT)
    return None


def _clean_list(raw: str) -> str:
    """Turn '["edible oil", "ghee"]' into 'edible oil, ghee'."""
    try:
        items = json.loads(raw)
        return ", ".join(items)
    except Exception:
        return raw.strip('[]"\'').replace('", "', ", ").replace("', '", ", ")


def _fmt_val(col_key: str, val: str) -> str:
    if col_key in {"food_terms_found", "adulteration_terms_found",
                   "location_terms_found", "action_terms_found"}:
        return _clean_list(val)
    if col_key == "is_duplicate":
        return "Yes" if val == "1" else "No"
    if col_key == "relevance_score":
        try:
            return f"{float(val):.2f}"
        except Exception:
            return val
    return val


def _write_sheet(ws, rows, cols, title_str):
    """Write header + data rows to a worksheet."""
    # ── Header row ────────────────────────────────────────────────────────────
    h_font   = Font(name="Calibri", bold=True, color=C_HEADER_FG, size=11)
    h_fill   = _header_fill()
    h_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    h_border = _thin_border()

    for c_idx, (_, display, _width) in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=c_idx, value=display)
        cell.font   = h_font
        cell.fill   = h_fill
        cell.alignment = h_align
        cell.border = h_border

    ws.row_dimensions[1].height = 28

    # ── Data rows ─────────────────────────────────────────────────────────────
    d_font   = Font(name="Calibri", size=10)
    d_align  = Alignment(vertical="top", wrap_text=True)
    d_border = _thin_border()
    alt_fill = PatternFill("solid", fgColor=C_ALT_ROW)

    col_keys = [k for k, _, _ in cols]
    rel_idx = col_keys.index("relevance_label") if "relevance_label" in col_keys else None

    for r_idx, row in enumerate(rows, start=2):
        label = row.get("relevance_label", "")
        row_fill = _rel_fill(label) or (alt_fill if r_idx % 2 == 0 else None)

        for c_idx, col_key in enumerate(col_keys, start=1):
            raw_val = row.get(col_key, "")
            val = _fmt_val(col_key, raw_val)
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font   = d_font
            cell.alignment = d_align
            cell.border = d_border
            if row_fill:
                cell.fill = row_fill

        ws.row_dimensions[r_idx].height = 45

    # ── Column widths ─────────────────────────────────────────────────────────
    for c_idx, (_, _, width) in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(c_idx)].width = width

    # ── Freeze header + auto-filter ───────────────────────────────────────────
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ── Sheet title ───────────────────────────────────────────────────────────
    ws.sheet_properties.tabColor = "1F4E79"


def _write_summary(ws, all_rows):
    """Write a summary sheet with breakdowns."""
    relevant = [r for r in all_rows if r["relevance_label"] == "relevant"
                and r["is_duplicate"] != "1"]

    h_font  = Font(name="Calibri", bold=True, color=C_HEADER_FG, size=11)
    h_fill  = _header_fill()
    b_font  = Font(name="Calibri", bold=True, size=11)
    n_font  = Font(name="Calibri", size=10)
    border  = _thin_border()

    def section(title, data_pairs, start_row):
        ws.cell(start_row, 1, title).font = b_font
        ws.merge_cells(start_row=start_row, start_column=1,
                       end_row=start_row, end_column=3)
        ws.cell(start_row, 1).fill = PatternFill("solid", fgColor="1F4E79")
        ws.cell(start_row, 1).font = Font(name="Calibri", bold=True,
                                          color="FFFFFF", size=11)
        ws.cell(start_row, 1).alignment = Alignment(horizontal="center")
        r = start_row + 1
        for label, count in data_pairs:
            ws.cell(r, 1, label).font = n_font
            ws.cell(r, 2, count).font = n_font
            ws.cell(r, 1).border = border
            ws.cell(r, 2).border = border
            r += 1
        return r + 1

    # Overview
    cur = 1
    ws.cell(cur, 1, "Metric").font = h_font
    ws.cell(cur, 1).fill = h_fill
    ws.cell(cur, 2, "Count").font = h_font
    ws.cell(cur, 2).fill = h_fill
    cur += 1
    for label, val in [
        ("Total articles crawled", len(all_rows)),
        ("Relevant (edible oil adulteration)", sum(1 for r in all_rows if r["relevance_label"] == "relevant")),
        ("Maybe relevant", sum(1 for r in all_rows if r["relevance_label"] == "maybe_relevant")),
        ("Irrelevant", sum(1 for r in all_rows if r["relevance_label"] == "irrelevant")),
        ("Duplicates marked", sum(1 for r in all_rows if r["is_duplicate"] == "1")),
        ("Unique relevant articles", len(relevant)),
    ]:
        ws.cell(cur, 1, label).font = n_font
        ws.cell(cur, 2, val).font = n_font
        ws.cell(cur, 1).border = border
        ws.cell(cur, 2).border = border
        cur += 1
    cur += 1

    # Top domains
    domain_counts = Counter(r["domain"] for r in relevant).most_common(15)
    cur = section("Top 15 Domains (relevant articles)", domain_counts, cur)

    # By state (from location_terms_found)
    STATES = ["Uttar Pradesh", "Gujarat", "Maharashtra", "Rajasthan", "Punjab",
              "Haryana", "Delhi", "West Bengal", "Madhya Pradesh", "Karnataka",
              "Tamil Nadu", "Kerala", "Andhra Pradesh", "Telangana", "Bihar",
              "Odisha", "Jharkhand", "Assam", "Chhattisgarh", "Uttarakhand"]
    state_counts = Counter()
    for r in relevant:
        locs = _clean_list(r.get("location_terms_found", ""))
        for state in STATES:
            if state.lower() in locs.lower():
                state_counts[state] += 1
    cur = section("Articles by State (relevant)", state_counts.most_common(20), cur)

    # By oil type
    OIL_TYPES = ["mustard oil", "ghee", "edible oil", "palm oil", "groundnut oil",
                 "soybean oil", "coconut oil", "sesame oil", "sunflower oil",
                 "rice bran oil", "vanaspati", "cottonseed oil"]
    oil_counts = Counter()
    for r in relevant:
        oils = _clean_list(r.get("food_terms_found", "")).lower()
        for oil in OIL_TYPES:
            if oil in oils:
                oil_counts[oil] += 1
    cur = section("Articles by Oil Type (relevant)", oil_counts.most_common(), cur)

    # By year
    year_counts = Counter()
    for r in relevant:
        d = r.get("publication_date", "")
        if d and len(d) >= 4:
            year_counts[d[:4]] += 1
    cur = section("Articles by Year (relevant)", sorted(year_counts.items()), cur)

    # Column widths
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.freeze_panes = "A2"


def main():
    print(f"Reading {CSV_IN} ...")
    with open(CSV_IN, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        all_rows = list(reader)

    relevant_rows = [r for r in all_rows
                     if r["relevance_label"] == "relevant"
                     and r["is_duplicate"] != "1"]
    non_dup_rows  = [r for r in all_rows if r["is_duplicate"] != "1"]

    print(f"  Total: {len(all_rows)}  |  Relevant (no dups): {len(relevant_rows)}")

    wb = openpyxl.Workbook()

    # Sheet 1: Relevant articles
    ws1 = wb.active
    ws1.title = "Relevant Articles"
    _write_sheet(ws1, relevant_rows, ARTICLE_COLS, "Relevant Articles")

    # Sheet 2: All articles (no duplicates)
    ws2 = wb.create_sheet("All Articles")
    _write_sheet(ws2, non_dup_rows, ARTICLE_COLS, "All Articles")

    # Sheet 3: Summary
    ws3 = wb.create_sheet("Summary")
    _write_summary(ws3, all_rows)
    ws3.sheet_properties.tabColor = "375623"

    # Tab colours
    ws1.sheet_properties.tabColor = "375623"   # green
    ws2.sheet_properties.tabColor = "1F4E79"   # navy

    wb.save(XLSX_OUT)
    print(f"Saved: {XLSX_OUT}")


if __name__ == "__main__":
    main()
